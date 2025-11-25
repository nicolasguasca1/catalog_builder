#!/usr/bin/env python3
import argparse, os, sys, math, json, re, time, tempfile, copy
from collections import defaultdict
from datetime import datetime
from urllib.parse import urlparse, parse_qsl, urlunparse, urlencode
from urllib.parse import urlparse
from contextlib import contextmanager
from dataclasses import dataclass, field
from typing import Any, Callable, Dict, List, Optional, Set, Tuple
from pathlib import Path

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles.fills import PatternFill
from roles import roles_dict
from catalog_schema import Catalog, Release, Track, Contributor
from exercise_tracker import (
    attach_context,
    current_http_context,
    finish_and_write,
    get_current_tracker,
    record_entity_attempt,
    start_exercise,
    set_verification_summary,
    update_metadata,
)
from verification_service import extract_expected_fields, run_verification

HTTP_POST_PAYLOADS: List[Dict[str, Any]] = []
HTTP_GET_PAYLOADS: List[Dict[str, Any]] = []
HTTP_RESPONSES: List[Dict[str, Any]] = []
HTTP_GET_RESPONSES: List[Dict[str, Any]] = []
_HTTP_REQUEST_SEQ = 0


def _next_request_id() -> str:
    """Produce a per-run unique identifier to correlate requests/responses."""
    global _HTTP_REQUEST_SEQ
    _HTTP_REQUEST_SEQ += 1
    millis = int(time.time() * 1000)
    return f"req-{millis}-{_HTTP_REQUEST_SEQ:05d}"


def _snapshot_x_headers(headers: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    """Capture only custom X-* headers for logging to avoid leaking auth."""
    try:
        return {k: v for k, v in (headers or {}).items() if isinstance(k, str) and k.lower().startswith("x-")}
    except Exception:
        return {}


def _serialize_response_body(resp: requests.Response, limit: int = 2000) -> Any:
    """Return JSON-decoded body when possible, otherwise a truncated string."""
    # Prefer decoding JSON before any size-based truncation so large API responses remain inspectable.
    content_type = ""
    try:
        content_type = resp.headers.get("content-type", "") or ""
    except Exception:
        content_type = ""
    try:
        text = resp.text or ""
    except Exception:
        text = ""
    if content_type.lower().startswith("application/json"):
        try:
            return resp.json()
        except Exception:
            pass
    if not text:
        return None
    if len(text) <= limit:
        return text
    truncated = text[:limit]
    return f"{truncated}...({len(text) - limit} chars truncated)"


def _log_http_response(request_id: str, method: str, url: str, resp: requests.Response, *, context: Optional[Dict[str, Any]] = None) -> None:
    entry = {
        "timestamp": round(time.time(), 3),
        "requestId": request_id,
        "method": method.upper(),
        "url": url,
        "status": resp.status_code,
        "ok": bool(resp.ok),
        "responseHeaders": {k: v for k, v in resp.headers.items() if isinstance(k, str) and k.lower() in ("content-type", "content-length")},
        "body": _serialize_response_body(resp)
    }
    attach_context(entry, context=context)
    try:
        HTTP_RESPONSES.append(entry)
        if method.upper() == "GET":
            HTTP_GET_RESPONSES.append(copy.deepcopy(entry))
    except Exception:
        pass

# ========= Config & constants =========

ARTIFACTS = Path("artifacts"); ARTIFACTS.mkdir(exist_ok=True)
ARTIFACTS_RAW = ARTIFACTS / "raw"; ARTIFACTS_RAW.mkdir(exist_ok=True)
ARTIFACTS_REPORTS = ARTIFACTS / "reports"; ARTIFACTS_REPORTS.mkdir(exist_ok=True)
ARTIFACTS_LOGS = ARTIFACTS / "logs"; ARTIFACTS_LOGS.mkdir(exist_ok=True)
TIMEOUT = 30

_ARTIFACT_REGISTRY: Dict[str, Dict[str, str]] = {}


def _artifact_base(category: str) -> Path:
    mapping = {
        "root": ARTIFACTS,
        "raw": ARTIFACTS_RAW,
        "reports": ARTIFACTS_REPORTS,
        "logs": ARTIFACTS_LOGS,
    }
    base = mapping.get(category, ARTIFACTS_RAW)
    base.mkdir(exist_ok=True)
    return base


def _register_artifact(category: str, path: Path, description: Optional[str] = None) -> None:
    try:
        rel = path.relative_to(ARTIFACTS)
    except ValueError:
        rel = path
    _ARTIFACT_REGISTRY[str(rel)] = {
        "category": category,
        "description": description or "",
    }


def write_json_artifact(category: str, filename: str, data: Any, description: str) -> Path:
    path = _artifact_base(category) / filename
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False))
    _register_artifact(category, path, description)
    return path


def write_text_artifact(category: str, filename: str, text: str, description: str) -> Path:
    path = _artifact_base(category) / filename
    path.write_text(text)
    _register_artifact(category, path, description)
    return path

# Track property map (API expects array[int])
TRACK_PROP_MAP = {
    "NONE APPLY": 1,
    "REMIX OR DERIVATIVE": 2,
    "SAMPLES OR STOCK": 3,
    "MIX OR COMPILATION": 4,
    "ALTERNATE VERSION": 5,
    "SPECIAL GENRE": 6,
    "NON MUSICAL CONTENT": 7,
    "INCLUDES AI": 8,
}

# Hard column limits (1-based inclusive) to skip spreadsheet noise past the template definitions.
SHEET_COLUMN_LIMITS: Dict[str, int] = {
    "1) Artists list": 10,
    "2) Labels list": 1,
    "3) Release_Label": 16,
    "4) Release_Artist(s)": 7,
    "5) Release_Track": 16,
    "6) Track_Artist(s)": 7,
    # Increase to 11 to read both contributor IPI (col 4) and publisher IPI (now col 10) when template provides extra spacer columns.
    "7) Comp ContributorPublisher li": 11,
    "8) Track_Composition(s)": 11,
    "9) Audio_Properties": 12,
}

# Build a name->id role map from roles.py (case-insensitive)
ROLE_FALLBACK = { (str(v).strip().lower()): k for k, v in roles_dict.items() }

# Fallback language & genre maps (used if API lookup fails)
def all_fields_values(df: 'pd.DataFrame', req_map: Dict[str, bool], cap_per_column: int = 50, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """Enumerate sample values per column with duplicate header handling.
    Special handling: if sheet_name matches the contributors/publishers sheet ("7) Comp ContributorPublisher li"),
    collapse duplicate IPI headers to at most two logical columns (Contributor, Publisher) and provide deterministic naming.
    """
    try:
        import pandas as pd  # type: ignore
    except Exception:
        pass
    start_row = int((req_map or {}).get('_effective_data_start') or 5)
    out: Dict[str, Any] = {}
    for col in [c for c in df.columns if isinstance(c, str)]:
        obj = df[col]
        # Simple column
        if hasattr(obj, 'items') and not isinstance(obj, type(df)):  # Series
            vals = []; uniques = []; seen = set(); taken = 0
            for idx, v in obj.items():
                if is_nan(v):
                    continue
                excel_row = start_row + int(idx)
                sval = str(v)
                if taken < cap_per_column:
                    vals.append({'row': excel_row, 'value': sval}); taken += 1
                if sval not in seen:
                    seen.add(sval)
                    if len(uniques) < 10:
                        uniques.append(sval)
            out[col] = {
                'values_sample': vals,
                'values_sample_capped': len(vals) >= cap_per_column,
                'unique_values_sample': uniques
            }
            continue
        # Duplicate header region (DataFrame of same-named columns)
        sub = obj if not hasattr(obj, 'items') else df[[col]]
        subcols = list(sub.columns)
        special_ipi_mode = sheet_name == "7) Comp ContributorPublisher li" and col.strip().lower() in ("ipi/cae","ipi cae","ipi")
        if special_ipi_mode and len(subcols) > 2:
            sub = sub.iloc[:, :2]; subcols = list(sub.columns)
        composite_vals = []; composite_uniques = []; composite_seen = set(); composite_taken = 0
        per_dup_col_samples: List[Dict[str, Any]] = []
        for dup_idx in range(len(subcols)):
            dup_vals = []; dup_uniques = []; dup_seen = set(); dup_taken = 0
            for ridx, row in sub.iterrows():
                try:
                    v = row.iloc[dup_idx]
                except Exception:
                    v = None
                if is_nan(v):
                    continue
                excel_row = start_row + int(ridx)
                sval = str(v)
                if dup_taken < cap_per_column:
                    dup_vals.append({'row': excel_row, 'value': sval}); dup_taken += 1
                if sval not in dup_seen:
                    dup_seen.add(sval)
                    if len(dup_uniques) < 10:
                        dup_uniques.append(sval)
                if composite_taken < cap_per_column and sval not in composite_seen:
                    composite_vals.append({'row': excel_row, 'value': sval, 'source_duplicate_index': dup_idx}); composite_taken += 1
                if sval not in composite_seen:
                    composite_seen.add(sval)
                    if len(composite_uniques) < 10:
                        composite_uniques.append(sval)
            per_dup_col_samples.append({
                'values_sample': dup_vals,
                'values_sample_capped': len(dup_vals) >= cap_per_column,
                'unique_values_sample': dup_uniques
            })
        out[col] = {
            'values_sample': composite_vals,
            'values_sample_capped': len(composite_vals) >= cap_per_column,
            'unique_values_sample': composite_uniques,
            'duplicates': len(subcols)
        }
        if not special_ipi_mode:
            for dup_idx in range(len(subcols)):
                suffix_name = f"{col}#{dup_idx+1}"
                if suffix_name in out:
                    suffix_name = f"{suffix_name}_{dup_idx+1}"
                out[suffix_name] = per_dup_col_samples[dup_idx]
        else:
            role_names = ["IPI/CAE (Contributor)", "IPI/CAE (Publisher)"]
            for dup_idx in range(len(subcols)):
                out[role_names[dup_idx]] = per_dup_col_samples[dup_idx]
    return out
LANGUAGE_FALLBACK = {
    1:"English",2:"Hebrew",3:"French",4:"Afrikaans",5:"Arabic",6:"Bulgarian",8:"Catalan",9:"Croatian",
    10:"Czech",11:"Danish",12:"Dutch",13:"Estonian",14:"Finnish",15:"German",16:"Greek",17:"Hindi",
    18:"Hungarian",19:"Icelandic",20:"Indonesian",21:"Italian",22:"Japanese",23:"Kazakh",24:"Korean",
    25:"Lao",26:"Latvian",27:"Lithuanian",28:"Malay",29:"Norwegian",30:"Polish",31:"Portuguese",
    32:"Romanian",34:"Russian",35:"Slovak",36:"Slovenian",37:"Spanish",38:"Swedish",39:"Tagalog",
    40:"Tamil",41:"Telugu",42:"Thai",43:"Turkish",44:"Ukrainian",45:"Urdu",46:"Vietnamese",47:"Zulu",
    48:"Instrumental",49:"Chinese Simplified",50:"Chinese Traditional",52:"Cantonese",53:"Bengali",
    54:"Haitian",55:"Irish",56:"Latin",57:"Persian",58:"Punjabi",59:"Sanskrit",60:"Spanish (Latin America)",
    61:"Amharic",62:"Oromo",63:"Tigrinya",66:"Abkhazian",67:"Afar",68:"Akan",69:"Albanian",70:"Aragonese",
    71:"Armenian",72:"Assamese",73:"Avaric",74:"Avestan",75:"Aymara",76:"Azerbaijani",77:"Bambara",78:"Bashkir",
    79:"Basque",80:"Belarusian",81:"Bihari languages",82:"Bislama",83:"Bosnian",84:"Breton",85:"Burmese",
    86:"Chamorro",87:"Chechen",88:"Chichewa",89:"Chuvash",90:"Cornish",91:"Corsican",92:"Cree",93:"Divehi",
    94:"Dzongkha",95:"Esperanto",96:"Ewe",97:"Faroese",98:"Fijian",99:"Fulah",100:"Galician",101:"Georgian",
    102:"Guarani",103:"Gujarati",104:"Hausa",105:"Herero",106:"Hiri Motu",107:"Interlingua",108:"Interlingue",
    109:"Igbo",110:"Inupiaq",111:"Ido",112:"Inuktitut",113:"Javanese",114:"Kalaallisut",115:"Kannada",116:"Kanuri",
    117:"Kashmiri",118:"Central Khmer",119:"Kikuyu",120:"Kinyarwanda",121:"Kirghiz",122:"Komi",123:"Kongo",
    124:"Kurdish",125:"Kuanyama",126:"Luxembourgish",127:"Ganda",128:"Limburgan",129:"Lingala",130:"Luba-Katanga",
    131:"Manx",132:"Macedonian",133:"Malagasy",134:"Malayalam",135:"Maltese",136:"Maori",137:"Marathi",
    138:"Marshallese",139:"Mongolian",140:"Nauru",141:"Navajo",142:"North Ndebele",143:"Nepali",144:"Ndonga",
    145:"Norwegian Bokmål",146:"Norwegian Nynorsk",147:"Sichuan Yi",148:"South Ndebele",149:"Occitan",150:"Ojibwa",
    151:"Church Slavic",152:"Oromo",153:"Oriya",154:"Ossetian",155:"Pali",156:"Pashto",157:"Quechua",158:"Romansh",
    159:"Rundi",160:"Sardinian",161:"Sindhi",162:"Northern Sami",163:"Samoan",164:"Sango",165:"Serbian",
    166:"Gaelic",167:"Shona",168:"Sinhala",169:"Somali",170:"Southern Sotho",171:"Sundanese",172:"Swahili",
    173:"Swati",174:"Tajik",175:"Tibetan",176:"Turkmen",177:"Tswana",178:"Tonga",179:"Tsonga",180:"Tatar",
    181:"Twi",182:"Tahitian",183:"Uighur",184:"Uzbek",185:"Venda",186:"Volapük",187:"Walloon",188:"Welsh",
    189:"Wolof",190:"Western Frisian",191:"Xhosa",192:"Yiddish",193:"Yoruba",194:"Zhuang",195:"Bhojpuri",
    196:"Haryanvi",197:"Konkani",198:"Rajasthani",199:"Bhojpuri",200:"Haryanvi",201:"Konkani",202:"Rajasthani",
}
# inverted name->id
LANG_NAME_TO_ID_FALLBACK = {v.lower(): k for k,v in LANGUAGE_FALLBACK.items()}

MUSICSTYLE_FALLBACK = {
    10:"Pop",11:"Rock",12:"Electronic",13:"Reggae",14:"Singer/Songwriter",15:"World",16:"Dance",
    17:"Salsa y Tropical",18:"Latin",19:"New Age",20:"Holiday",21:"Arabic",22:"Jazz",23:"Children's Music",
    24:"R&B/Soul",25:"Alternative",26:"Anime",28:"Blues",29:"Brazilian",30:"Chinese",31:"Christian & Gospel",
    32:"Classical",33:"Comedy",34:"Country",35:"Folk",37:"Fitness & Workout",38:"French Pop",39:"German Folk",
    40:"German Pop",41:"Hip Hop/Rap",43:"Indian",45:"J-Pop",46:"K-Pop",47:"Karaoke",48:"Korean",49:"Opera",
    52:"Soundtrack",53:"Vocal",54:"Disney",55:"Easy Listening",56:"Inspirational",57:"Instrumental",
    58:"Marching Bands",59:"Spoken Word",60:"College Rock",61:"Goth Rock",62:"Grunge",63:"Indie Rock",
    64:"New Wave",65:"Punk",
    # (trimmed for brevity; you can paste the full list here if you prefer strict offline fallback)
}
MUSICSTYLE_NAME_TO_ID_FALLBACK = {v.lower(): k for k,v in MUSICSTYLE_FALLBACK.items()}

# ========= Helpers =========

def autofill_release_artist_roles(
    df: 'pd.DataFrame',
    req_map: Dict[str, bool],
    *,
    upc_col: str = "UPC / EAN / JAN",
    artist_col: str = "ARTIST",
    role_col: str = "ARTIST ROLE",
) -> List[Dict[str, Any]]:
    """Fill missing release-artist roles with conservative defaults.

    Preflight previously forced every row to provide an `ARTIST ROLE`, which
    blocked dry runs when vendors left the column empty. We now infer a role so
    the ingest can proceed while logging the rows we adjusted.
    """

    if df is None or df.empty:
        return []
    if role_col not in df.columns:
        return []

    autofills: List[Dict[str, Any]] = []
    primary_seen: Dict[str, bool] = {}

    # Preserve explicitly supplied primaries per release key
    for idx, row in df.iterrows():
        role = norm_str(row.get(role_col))
        if role and role.strip().lower() == "main primary artist":
            upc_key = norm_str(row.get(upc_col)) or f"row_{idx}"
            primary_seen[upc_key] = True

    start_row = int(req_map.get("_effective_data_start", 5) or 5)

    for idx, row in df.iterrows():
        role = norm_str(row.get(role_col))
        if role:
            continue
        artist = norm_str(row.get(artist_col))
        if not artist:
            # Let the usual required-field validation surface the blank artist
            continue
        upc = norm_str(row.get(upc_col))
        upc_key = upc or f"row_{idx}"
        has_primary = primary_seen.get(upc_key, False)
        fallback = "Featured Artist" if has_primary else "Main Primary Artist"
        primary_seen[upc_key] = True
        df.at[idx, role_col] = fallback
        autofills.append({
            "row": start_row + idx,
            "upc": upc,
            "artist": artist,
            "autoRole": fallback,
            "reason": "missing_role_autofilled",
        })

    return autofills

class Progress:
    """Lightweight step tracker for transparent progress & debugging.
    Usage:
        progress = Progress()
        with progress.step("Read sheets") as s:
            # ... work ...
            s.info(sheets=9)
        progress.write_log()
    """
    def __init__(self):
        self.records: List[Dict[str, Any]] = []

    @contextmanager
    def step(self, name: str):
        start = time.time()
        rec: Dict[str, Any] = {"name": name, "status": "running", "start_ts": start}
        print(f"[STEP] → {name}")
        class StepCtx:
            def __init__(self, rec: Dict[str, Any]):
                self._rec = rec
            def info(self, **kwargs):
                self._rec.setdefault("meta", {}).update(kwargs)
                if kwargs:
                    kv = ", ".join(f"{k}={v}" for k,v in kwargs.items())
                    print(f"[INFO] {name}: {kv}")
        ctx = StepCtx(rec)
        try:
            yield ctx
            rec["status"] = "ok"
        except Exception as e:
            rec["status"] = "error"
            rec["error"] = str(e)
            raise
        finally:
            rec["duration_sec"] = round(time.time() - start, 3)
            self.records.append(rec)
            print(f"[STEP] ✓ {name} → {rec['status']} in {rec['duration_sec']}s")

    def write_log(self, path: Path = ARTIFACTS / "run_log.json"):
        try:
            path.write_text(json.dumps(self.records, indent=2))
            print(f"[LOG] Wrote step log to {path.resolve()}")
        except Exception as e:
            print(f"[WARN] Failed writing run log: {e}")

def add_debug_sample(bucket: List[Any], record: Any, limit: int = 25) -> None:
    """Store up to `limit` samples per bucket for debugging."""
    if len(bucket) < limit:
        bucket.append(record)

def normalize_role_key(name: Optional[str]) -> str:
    if not name:
        return ""
    return re.sub(r"[^a-z0-9]+", " ", str(name).lower()).strip()


def normalize_isni(value: Optional[str]) -> Tuple[Optional[str], Optional[str]]:
    s = norm_str(value)
    if not s:
        return None, None
    cleaned = re.sub(r"[^0-9A-Za-z]", "", s.upper())
    if len(cleaned) != 16:
        return None, "invalid_length"
    if not re.match(r"^[0-9A-Z]{15}[0-9X]$", cleaned):
        return None, "invalid_format"
    return cleaned, None


def normalize_ipi_cae(value: Optional[str]) -> Tuple[Optional[str], Optional[str]]:
    s = norm_str(value)
    if not s:
        return None, None
    digits = re.sub(r"\D", "", s)
    if len(digits) not in (9, 11):
        return None, "invalid_length"
    return digits, None


def normalize_iswc(value: Optional[str]) -> Tuple[Optional[str], Optional[str]]:
    s = norm_str(value)
    if not s:
        return None, None
    cleaned = re.sub(r"[^0-9A-Za-z]", "", s.upper())
    if not cleaned.startswith("T") or len(cleaned) != 11:
        return None, "invalid_format"
    if not re.match(r"^T\d{10}$", cleaned):
        return None, "invalid_format"
    return cleaned, None

def getenv_required(name: str) -> str:
    v = os.getenv(name)
    if not v:
        print(f"[FATAL] Missing env var: {name}")
        sys.exit(2)
    return v

def http(session: requests.Session, method: str, url: str, token: str, json_body=None, params=None, headers=None) -> requests.Response:
    global HTTP_POST_PAYLOADS, HTTP_GET_PAYLOADS
    request_id = _next_request_id()
    method_upper = method.upper()
    h = {"Authorization": f"Bearer {token}"}
    if headers:
        h.update(headers)
    header_snapshot = _snapshot_x_headers(h)
    context_snapshot = current_http_context()
    try:
        if method_upper == "POST":
            payload_snapshot = {
                "timestamp": round(time.time(), 3),
                "requestId": request_id,
                "method": method_upper,
                "url": url,
                "json": copy.deepcopy(json_body) if json_body is not None else None,
                "params": copy.deepcopy(params) if params is not None else None,
                "headers": header_snapshot,
            }
            attach_context(payload_snapshot, context=context_snapshot)
            HTTP_POST_PAYLOADS.append(payload_snapshot)
        elif method_upper == "GET":
            payload_snapshot = {
                "timestamp": round(time.time(), 3),
                "requestId": request_id,
                "method": method_upper,
                "url": url,
                "json": None,
                "params": copy.deepcopy(params) if params is not None else None,
                "headers": header_snapshot,
            }
            attach_context(payload_snapshot, context=context_snapshot)
            HTTP_GET_PAYLOADS.append(payload_snapshot)
    except Exception:
        pass
    try:
        resp = session.request(method, url, json=json_body, params=params, headers=h, timeout=TIMEOUT)
    except Exception as exc:
        try:
            error_entry = {
                "timestamp": round(time.time(), 3),
                "requestId": request_id,
                "method": method_upper,
                "url": url,
                "error": str(exc)
            }
            attach_context(error_entry, context=context_snapshot)
            HTTP_RESPONSES.append(error_entry)
            if method_upper == "GET":
                HTTP_GET_RESPONSES.append(copy.deepcopy(error_entry))
        except Exception:
            pass
        raise
    setattr(resp, "_catalog_request_id", request_id)
    _log_http_response(request_id, method_upper, url, resp, context=context_snapshot)
    return resp

def _record_http_error(http_errors: Optional[List[Dict[str, Any]]], record: Dict[str, Any]) -> None:
    if http_errors is None:
        return
    http_errors.append(record)
    try:
        write_json_artifact("root", "http_errors.json", http_errors, "HTTP errors captured during ingest")
    except Exception:
        pass

def ensure_http_ok(resp: requests.Response, *, context: str, endpoint: str, request: Optional[Any], http_errors: Optional[List[Dict[str, Any]]]) -> None:
    if resp.ok:
        return
    record = {
        "when": context,
        "endpoint": endpoint,
        "status": resp.status_code,
        "request": request,
        "response": (resp.text or "")[:1500]
    }
    req_id = getattr(resp, "_catalog_request_id", None)
    if req_id:
        record["requestId"] = req_id
    _record_http_error(http_errors, record)
    snippet = (resp.text or "")[:300]
    raise SystemExit(f"[FATAL] {context} failed ({resp.status_code}). Response: {snippet}")

def fetch_all_labels(session: requests.Session, base_url: str, token: str, headers: Dict[str,str]) -> Dict[str, Dict[str,Any]]:
    """Return a map name.lower() -> label object.
    Supports both paginated { items, totalItemsCount } and plain array responses.
    """
    out: Dict[str, Dict[str,Any]] = {}
    page = 1; page_size = 100
    while True:
        url = f"{base_url}/content/label/all"
        resp = http(session, "GET", url, token, params={"pageNumber": page, "pageSize": page_size}, headers=headers)
        if not resp.ok:
            break
        try:
            data = resp.json()
        except Exception:
            data = None
        # Handle plain array or paginated object
        if isinstance(data, list):
            items = data
            total = len(items)
            paginated = False
        else:
            data = data or {}
            items = (data.get("items") or data.get("data") or [])
            total = data.get("totalItemsCount", len(items))
            paginated = True if "items" in data or "totalItemsCount" in data else False
        for it in items or []:
            try:
                name = (it.get("name") or "").strip()
                if name:
                    out[name.lower()] = it
            except Exception:
                continue
        # Stop if no pagination or end reached
        if not items:
            break
        if not paginated or (page * page_size) >= (total or 0):
            break
        page += 1
    # Fallback 1: if empty, retry without pagination params
    if not out:
        try:
            url = f"{base_url}/content/label/all"
            resp = http(session, "GET", url, token, headers=headers)
            if resp.ok:
                data = resp.json()
                items = data if isinstance(data, list) else (data or {}).get("items") or (data or {}).get("data") or []
                for it in items or []:
                    name = (it.get("name") or "").strip()
                    if name:
                        out[name.lower()] = it
        except Exception:
            pass
    # Fallback 2: alternative plural path
    if not out:
        try:
            url = f"{base_url}/content/labels/all"
            resp = http(session, "GET", url, token, headers=headers)
            if resp.ok:
                data = resp.json()
                items = data if isinstance(data, list) else (data or {}).get("items") or (data or {}).get("data") or []
                for it in items or []:
                    name = (it.get("name") or "").strip()
                    if name:
                        out[name.lower()] = it
        except Exception:
            pass
    return out

def fetch_all_publishers(session: requests.Session, base_url: str, token: str, headers: Optional[Dict[str,str]]) -> Dict[str, Dict[str,Any]]:
    out: Dict[str, Dict[str,Any]] = {}
    page = 1; page_size = 100
    while True:
        url = f"{base_url}/content/publisher/all"
        resp = http(session, "GET", url, token, params={"pageNumber": page, "pageSize": page_size}, headers=headers)
        if not resp.ok:
            break
        try:
            data = resp.json()
        except Exception:
            data = None
        if isinstance(data, list):
            items = data; total = len(items); paginated = False
        else:
            data = data or {}
            items = (data.get("items") or data.get("data") or [])
            total = data.get("totalItemsCount", len(items))
            paginated = True if "items" in data or "totalItemsCount" in data else False
        for it in items or []:
            try:
                name = (it.get("name") or "").strip()
                if name:
                    out[name.lower()] = it
            except Exception:
                continue
        if not items:
            break
        if not paginated or (page * page_size) >= (total or 0):
            break
        page += 1
    if not out:
        try:
            resp = http(session, "GET", f"{base_url}/content/publisher/all", token, headers=headers)
            if resp.ok:
                data = resp.json()
                items = data if isinstance(data, list) else (data or {}).get("items") or (data or {}).get("data") or []
                for it in items or []:
                    name = (it.get("name") or "").strip()
                    if name:
                        out[name.lower()] = it
        except Exception:
            pass
    if not out:
        try:
            resp = http(session, "GET", f"{base_url}/content/publishers/all", token, headers=headers)
            if resp.ok:
                data = resp.json()
                items = data if isinstance(data, list) else (data or {}).get("items") or (data or {}).get("data") or []
                for it in items or []:
                    name = (it.get("name") or "").strip()
                    if name:
                        out[name.lower()] = it
        except Exception:
            pass
    return out

def fetch_all_composers(session: requests.Session, base_url: str, token: str, headers: Optional[Dict[str,str]]) -> Dict[str, Dict[str,Any]]:
    out: Dict[str, Dict[str,Any]] = {}
    page = 1; page_size = 100
    while True:
        url = f"{base_url}/content/composer/all"
        resp = http(session, "GET", url, token, params={"pageNumber": page, "pageSize": page_size}, headers=headers)
        if not resp.ok:
            break
        try:
            data = resp.json()
        except Exception:
            data = None
        if isinstance(data, list):
            items = data; total = len(items); paginated = False
        else:
            data = data or {}
            items = (data.get("items") or data.get("data") or [])
            total = data.get("totalItemsCount", len(items))
            paginated = True if "items" in data or "totalItemsCount" in data else False
        for it in items or []:
            try:
                name = (it.get("name") or "").strip()
                if name:
                    out[name.lower()] = it
            except Exception:
                continue
        if not items:
            break
        if not paginated or (page * page_size) >= (total or 0):
            break
        page += 1
    if not out:
        try:
            resp = http(session, "GET", f"{base_url}/content/composer/all", token, headers=headers)
            if resp.ok:
                data = resp.json()
                items = data if isinstance(data, list) else (data or {}).get("items") or (data or {}).get("data") or []
                for it in items or []:
                    name = (it.get("name") or "").strip()
                    if name:
                        out[name.lower()] = it
        except Exception:
            pass
    if not out:
        try:
            resp = http(session, "GET", f"{base_url}/content/composers/all", token, headers=headers)
            if resp.ok:
                data = resp.json()
                items = data if isinstance(data, list) else (data or {}).get("items") or (data or {}).get("data") or []
                for it in items or []:
                    name = (it.get("name") or "").strip()
                    if name:
                        out[name.lower()] = it
        except Exception:
            pass
    return out

def fetch_contributor_roles(session: requests.Session, base_url: str, token: str, headers: Optional[Dict[str,str]]) -> Dict[str, Dict[str,Any]]:
    out: Dict[str, Dict[str,Any]] = {}
    try:
        resp = http(session, "GET", f"{base_url}/common/lookup/contributorRoles", token, headers=headers)
        if not resp.ok:
            return out
        data = resp.json() or []
        for it in data:
            try:
                group_id = it.get("contributorRoleGroupId")
                role_id = it.get("contributorRoleId") or it.get("roleId")
                name = (it.get("name") or "").strip()
                if group_id is not None and int(group_id) == 4 and role_id and name:
                    key = normalize_role_key(name)
                    out[key] = {"roleId": int(role_id), "name": name, "raw": it}
            except Exception:
                continue
    except Exception:
        pass
    return out

def lookup_artist_by_name(session: requests.Session, base_url: str, token: str, enterpriseId: int, name: str, headers: Dict[str,str]) -> Tuple[Optional[int], Optional[Dict[str, Any]]]:
    if not name:
        return None, None
    url = f"{base_url}/api/enterprises/{enterpriseId}/artists"
    resp = http(session, "GET", url, token, params={"name": name, "pageSize": 1}, headers=headers)
    if not resp.ok:
        return None, None
    data = resp.json() or {}
    items = data.get("items", []) or []
    if not items:
        return None, None
    for it in items:
        if (it.get("name") or "").strip().lower() == name.strip().lower():
            return it.get("artistId"), it
    return None, None

def create_or_reuse_artists(session: requests.Session, base_url: str, token: str, headers: Dict[str,str], enterpriseId: int, artists_payload: List[Dict[str,Any]], http_errors: List[Dict[str,Any]]):
    name_to_id: Dict[str,int] = {}
    name_to_external_ids: Dict[str, List[Dict[str, Any]]] = {}
    created = 0; reused = 0; failed = 0
    failed_entries: List[Dict[str, Any]] = []
    artist_detail_cache: Dict[int, Dict[str, Any]] = {}

    def cached_artist_details(artist_id: Optional[int]) -> Optional[Dict[str, Any]]:
        if artist_id is None:
            return None
        try:
            artist_id_int = int(artist_id)
        except Exception:
            return None
        if artist_id_int in artist_detail_cache:
            return artist_detail_cache[artist_id_int]
        detail = fetch_artist_details(session, base_url, token, headers, enterpriseId, artist_id_int)
        if detail:
            artist_detail_cache[artist_id_int] = detail
        return detail

    # Resolve existing artists first; always POST to /artists so dry-run/live parity is maintained.
    for it in artists_payload:
        name = (it.get("name") or "").strip()
        if not name:
            continue
        existing_id, existing_obj = lookup_artist_by_name(session, base_url, token, enterpriseId, name, headers)
        payload = copy.deepcopy(it) if isinstance(it, dict) else {}
        payload["name"] = name
        existing_ext = []
        if existing_obj:
            existing_ext = existing_obj.get("artistExternalIds") or []
        elif existing_id is not None:
            detail = cached_artist_details(existing_id)
            if detail:
                existing_ext = detail.get("artistExternalIds") or []

        sheet_ext = copy.deepcopy(it.get("artistExternalIds") or [])
        merged_ext = merge_external_ids(existing_ext, sheet_ext)
        if merged_ext:
            payload["artistExternalIds"] = merged_ext
        elif existing_ext:
            payload["artistExternalIds"] = existing_ext

        if existing_id is not None:
            try:
                payload["artistId"] = int(existing_id)
            except Exception:
                payload["artistId"] = existing_id

        endpoint = f"{base_url}/artists"
        resp = http(session, "POST", endpoint, token, json_body=payload, headers=headers)
        if not resp.ok:
            if existing_id is not None and resp.status_code in (400, 409, 412):
                record = {
                    "when": f"reuse_artist[{name}]",
                    "endpoint": endpoint,
                    "status": resp.status_code,
                    "request": payload,
                    "response": (resp.text or "")[:600]
                }
                req_id = getattr(resp, "_catalog_request_id", None)
                if req_id:
                    record["requestId"] = req_id
                _record_http_error(http_errors, record)
                detail = cached_artist_details(existing_id)
                response_data = detail or {}
                failure_rec: Dict[str, Any] = {"name": name, "status": resp.status_code}
                if req_id:
                    failure_rec["requestId"] = req_id
                failed_entries.append(failure_rec)
                failed += 1
            else:
                ensure_http_ok(resp, context=f"create_artist[{name}]", endpoint=endpoint, request=payload, http_errors=http_errors)
                response_data = {}
        else:
            try:
                response_data = resp.json() or {}
            except Exception:
                response_data = {}

        final_id = response_data.get("artistId") or existing_id
        try:
            if final_id is not None:
                final_id = int(final_id)
        except Exception:
            pass

        final_ext = response_data.get("artistExternalIds")
        if not final_ext:
            final_ext = merged_ext or existing_ext or sheet_ext

        if isinstance(it, dict):
            if final_ext:
                it["artistExternalIds"] = copy.deepcopy(final_ext)
            elif "artistExternalIds" in it:
                it.pop("artistExternalIds", None)

        if final_ext:
            name_to_external_ids[name.lower()] = copy.deepcopy(final_ext)

        if final_id is not None:
            name_to_id[name.lower()] = final_id

        success_flag = bool(resp.ok or (existing_id is not None and resp.status_code in (400, 409, 412)))
        expected_fields = extract_expected_fields("artist", payload)
        note = None
        if existing_id is not None and not resp.ok and resp.status_code in (400, 409, 412):
            note = "duplicate_reuse"
        record_entity_attempt(
            kind="artist",
            endpoint=endpoint,
            request_payload=copy.deepcopy(payload),
            response_payload=response_data if isinstance(response_data, dict) else {},
            status_code=resp.status_code,
            success=success_flag,
            identifiers={"artistId": final_id, "name": name},
            expected=expected_fields,
            request_id=getattr(resp, "_catalog_request_id", None),
            notes=note,
        )

        if existing_id is not None:
            reused += 1
        else:
            created += 1

    return name_to_id, name_to_external_ids, created, reused, failed, failed_entries

def create_or_reuse_labels(session: requests.Session, base_url: str, token: str, headers: Dict[str,str], labels_payload: List[Dict[str,Any]], http_errors: List[Dict[str,Any]]):
    existing = fetch_all_labels(session, base_url, token, headers)
    name_to_id: Dict[str,int] = {}
    created = 0; reused = 0; failed = 0
    for it in labels_payload:
        name = (it.get("name") or "").strip()
        if not name:
            continue
        key = name.lower()
        if key in existing:
            name_to_id[key] = int(existing[key].get("labelId"))
            reused += 1
            continue
        endpoint = f"{base_url}/content/label/save"
        resp = http(session, "POST", endpoint, token, json_body=it, headers=headers)
        ensure_http_ok(resp, context=f"create_label[{name}]", endpoint=endpoint, request=it, http_errors=http_errors)
        try:
            response_payload = resp.json() or {}
        except Exception:
            response_payload = {}
        expected = extract_expected_fields("label", it)
        record_entity_attempt(
            kind="label",
            endpoint=endpoint,
            request_payload=copy.deepcopy(it),
            response_payload=response_payload if isinstance(response_payload, dict) else {},
            status_code=resp.status_code,
            success=bool(resp.ok),
            identifiers={"labelId": (response_payload.get("labelId") if isinstance(response_payload, dict) else None), "name": name},
            expected=expected,
            request_id=getattr(resp, "_catalog_request_id", None),
        )
        try:
            lid = int((response_payload or {}).get("labelId"))
            name_to_id[key] = lid
            created += 1
        except Exception:
            created += 1
    return name_to_id, created, reused, failed

def create_or_reuse_publishers(session: requests.Session, base_url: str, token: str, headers: Dict[str,str], publishers_payload: List[Dict[str,Any]], http_errors: List[Dict[str,Any]], existing: Optional[Dict[str, Dict[str, Any]]] = None):
    existing_map = existing if existing is not None else fetch_all_publishers(session, base_url, token, headers)
    name_to_id: Dict[str, Optional[int]] = {}
    created = 0; reused = 0; failed = 0
    created_objs: List[Dict[str, Any]] = []
    for it in publishers_payload:
        name = (it.get("name") or "").strip() if isinstance(it, dict) else ""
        if not name:
            continue
        key = name.lower()
        existing_obj = existing_map.get(key) if existing_map else None
        if existing_obj:
            try:
                pid = existing_obj.get("publisherId") or existing_obj.get("id")
                pid = int(pid) if pid is not None else None
            except Exception:
                pid = existing_obj.get("publisherId") or existing_obj.get("id")
            name_to_id[key] = pid
            reused += 1
            continue
        endpoint = f"{base_url}/content/publisher/save"
        resp = http(session, "POST", endpoint, token, json_body=it, headers=headers)
        ensure_http_ok(resp, context=f"create_publisher[{name}]", endpoint=endpoint, request=it, http_errors=http_errors)
        try:
            data = resp.json() or {}
        except Exception:
            data = {}
        expected = extract_expected_fields("publisher", it)
        record_entity_attempt(
            kind="publisher",
            endpoint=endpoint,
            request_payload=copy.deepcopy(it),
            response_payload=data if isinstance(data, dict) else {},
            status_code=resp.status_code,
            success=bool(resp.ok),
            identifiers={"publisherId": (data.get("publisherId") if isinstance(data, dict) else None), "name": name},
            expected=expected,
            request_id=getattr(resp, "_catalog_request_id", None),
        )
        created_objs.append({"request": it, "response": data})
        pid = data.get("publisherId") or data.get("id")
        try:
            pid_int = int(pid) if pid is not None else None
        except Exception:
            pid_int = pid
        if pid_int is not None:
            name_to_id[key] = pid_int
            if existing_map is not None:
                existing_map[key] = {"name": name, "publisherId": pid_int}
        created += 1
    return name_to_id, created, reused, failed, created_objs

def create_or_reuse_composers(session: requests.Session, base_url: str, token: str, headers: Dict[str,str], composers_payload: List[Dict[str,Any]], http_errors: List[Dict[str,Any]], existing: Optional[Dict[str, Dict[str, Any]]] = None):
    existing_map = existing if existing is not None else fetch_all_composers(session, base_url, token, headers)
    name_to_id: Dict[str, Optional[int]] = {}
    created = 0; reused = 0; failed = 0
    created_objs: List[Dict[str, Any]] = []
    for it in composers_payload:
        name = (it.get("name") or "").strip() if isinstance(it, dict) else ""
        if not name:
            continue
        key = name.lower()
        existing_obj = existing_map.get(key) if existing_map else None
        if existing_obj:
            try:
                cid = existing_obj.get("composerId") or existing_obj.get("id")
                cid = int(cid) if cid is not None else None
            except Exception:
                cid = existing_obj.get("composerId") or existing_obj.get("id")
            name_to_id[key] = cid
            reused += 1
            continue
        endpoint = f"{base_url}/content/composer/save"
        resp = http(session, "POST", endpoint, token, json_body=it, headers=headers)
        ensure_http_ok(resp, context=f"create_composer[{name}]", endpoint=endpoint, request=it, http_errors=http_errors)
        try:
            data = resp.json() or {}
        except Exception:
            data = {}
        expected = extract_expected_fields("composer", it)
        record_entity_attempt(
            kind="composer",
            endpoint=endpoint,
            request_payload=copy.deepcopy(it),
            response_payload=data if isinstance(data, dict) else {},
            status_code=resp.status_code,
            success=bool(resp.ok),
            identifiers={"composerId": (data.get("composerId") if isinstance(data, dict) else None), "name": name},
            expected=expected,
            request_id=getattr(resp, "_catalog_request_id", None),
        )
        created_objs.append({"request": it, "response": data})
        cid = data.get("composerId") or data.get("id")
        try:
            cid_int = int(cid) if cid is not None else None
        except Exception:
            cid_int = cid
        if cid_int is not None:
            name_to_id[key] = cid_int
            if existing_map is not None:
                existing_map[key] = {"name": name, "composerId": cid_int}
        created += 1
    return name_to_id, created, reused, failed, created_objs


def _normalize_code(value: Optional[Any]) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    return s.upper()


def _extract_items(payload: Any) -> List[Dict[str, Any]]:
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]
    if isinstance(payload, dict):
        for key in ("items", "data", "results"):
            if key in payload:
                nested = _extract_items(payload[key])
                if nested:
                    return nested
    return []


def search_collection(session: requests.Session, base_url: str, token: str, headers: Dict[str,str], path: str, search_text: str, page_size: int = 25) -> List[Dict[str, Any]]:
    if not search_text:
        return []
    params = {"searchText": search_text, "pageNumber": 1, "pageSize": page_size}
    resp = http(session, "GET", f"{base_url}{path}", token, params=params, headers=headers)
    if not resp.ok:
        return []
    try:
        data = resp.json()
    except Exception:
        data = None
    return _extract_items(data)


def find_release_duplicate(session: requests.Session, base_url: str, token: str, headers: Dict[str,str], name: Optional[str], upc: Optional[str], cache: Dict[str, List[Dict[str, Any]]]) -> Optional[Dict[str, Any]]:
    norm_upc = _normalize_code(upc)
    if not name or not norm_upc:
        return None
    key = name.strip().lower()
    if key not in cache:
        cache[key] = search_collection(session, base_url, token, headers, "/content/release/all", name)
    for item in cache.get(key, []):
        item_upc = _normalize_code(item.get("upc") or item.get("UPC"))
        if item_upc == norm_upc:
            return item
    return None


def _track_item_has_isrc(item: Dict[str, Any], target_isrc: str) -> bool:
    norm_target = _normalize_code(target_isrc)
    if not norm_target:
        return False
    for key in ("isrc", "ISRC", "trackIsrc", "recordingIsrc"):
        val = _normalize_code(item.get(key))
        if val == norm_target:
            return True
    for container in ("trackRecordingVersions", "recordingVersions", "versions"):
        arr = item.get(container)
        if isinstance(arr, list):
            for entry in arr:
                if not isinstance(entry, dict):
                    continue
                for key in ("isrc", "ISRC"):
                    val = _normalize_code(entry.get(key))
                    if val == norm_target:
                        return True
    return False


def find_track_duplicate(session: requests.Session, base_url: str, token: str, headers: Dict[str,str], name: Optional[str], isrc: Optional[str], cache: Dict[str, List[Dict[str, Any]]]) -> Optional[Dict[str, Any]]:
    norm_isrc = _normalize_code(isrc)
    if not name or not norm_isrc:
        return None
    key = name.strip().lower()
    if key not in cache:
        cache[key] = search_collection(session, base_url, token, headers, "/content/track/all", name)
    for item in cache.get(key, []):
        if _track_item_has_isrc(item, norm_isrc):
            return item
    return None


def merge_external_ids(existing: Optional[List[Dict[str, Any]]], additional: Optional[List[Dict[str, Any]]]) -> List[Dict[str, Any]]:
    merged: Dict[Tuple[int, str], Dict[str, Any]] = {}

    def _add(items: Optional[List[Dict[str, Any]]], prefer_new: bool = False) -> None:
        if not items:
            return
        for item in items:
            if not isinstance(item, dict):
                continue
            ds = item.get("distributorStoreId")
            profile = item.get("profileId")
            try:
                ds_int = int(ds)
            except Exception:
                continue
            profile_str = str(profile).strip() if profile is not None else ""
            if not profile_str:
                continue
            key = (ds_int, profile_str.lower())
            cleaned = {k: v for k, v in item.items() if v is not None}
            cleaned["distributorStoreId"] = ds_int
            cleaned["profileId"] = profile_str
            if prefer_new or key not in merged:
                merged[key] = cleaned

    _add(existing or [], prefer_new=False)
    _add(additional or [], prefer_new=True)

    return [merged[k] for k in sorted(merged.keys(), key=lambda kv: (kv[0], kv[1]))]


def fetch_artist_details(session: requests.Session, base_url: str, token: str, headers: Dict[str,str], enterprise_id: int, artist_id: int) -> Optional[Dict[str, Any]]:
    candidate_paths = [
        f"/artists/{artist_id}",
        f"/api/enterprises/{enterprise_id}/artists/{artist_id}",
        f"/api/artists/{artist_id}",
        f"/content/artist/{artist_id}"
    ]
    for path in candidate_paths:
        url = f"{base_url}{path}"
        try:
            resp = http(session, "GET", url, token, headers=headers)
        except Exception:
            continue
        if resp.ok:
            try:
                data = resp.json()
            except Exception:
                data = None
            if isinstance(data, dict):
                return data
    return None


def yes_no(prompt: str) -> bool:
    while True:
        ans = input(f"{prompt} [y/n]: ").strip().lower()
        if ans in ("y","yes"): return True
        if ans in ("n","no"): return False

def is_nan(x): 
    return x is None or (isinstance(x, float) and math.isnan(x)) or (isinstance(x, str) and x.strip()=="")

def norm_bool(x) -> Optional[bool]:
    if x is None:
        return None
    # Numeric handling (covers 1.0/0.0 and ints)
    if isinstance(x, (int, float)):
        try:
            return bool(int(round(float(x))))
        except Exception:
            return None
    s = str(x).strip().lower()
    if s in ("1","true","yes","y","x","✓","✔","t","on"):
        return True
    if s in ("0","false","no","n","off"):
        return False
    # Attempt float parse like "1.0"/"0.0"
    try:
        return bool(int(round(float(s))))
    except Exception:
        return None

def norm_int(x) -> Optional[int]:
    if is_nan(x): return None
    try: return int(float(str(x).strip()))
    except: return None

def norm_float(x) -> Optional[float]:
    if is_nan(x): return None
    try: return float(str(x).strip())
    except: return None

def norm_str(x) -> Optional[str]:
    if is_nan(x): return None
    s = str(x)
    # normalize non-breaking spaces and collapse runs of whitespace
    s = s.replace("\u00A0", " ")
    s = re.sub(r"\s+", " ", s)
    s = s.strip()
    return s if s != "" else None

def resolve_rights_id(raw: Optional[str]) -> Tuple[int, Optional[str]]:
    """Return (rightsId, reason) where reason is None when mapping was confident.

    Known mappings:
      1 -> self-published / copyright control
      2 -> published under a publisher / administered by publisher
      3 -> public domain / no publisher

    Any unrecognized or missing value defaults to 1 with a reason flag so we can log it."""
    s = norm_str(raw)
    if not s:
        return 1, "missing"
    lowered = s.lower()
    digits = re.sub(r"[^0-9]", "", lowered)
    if digits in {"1", "2", "3"}:
        return int(digits), None

    tokens = [tok for tok in re.split(r"[^a-z0-9]+", lowered) if tok]
    normalized = " ".join(tokens)
    token_set = set(tokens)

    def has_tokens(*needed: str) -> bool:
        return all(tok in token_set for tok in needed)

    if "public domain" in normalized or has_tokens("public", "domain") or token_set.intersection({"pd"}):
        return 3, None
    if has_tokens("no", "publisher") or has_tokens("without", "publisher"):
        return 3, None

    if "copyright control" in normalized or has_tokens("copyright", "control"):
        return 1, None
    if has_tokens("self", "published") or has_tokens("self", "publish") or has_tokens("self", "publishing"):
        return 1, None
    if has_tokens("yes", "self"):
        return 1, None

    if "managed by a publisher" in normalized or has_tokens("publisher", "managed") or has_tokens("publisher", "administered"):
        return 2, None
    if has_tokens("yes", "publisher"):
        return 2, None
    if "publisher" in token_set and "no" not in token_set:
        return 2, None
    if "published" in token_set and "self" not in token_set:
        return 2, None

    return 1, "unrecognized"

def parse_header_and_requirements(xlsx_path: str, sheet_name: str) -> Tuple[List[str], Dict[str,bool], int]:
    """Return (headers, required_map, data_start_row_index) using row 3 as headers and scanning rows 1-4 for notes.
       required_map[col] = True if REQUIRED FIELD (based on text or fill heuristics). Also include:
       - required_map["_optional_flags"]: {unique_column_key -> True if explicitly marked OPTIONAL}
       - required_map["_optional_columns"]: List of column metadata dicts with display name, position, and notes
       - required_map["_notes_row_text"]: {unique_column_key -> {rowN: raw_text}}
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    header_row = 3
    # Template has information across rows 1-4; row 3 is the visible header, row 2/4 often include notes
    req_row = 4
    data_start = 5
    headers: List[str] = []
    col_index_map: Dict[int, str] = {}
    required: Dict[str, bool] = {}
    optional_flags: Dict[str, bool] = {}
    optional_columns: Dict[str, Dict[str, Any]] = {}
    notes_text: Dict[str, Dict[str, Optional[str]]] = {}
    name_occurrence: Dict[str, int] = {}

    def _row_text(cell_value) -> Optional[str]:
        if cell_value is None:
            return None
        try:
            s = str(cell_value).strip()
        except Exception:
            s = str(cell_value)
        return s or None

    requested_limit = SHEET_COLUMN_LIMITS.get(sheet_name)
    max_column = ws.max_column if requested_limit is None else min(ws.max_column, max(0, requested_limit))

    for c in range(1, max_column+1):
        row_vals: Dict[int, Optional[str]] = {}
        for rr in (1,2,3,4):
            try:
                row_vals[rr] = _row_text(ws.cell(row=rr, column=c).value)
            except Exception:
                row_vals[rr] = None
        head = row_vals.get(3) or row_vals.get(4) or row_vals.get(2) or row_vals.get(1) or f"Column_{c}"
        headers.append(head)
        col_index_map[c] = head
        occ = name_occurrence.get(head.lower(), 0)
        name_occurrence[head.lower()] = occ + 1
        col_key = f"{head}__col{c}"

        # Marker detection across rows
        is_required = False
        status_opt = False
        for rr in (1,2,3,4):
            tx = (row_vals.get(rr) or "").upper()
            if "= REQUIRED FIELD" in tx:
                is_required = True
            if "= OPTIONAL FIELD" in tx:
                status_opt = True
        if not is_required:
            try:
                for rr in (1,2,3,4):
                    cell_rr = ws.cell(row=rr, column=c)
                    fill_rr: PatternFill = cell_rr.fill
                    fg = getattr(fill_rr, "fgColor", None)
                    rgb = getattr(fg, "rgb", None) if fg else None
                    if rgb and (rgb.startswith("FFFF00") or rgb.endswith("FF00")):
                        is_required = True
                        break
            except Exception:
                pass

        prev_required = required.get(head, False)
        required[head] = bool(prev_required or is_required)
        optional_flags[col_key] = bool(status_opt)
        row_notes = {f"row{rr}": row_vals.get(rr) for rr in (1,2,3,4) if row_vals.get(rr)}
        notes_text[col_key] = row_notes
        notes_text.setdefault(head, row_notes)
        optional_columns[col_key] = {
            "key": col_key,
            "name": head,
            "column_index": c-1,
            "occurrence_index": occ,
            "optional": bool(status_opt),
            "row_texts": row_notes,
        }

    # stash meta inside req map for logging later
    required["_header_row_value"] = header_row
    required["_data_start_value"] = data_start
    required["_optional_flags"] = optional_flags
    required["_optional_columns"] = list(optional_columns.values())
    required["_notes_row_text"] = notes_text
    required["_column_limit_requested"] = requested_limit
    required["_column_limit_applied"] = max_column
    return headers, required, data_start

def df_from_sheet(xlsx_path: str, sheet_name: str) -> Tuple[pd.DataFrame, Dict[str,bool]]:
    headers, req_map, data_start = parse_header_and_requirements(xlsx_path, sheet_name)
    df = pd.read_excel(xlsx_path, sheet_name=sheet_name, header=None)
    max_cols = len(headers)
    if max_cols:
        df = df.iloc[:, :max_cols]
    # build rename map
    rename = {}
    for idx, h in enumerate(headers):
        if h:
            rename[idx] = h
    # Force data to start at row 5 per template (rows 1-4 are fixed header/subheaders)
    effective_start = 5
    df = df.iloc[effective_start-1:, :]  # pandas 1-index vs 0-index care
    df = df.rename(columns=rename)
    # Keep all columns after renaming so optional subheaders and duplicates remain available
    df = df.loc[:, [c for c in df.columns if isinstance(c, str)]]
    # drop fully empty rows
    df = df.dropna(how="all")
    # stash effective start for logging
    req_map["_effective_data_start"] = effective_start
    req_map["_column_limit_effective"] = max_cols
    return df.reset_index(drop=True), req_map

def require_columns(df: pd.DataFrame, req_map: Dict[str,bool]) -> List[Tuple[int,str]]:
    errs = []
    required_cols = [c for c,req in req_map.items() if req and c in df.columns]
    for i,row in df.iterrows():
        for col in required_cols:
            v = row.get(col, None)
            if is_nan(v):
                errs.append((i+2, f"Missing required '{col}'"))
    return errs

def parse_year_holder(year, holder) -> Optional[str]:
    """Combine year/holder into a single copyright line.

    Accept partial data so that we still emit whichever portion is present.
    """
    y_int = norm_int(year)
    y = str(y_int) if y_int is not None else norm_str(year)
    h = norm_str(holder)
    parts = [p for p in (y, h) if p]
    if not parts:
        return None
    return " ".join(parts)

def resolve_language_id(name: Optional[str], session: requests.Session, base_url: str, token: str) -> Optional[int]:
    if not name: return None
    try:
        resp = http(session, "GET", f"{base_url}/common/lookup/languages", token)
        if resp.ok:
            items = resp.json()
            for it in items:
                if it.get("name","").strip().lower() == name.strip().lower():
                    return int(it.get("languageId"))
    except Exception:
        pass
    return LANG_NAME_TO_ID_FALLBACK.get(name.strip().lower())

def resolve_musicstyle_id(name: Optional[str], session: requests.Session, base_url: str, token: str) -> Optional[int]:
    if not name: return None
    try:
        resp = http(session, "GET", f"{base_url}/common/lookup/musicstyles", token)
        if resp.ok:
            items = resp.json()
            for it in items:
                if it.get("name","").strip().lower() == name.strip().lower():
                    return int(it.get("musicStyleId"))
    except Exception:
        pass
    return MUSICSTYLE_NAME_TO_ID_FALLBACK.get(name.strip().lower())

def ingest_image_by_url(url: str, session: requests.Session, base_url: str, token: str) -> Optional[Dict[str,Any]]:
    if not url: return None
    # If Revelator has image pull-by-URL, use it; else, just store the URL in dry-run.
    # Placeholder: assuming upload by URL-like endpoint isn’t public; we keep URL in dry-run and return mock structure.
    return {"fileId": None, "filename": os.path.basename(url), "sourceUrl": url}

def ingest_audio_by_url(url: str, filetype: str, session: requests.Session, base_url: str, token: str, live: bool, headers: Dict[str,str], isrc: Optional[str]=None, upload_log: Optional[List[Dict[str,Any]]]=None) -> Optional[Dict[str,Any]]:
    """Delegate audio ingestion to the pull-external endpoint and return a simple descriptor.
    This preserves previous call sites but routes uploads via /media/audio/pullexternal/{ext}.
    """
    if not url:
        return None
    fmt = (filetype or "").strip().upper()
    fileFormat = {"WAV": 1, "FLAC": 2, "MP3": 3}.get(fmt)
    filename = os.path.basename(urlparse(url).path) or os.path.basename(url)
    result_stub = {"audioId": None, "audioFilename": filename, "fileFormat": fileFormat, "sourceUrl": url}
    if not live:
        return result_stub
    # Use pull-external to let the API fetch from the URL
    audio_id, rec = upload_audio_by_url(session, base_url, token, headers or {}, url)
    if upload_log is not None and isinstance(rec, dict):
        # annotate with isrc for traceability
        rec = {**rec, "isrc": isrc}
        upload_log.append(rec)
    if audio_id:
        return {"audioId": audio_id, "audioFilename": filename, "fileFormat": fileFormat}
    return result_stub

def extract_spotify_artist_id(val: Optional[str]) -> Optional[str]:
    """Return the canonical Spotify artist ID from a URI or URL.
    Examples:
      'spotify:artist:1Gnh4...' -> '1Gnh4...'
      'https://open.spotify.com/artist/1Gnh4...?si=...' -> '1Gnh4...'
      '1Gnh4...' -> '1Gnh4...'
    """
    if not val:
        return None
    s = str(val).strip()
    if s.startswith("spotify:"):
        parts = s.split(":")
        return parts[-1] or None
    low = s.lower()
    if "open.spotify.com/artist/" in low:
        try:
            tail = s.split("/artist/")[1]
            tail = tail.split("?")[0]
            tail = tail.split("/")[0]
            return tail or None
        except Exception:
            return None
    return s or None

def normalize_audio_url(url: Optional[str]) -> Optional[str]:
    """Normalize known share URLs to direct-download when possible (e.g., Dropbox dl=1)."""
    s = norm_str(url)
    if not s:
        return None
    try:
        low = s.lower()
        # Dropbox: ensure dl=1 for direct download, but don't alter host or other params
        if "dropbox.com/" in low:
            if "?" in s:
                base, qs = s.split("?", 1)
                # preserve existing params but force dl=1
                params = []
                seen_dl = False
                for part in qs.split("&"):
                    if part.startswith("dl="):
                        params.append("dl=1"); seen_dl = True
                    else:
                        params.append(part)
                if not seen_dl:
                    params.append("dl=1")
                s = base + "?" + "&".join(params)
            else:
                s = s + "?dl=1"
        return s
    except Exception:
        return url


def parse_duration_seconds(value: Optional[Any]) -> Optional[float]:
    """Best-effort conversion of spreadsheet duration columns to seconds."""
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    if isinstance(value, (int, float)):
        try:
            return float(value)
        except Exception:
            return None
    s = norm_str(value)
    if not s:
        return None
    # Handle mm:ss or hh:mm:ss patterns
    if ":" in s:
        parts = s.split(":")
        try:
            numbers = [float(p.strip()) for p in parts]
        except Exception:
            return None
        seconds = 0.0
        for num in numbers:
            seconds = seconds * 60 + num
        return seconds
    try:
        return float(s)
    except Exception:
        return None


def _clone_extra(extra: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    if not extra:
        return {}
    try:
        return copy.deepcopy(extra)
    except Exception:
        return dict(extra)


def _contributor_from_raw(raw: Dict[str, Any]) -> Contributor:
    name = raw.get("name") or raw.get("artistName") or ""
    role_name = raw.get("roleName")
    contributor_type = raw.get("contributorType")
    musician_type = raw.get("musicianContributorType")
    raw_line = raw.get("rawCreditLine")
    extra = _clone_extra(raw.get("extra"))
    for key in (
        "roleId",
        "artistId",
        "artistName",
        "share",
        "shareNumeric",
        "rightsId",
        "publisherName",
        "publisherId",
        "composerId",
        "iswc",
        "isni",
        "ipiCae",
    ):
        if key in raw and raw[key] is not None and key not in extra:
            extra[key] = raw[key]
    return Contributor(
        name=name,
        roleName=role_name,
        contributorType=contributor_type,
        musicianContributorType=musician_type,
        rawCreditLine=raw_line,
        extra=extra,
    )


def _track_from_meta(meta: Dict[str, Any]) -> Track:
    contributors: List[Contributor] = []
    for raw in meta.get("contributors_raw", []):
        contributors.append(_contributor_from_raw(raw))
    for raw in meta.get("composer_raw", []):
        contributors.append(_contributor_from_raw(raw))
    return Track(
        vendorTrackId=meta.get("vendorTrackId"),
        isrc=meta.get("isrc"),
        numberInRelease=meta.get("numberInRelease"),
        title=meta.get("title"),
        versionTitle=meta.get("versionTitle"),
        artistName=meta.get("artistName"),
        lengthSeconds=meta.get("lengthSeconds"),
        audioLanguageCode=meta.get("audioLanguageCode"),
        metadataLanguageCode=meta.get("metadataLanguageCode"),
        genreDesc=meta.get("genreDesc"),
        parentalAdvisory=meta.get("parentalAdvisory"),
        originalReleaseDate=meta.get("originalReleaseDate"),
        contributors=contributors,
        extra=_clone_extra(meta.get("extra")),
    )


def build_catalog_from_meta(
    release_order: List[str],
    release_meta_by_key: Dict[str, Dict[str, Any]],
    track_meta_by_isrc: Dict[str, Dict[str, Any]],
    track_isrcs_by_release_key: Dict[str, List[str]],
) -> Catalog:
    releases: List[Release] = []
    for release_key in release_order:
        meta = release_meta_by_key.get(release_key)
        if not meta:
            continue
        track_isrcs = track_isrcs_by_release_key.get(release_key, [])
        tracks: List[Track] = []
        for isrc in track_isrcs:
            tmeta = track_meta_by_isrc.get(isrc)
            if not tmeta:
                continue
            tracks.append(_track_from_meta(tmeta))
        total_tracks = meta.get("totalTracks")
        if total_tracks is None and tracks:
            total_tracks = len(tracks)
        release_contributors = [_contributor_from_raw(raw) for raw in meta.get("contributors_raw", [])]
        releases.append(
            Release(
                vendorReleaseId=meta.get("vendorReleaseId") or release_key,
                upc=meta.get("upc"),
                grid=meta.get("grid"),
                title=meta.get("title"),
                versionTitle=meta.get("versionTitle"),
                artistName=meta.get("artistName"),
                originalReleaseDate=meta.get("originalReleaseDate"),
                status=meta.get("status"),
                genreDesc=meta.get("genreDesc"),
                metadataLanguageCode=meta.get("metadataLanguageCode"),
                parentalAdvisory=meta.get("parentalAdvisory"),
                pLine=meta.get("pLine"),
                cLine=meta.get("cLine"),
                totalTracks=total_tracks,
                tracks=tracks,
                contributors=release_contributors,
                extra=_clone_extra(meta.get("extra")),
            )
        )
    return Catalog(releases=releases)


def catalog_to_api_payloads(catalog: Catalog, artist_name_to_obj: Optional[Dict[str, Dict[str, Any]]] = None) -> Tuple[List[Dict[str, Any]], List[Tuple[Optional[str], Dict[str, Any]]]]:
    """Convert catalog model to API payloads, embedding full artist objects in contributors when available."""
    releases_payload: List[Dict[str, Any]] = []
    tracks_payload: List[Tuple[Optional[str], Dict[str, Any]]] = []
    if artist_name_to_obj is None:
        artist_name_to_obj = {}

    for release in catalog.releases:
        extra = release.extra or {}
        payload: Dict[str, Any] = {
            "name": release.title,
            "version": release.versionTitle,
            "previouslyReleased": bool(extra.get("previouslyReleased")),
            "releaseDate": extra.get("releaseDate"),
        }
        if release.upc:
            payload["upc"] = release.upc
        if release.pLine:
            payload["copyrightP"] = release.pLine
        if release.cLine:
            payload["copyrightC"] = release.cLine
        if extra.get("releaseLocals"):
            payload["releaseLocals"] = copy.deepcopy(extra["releaseLocals"])
        if extra.get("languageId") is not None:
            payload["languageId"] = extra.get("languageId")
        if extra.get("primaryMusicStyleId") is not None:
            payload["primaryMusicStyleId"] = extra.get("primaryMusicStyleId")
        if extra.get("secondaryMusicStyleId") is not None:
            payload["secondaryMusicStyleId"] = extra.get("secondaryMusicStyleId")
        if extra.get("hasRecordLabel"):
            payload["hasRecordLabel"] = extra.get("hasRecordLabel")
        if extra.get("labelName"):
            payload["labelName"] = extra.get("labelName")
        if extra.get("image"):
            payload["image"] = copy.deepcopy(extra["image"])
        if extra.get("imageSourceUrl"):
            payload["imageSourceUrl"] = extra.get("imageSourceUrl")
        if extra.get("artistExternalIds"):
            payload["artistExternalIds"] = copy.deepcopy(extra["artistExternalIds"])
        if release.artistName and not payload.get("artistName"):
            payload["artistName"] = release.artistName

        if release.contributors:
            contributors_payload: List[Dict[str, Any]] = []
            for contributor in release.contributors:
                role_id = contributor.extra.get("roleId")
                if not role_id:
                    continue
                # Always use full artist object if available from artist_name_to_obj
                artist_entry: Dict[str, Any] = {}
                contributor_name_key = (contributor.name or "").strip().lower()
                full_artist = artist_name_to_obj.get(contributor_name_key)
                if full_artist:
                    # Use complete artist object (name, artistExternalIds, isni, image)
                    artist_entry = copy.deepcopy(full_artist)
                elif contributor.extra.get("artistId"):
                    artist_entry["artistId"] = contributor.extra.get("artistId")
                else:
                    artist_entry["name"] = contributor.name
                contributors_payload.append({"roleId": role_id, "artist": artist_entry})
            if contributors_payload:
                payload["contributors"] = contributors_payload

        releases_payload.append(payload)

        for track in release.tracks:
            textra = track.extra or {}
            track_payload: Dict[str, Any] = {
                "name": track.title,
                "version": track.versionTitle,
                "languageId": textra.get("languageId"),
                "explicit": textra.get("explicit"),
                "trackType": textra.get("trackType"),
                "trackNumber": track.numberInRelease,
                "previewStartSeconds": textra.get("previewStartSeconds"),
                "trackRecordingVersions": copy.deepcopy(textra.get("trackRecordingVersions", [])),
            }
            if track.artistName:
                track_payload["artistName"] = track.artistName
            if textra.get("artistExternalIds"):
                track_payload["artistExternalIds"] = copy.deepcopy(textra["artistExternalIds"])
            if textra.get("trackProperties"):
                track_payload["trackProperties"] = copy.deepcopy(textra["trackProperties"])
            else:
                track_payload["trackProperties"] = [1]
            if textra.get("compositions"):
                track_payload["compositions"] = copy.deepcopy(textra["compositions"])
            composer_contents = textra.get("composerContentsDTO")
            if composer_contents is not None:
                track_payload["composerContentsDTO"] = copy.deepcopy(composer_contents)
            else:
                track_payload["composerContentsDTO"] = []
            if textra.get("trackLocals"):
                track_payload["trackLocals"] = copy.deepcopy(textra["trackLocals"])
            else:
                if track.title and textra.get("languageId"):
                    track_payload.setdefault("trackLocals", [{"name": track.title, "languageId": textra.get("languageId")}])

            contributor_payloads: List[Dict[str, Any]] = []
            for contributor in track.contributors:
                role_id = contributor.extra.get("roleId")
                if not role_id:
                    continue
                # Always use full artist object if available from artist_name_to_obj
                artist_entry: Dict[str, Any] = {}
                contributor_name_key = (contributor.name or "").strip().lower()
                full_artist = artist_name_to_obj.get(contributor_name_key)
                if full_artist:
                    # Use complete artist object (name, artistExternalIds, isni, image)
                    artist_entry = copy.deepcopy(full_artist)
                elif contributor.extra.get("artistId"):
                    artist_entry["artistId"] = contributor.extra.get("artistId")
                else:
                    artist_entry["name"] = contributor.name
                contributor_payloads.append({"roleId": role_id, "artist": artist_entry})
            if contributor_payloads:
                track_payload["contributors"] = contributor_payloads
            else:
                track_payload.setdefault("contributors", [])

            if track_payload.get("languageId") is None:
                track_payload.pop("languageId", None)

            if not track_payload.get("trackRecordingVersions"):
                track_payload["trackRecordingVersions"] = [{
                    "isrc": track.isrc,
                    "recordingVersionType": textra.get("trackType"),
                    "audioFiles": []
                }]

            track_upc = textra.get("upc") or release.upc
            tracks_payload.append((track_upc, track_payload))

    return releases_payload, tracks_payload

def _filename_from_url(url: str) -> str:
    try:
        p = urlparse(url)
        name = os.path.basename(p.path)
        return name or "file"
    except Exception:
        return os.path.basename(url) or "file"

def _audio_ext_from_url(url: str) -> str:
    name = _filename_from_url(url).lower()
    for ext in ("flac","wav","mp3","m4a","aac","aiff","aif"):
        if name.endswith("."+ext):
            return ext
    return "wav"

def upload_audio_by_url(session: requests.Session, base_url: str, token: str, headers_common: Dict[str, str], source_url: str) -> Tuple[Optional[str], Dict[str, Any]]:
    """
    Upload audio by instructing the API to pull from an external URL.

    IMPORTANT API QUIRK (temporary, expected to be refactored):
    - Even when the source file is FLAC, the API expects the endpoint path
      'media/audio/pullexternal/wav'. In other words, both FLAC and WAV must
      be sent to the 'wav' variant of the endpoint. We still pass a correct
      fileName (with .flac or .wav) in the body so the server can record it.

    Returns (audioId, log_record). The log_record captures endpoint, request
    body (externalUrl + fileName), status, and truncated response text.
    """
    s_url = normalize_audio_url(source_url)
    ext = _audio_ext_from_url(s_url or source_url)
    # Force FLAC to use the 'wav' endpoint as per current API behavior.
    endpoint_ext = "wav" if ext in ("flac", "wav") else ext
    endpoint = f"{base_url}/media/audio/pullexternal/{endpoint_ext}"
    # Derive filename from URL; if no extension, synthesize one using ext
    raw_name = _filename_from_url(s_url or source_url)
    if not raw_name or "." not in raw_name:
        raw_name = f"audio.{ext}"
    # Ensure extension consistency: if raw_name has mismatched extension, align with ext
    base, dot, suffix = raw_name.rpartition(".")
    if base and dot and suffix.lower() != ext.lower():
        file_name = f"{base}.{ext}"
    else:
        file_name = raw_name
    # Build body with externalUrl + fileName as requested
    body = {"externalUrl": s_url, "fileName": file_name}
    try:
        resp = http(session, "POST", endpoint, token, json_body=body, headers=headers_common)
        rec = {
            "endpoint": endpoint,
            "request": body,
            "status": resp.status_code,
            "responseText": (resp.text or "")[:2000],
            "requestId": getattr(resp, "_catalog_request_id", None)
        }
        if resp.ok:
            # Accept both JSON object and JSON scalar responses
            if resp.headers.get("content-type", "").startswith("application/json"):
                try:
                    data = resp.json()
                except ValueError:
                    data = None
            else:
                data = None
            audio_id = None
            if isinstance(data, dict):
                audio_id = data.get("audioId") or data.get("fileId") or data.get("id")
            elif isinstance(data, (str, int)):
                audio_id = str(data).strip().strip('"')
            if not audio_id:
                m = re.search(r"[0-9a-fA-F\-]{8,}", resp.text or "") or re.search(r"\b\d{4,}\b", resp.text or "")
                if m:
                    audio_id = m.group(0)
            if audio_id:
                rec["audioId"] = audio_id
                return str(audio_id), rec
        return None, rec
    except Exception as e:
        return None, {"endpoint": endpoint, "request": body, "error": str(e), "requestId": getattr(locals().get("resp"), "_catalog_request_id", None)}

def download_file(session: requests.Session, url: str) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    """Download URL to a temp file. Returns (path, filename, error)."""
    try:
        s_url = normalize_audio_url(url) or url
        fn = _filename_from_url(s_url)
        with session.get(s_url, stream=True, timeout=TIMEOUT) as r:
            r.raise_for_status()
            fd, tmp_path = tempfile.mkstemp(prefix="ingest_", suffix="_"+fn)
            with os.fdopen(fd, "wb") as f:
                for chunk in r.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
        return tmp_path, fn, None
    except Exception as e:
        return None, None, str(e)

def upload_image_file(session: requests.Session, base_url: str, token: str, headers_common: Dict[str,str], file_path: str, filename: str, *, cover: bool = True) -> Tuple[Optional[str], Dict[str,Any]]:
    """POST multipart/form-data to /media/image/upload. Returns (fileId, log_record)."""
    endpoint = f"{base_url}/media/image/upload"
    params = {"cover": "true" if cover else "false"}
    headers = {"Authorization": f"Bearer {token}", **headers_common}
    request_id = _next_request_id()
    header_snapshot = _snapshot_x_headers(headers)
    context_snapshot = current_http_context()
    try:
        with open(file_path, "rb") as f:
            files = {"file": (filename, f)}
            try:
                payload_entry = {
                    "timestamp": round(time.time(), 3),
                    "requestId": request_id,
                    "method": "POST",
                    "url": endpoint,
                    "json": None,
                    "params": copy.deepcopy(params),
                    "headers": header_snapshot,
                    "files": {"file": filename}
                }
                attach_context(payload_entry, context=context_snapshot, extra={"phase": "image_upload"})
                HTTP_POST_PAYLOADS.append(payload_entry)
            except Exception:
                pass
            resp = session.post(endpoint, params=params, headers=headers, files=files, timeout=TIMEOUT)
        setattr(resp, "_catalog_request_id", request_id)
        _log_http_response(request_id, "POST", endpoint, resp, context=context_snapshot)
        rec = {
            "endpoint": endpoint,
            "params": params,
            "filename": filename,
            "status": resp.status_code,
            "contentType": resp.headers.get("content-type", ""),
            "responseText": (resp.text or "")[:2000],
            "requestId": request_id
        }
        attach_context(rec, extra={"phase": "image_upload"}, context=context_snapshot)
        if resp.ok:
            # Handle both JSON object and JSON string bodies
            if resp.headers.get("content-type","" ).startswith("application/json"):
                try:
                    data = resp.json()
                except ValueError:
                    data = None
            else:
                data = None
            file_id = None
            if isinstance(data, dict):
                file_id = data.get("fileId") or data.get("id")
            elif isinstance(data, (str, int)):
                # Some APIs return a bare JSON string with the id
                file_id = str(data).strip().strip('"')
            if not file_id:
                # Fallback: try to extract a GUID or numeric id from the text
                m = re.search(r"[0-9a-fA-F\-]{8,}", resp.text or "") or re.search(r"\b\d{4,}\b", resp.text or "")
                if m:
                    file_id = m.group(0)
            if file_id:
                rec["fileId"] = file_id
                return str(file_id), rec
        return None, rec
    except Exception as e:
        error_rec = {"endpoint": endpoint, "error": str(e), "filename": filename, "requestId": request_id}
        attach_context(error_rec, extra={"phase": "image_upload"}, context=context_snapshot)
        return None, error_rec

def map_track_properties(row: Dict[str,Any]) -> Tuple[Optional[List[int]], Dict[str, Any]]:
    """Map a Track Properties row to API IDs with diagnostics.
    Returns (ids or [1] when defaulted, diag).
    """
    # Normalize incoming row keys: collapse non-alphanumerics to spaces for resilient matching
    def norm_key(k: str) -> str:
        try:
            return re.sub(r"[^a-z0-9]+", " ", str(k or "").strip().lower()).strip()
        except Exception:
            return str(k or "").strip().lower()

    # Build a reverse map of normalized key -> (original key, value, tokens)
    def tokens(s: str) -> List[str]:
        return [t for t in norm_key(s).split() if t]

    row_norm_map: Dict[str, Tuple[str, Any, List[str]]] = {}
    for k, v in row.items():
        nk = norm_key(k)
        row_norm_map[nk] = (k, v, tokens(k))

    labels = [
        "REMIX OR DERIVATIVE","SAMPLES OR STOCK","MIX OR COMPILATION","ALTERNATE VERSION",
        "SPECIAL GENRE","NON MUSICAL CONTENT","INCLUDES AI","NONE APPLY","NONE"
    ]

    set_ids: List[int] = []
    any_true = False
    diag_values: Dict[str, Dict[str, Any]] = {}

    STOPWORDS = {"or","and","of","the"}
    def label_tokens(label: str) -> List[str]:
        return [t for t in tokens(label) if t not in STOPWORDS]

    for lab in labels:
        tgt_tokens = set(label_tokens(lab))
        matched_col = None; raw_val = None; parsed = None
        # Try exact normalized key first
        nk = norm_key(lab)
        if nk in row_norm_map:
            matched_col, raw_val, col_tokens = row_norm_map[nk]
        else:
            # Token-subset match: find a column whose tokens superset target tokens
            for _, (orig_key, val, col_tokens) in row_norm_map.items():
                col_set = set([t for t in col_tokens if t not in STOPWORDS])
                if tgt_tokens and tgt_tokens.issubset(col_set):
                    matched_col, raw_val = orig_key, val
                    break
            # Fuzzy fallback: choose the column with max token overlap if we didn't find a perfect superset
            if matched_col is None and tgt_tokens:
                best = (0, None, None)  # (overlap_count, orig_key, val)
                for _, (orig_key, val, col_tokens) in row_norm_map.items():
                    col_set = set([t for t in col_tokens if t not in STOPWORDS])
                    overlap = len(tgt_tokens & col_set)
                    if overlap > best[0]:
                        best = (overlap, orig_key, val)
                # Require at least 1 common token to consider it a match
                if best[0] >= 1:
                    matched_col, raw_val = best[1], best[2]
        parsed = norm_bool(raw_val)
        diag_values[lab] = {"matched_column": matched_col, "raw": raw_val, "bool": parsed}
        if parsed:
            any_true = True
            key = lab.upper()
            if key in ("NONE APPLY", "NONE"):
                set_ids = [1]
                break
            set_ids.append(TRACK_PROP_MAP[key if key != "NONE APPLY" else "NONE"])

    defaulted = False
    if not any_true:
        # Default to NONE per API requirement: track must have properties
        set_ids = [1]
        defaulted = True

    # dedupe & sort
    set_ids = sorted(set(set_ids))
    diag = {"values": diag_values, "defaulted_to_none": defaulted, "result_ids": set_ids}
    return set_ids, diag

# Column name normalization (case + whitespace resilient)
COLSPACE_RE = re.compile(r"[\s_]+")
def norm_colkey(s) -> str:
    """Normalize a column key to a lowercase, space-collapsed string.
    Accepts non-strings (e.g., numeric column indices) by stringifying first.
    """
    if s is None:
        return ""
    try:
        s = str(s)
    except Exception:
        # Fallback best-effort
        s = ""
    return COLSPACE_RE.sub(" ", s.strip()).lower()

# Build a map of normalized column names to real names
def make_colmap(df: pd.DataFrame) -> Dict[str,str]:
    return {norm_colkey(c): c for c in df.columns if isinstance(c, str)}

def has_col(df: pd.DataFrame, *names: str) -> bool:
    cmap = make_colmap(df)
    return any(norm_colkey(n) in cmap for n in names)

def get_val(row: pd.Series, cmap: Dict[str,str], *names: str):
    for n in names:
        col = cmap.get(norm_colkey(n))
        if col is not None:
            return row.get(col)
    return None

# Resolve a column by trying exact normalized name then token-based partial match
TOK_RE = re.compile(r"[a-z0-9]+")
def _tokens(s: str) -> List[str]:
    return TOK_RE.findall(norm_colkey(s))

def resolve_colkey(df: pd.DataFrame, *candidates: str) -> Optional[str]:
    if df is None or df.empty:
        return None
    cmap = make_colmap(df)
    # Exact first
    for n in candidates:
        col = cmap.get(norm_colkey(n))
        if col is not None:
            return col
    # Token containment fallback
    cand_tokens = [(n, set(_tokens(n))) for n in candidates]
    best: Tuple[float, Optional[str]] = (0.0, None)
    for norm_name, real in cmap.items():
        real_tokens = set(_tokens(real))
        for name, toks in cand_tokens:
            if not toks:
                continue
            # score: fraction of target tokens present in real col
            inter = len(toks & real_tokens)
            score = inter / len(toks)
            if score == 1.0:  # all tokens present
                # prefer the candidate with most tokens matched (tie by real length)
                weight = (len(toks), len(real_tokens))
                # encode into a float while preserving ordering; simpler keep best by score then token count
                if score > best[0] or (score == best[0] and len(toks) > 0):
                    best = (score, real)
    return best[1]

def first_nonempty(df: pd.DataFrame, col: Optional[str]) -> Optional[str]:
    if not col or col not in df.columns:
        return None
    for v in df[col].tolist():
        nv = norm_str(v)
        if nv is not None:
            return nv
    return None

def resolve_and_peek(df: pd.DataFrame, *candidates: str) -> Tuple[Optional[str], Optional[str]]:
    col = resolve_colkey(df, *candidates)
    return col, first_nonempty(df, col)

ACCOUNT_ID_NUMBER_RE = re.compile(r"(-?\d+(?:\.\d+)?)")

def _extract_first_int_token(text: Any) -> Optional[int]:
    if text is None:
        return None
    match = ACCOUNT_ID_NUMBER_RE.search(str(text))
    if not match:
        return None
    try:
        return int(float(match.group(1)))
    except Exception:
        return None


def _cell_value_to_int(value: Any) -> Optional[int]:
    """Convert a spreadsheet cell value to an int, tolerating floats/strings."""
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        try:
            return int(round(value))
        except Exception:
            return None
    s = str(value).strip()
    if not s:
        return None
    # Handle stray formulas like "= 332920" or "TenantId= 332920"
    s = re.sub(r"[^0-9]+", " ", s)
    s = s.strip()
    if not s:
        return None
    try:
        return int(float(s.split()[0]))
    except Exception:
        return None


def ensure_sheet_account_match(xlsx_path: str, sheet_name: str, enterprise_id: int, tenant_id: int, *, column_index_one_based: int = 10) -> Tuple[Optional[int], Optional[int]]:
    """Hard-stop the script if the Artists sheet metadata does not match the provided IDs.

    Returns the spreadsheet enterprise and tenant identifiers for downstream logging.
    """
    try:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    except Exception as exc:
        print(f"[FATAL] Could not open workbook '{xlsx_path}' for account validation: {exc}")
        sys.exit(2)
    try:
        try:
            ws = wb[sheet_name]
        except KeyError:
            print(f"[FATAL] Sheet '{sheet_name}' not found while validating account metadata.")
            sys.exit(2)

        enterprise_cell = ws.cell(row=1, column=column_index_one_based).value
        tenant_cell = ws.cell(row=2, column=column_index_one_based).value
    finally:
        wb.close()

    sheet_enterprise = _cell_value_to_int(enterprise_cell)
    sheet_tenant = _cell_value_to_int(tenant_cell)

    if sheet_enterprise is not None and sheet_enterprise != enterprise_id:
        print("[FATAL] EnterpriseId mismatch between input and spreadsheet metadata.")
        print(f"        Sheet value: {sheet_enterprise} (row 1, column {column_index_one_based})")
        print(f"        Input value: {enterprise_id}")
        sys.exit(2)

    if sheet_tenant is None:
        print("[FATAL] TenantId is missing in spreadsheet metadata (row 2, column 10).")
        sys.exit(2)

    if sheet_tenant != tenant_id:
        print("[FATAL] TenantId mismatch between input and spreadsheet metadata.")
        print(f"        Sheet value: {sheet_tenant} (row 2, column {column_index_one_based})")
        print(f"        Input value: {tenant_id}")
        sys.exit(2)

    return sheet_enterprise, sheet_tenant


def build_dry_run_payload_doc(
    base_url: str,
    enterprise_id: int,
    tenant_id: int,
    *,
    artist_image_tasks: Optional[List[Dict[str, Any]]] = None,
    artists_payload: Optional[List[Dict[str, Any]]] = None,
    labels_payload: Optional[List[Dict[str, Any]]] = None,
    publishers_payload: Optional[List[Dict[str, Any]]] = None,
    composers_payload: Optional[List[Dict[str, Any]]] = None,
    releases_payload: Optional[List[Dict[str, Any]]] = None,
    tracks_payload: Optional[List[Tuple[str, Dict[str, Any]]]] = None,
    audio_url_map: Optional[Dict[str, Optional[str]]] = None,
) -> str:
    """Return a Markdown document that outlines the HTTP payloads a live run will issue."""

    timestamp = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%SZ")
    lines: List[str] = []

    lines.append("# Dry-Run HTTP Payload Simulation")
    lines.append("")
    lines.append(f"- Generated: {timestamp}")
    lines.append(f"- Base URL: `{base_url}`")
    lines.append(f"- EnterpriseId: {enterprise_id}")
    lines.append(f"- TenantId: {tenant_id}")
    lines.append("")
    lines.append("Full payload lists are exported as JSON under `artifacts/`. This document highlights the HTTP requests that a live run will send.")
    lines.append("")
    artifact_refs = [
        "artifacts/artists.json",
        "artifacts/labels.json",
        "artifacts/publishers.json",
        "artifacts/composers.json",
        "artifacts/releases.json",
        "artifacts/tracks.json",
        "artifacts/audio_urls.json",
    ]
    lines.append("Key payload artifacts:")
    for ref in artifact_refs:
        lines.append(f"- `{ref}`")
    lines.append("")

    def _add_section(title: str, method: str, endpoint_hint: str, requests: List[Dict[str, Any]], *, sample_limit: int = 2, notes: Optional[List[str]] = None) -> None:
        lines.append(f"## {title}")
        lines.append("")
        if not requests:
            lines.append("- No requests in this stage.")
            lines.append("")
            return
        lines.append(f"- Method: `{method}`")
        lines.append(f"- Endpoint: {endpoint_hint}")
        lines.append(f"- Requests: {len(requests)}")
        if notes:
            for note in notes:
                lines.append(f"- {note}")
        sample = requests[:max(1, min(sample_limit, len(requests)))]
        lines.append("")
        lines.append("Sample payloads:")
        lines.append("```json")
        lines.append(json.dumps(sample, indent=2, ensure_ascii=False))
        lines.append("```")
        lines.append("")

    def _safe_filename(url: Optional[str], fallback_prefix: str, index: int) -> str:
        if not url:
            return f"{fallback_prefix}_{index}.dat"
        name = _filename_from_url(url)
        return name or f"{fallback_prefix}_{index}.dat"

    audio_requests: List[Dict[str, Any]] = []
    if audio_url_map:
        for isrc, url in audio_url_map.items():
            if not url:
                continue
            normalized = normalize_audio_url(url) or url
            ext = _audio_ext_from_url(normalized)
            endpoint_ext = "wav" if ext in ("flac", "wav") else ext
            raw_name = _filename_from_url(normalized)
            if not raw_name or "." not in raw_name:
                raw_name = f"audio.{ext}"
            else:
                base, dot, suffix = raw_name.rpartition(".")
                if base and dot and suffix.lower() != ext.lower():
                    raw_name = f"{base}.{ext}"
            body = {
                "externalUrl": normalized,
                "fileName": raw_name,
            }
            audio_requests.append({
                "endpoint": f"{base_url}/media/audio/pullexternal/{endpoint_ext}",
                "body": body,
                "isrc": isrc,
            })

    artist_image_requests: List[Dict[str, Any]] = []
    for idx, task in enumerate(artist_image_tasks or []):
        url = task.get("url") if isinstance(task, dict) else None
        name = task.get("name") if isinstance(task, dict) else None
        if not url:
            continue
        artist_image_requests.append({
            "endpoint": f"{base_url}/media/image/upload?cover=false",
            "artistName": name,
            "filename": _safe_filename(url, "artist_image", idx + 1),
            "sourceUrl": url,
        })

    release_image_requests: List[Dict[str, Any]] = []
    for idx, rel in enumerate(releases_payload or []):
        url = rel.get("imageSourceUrl") if isinstance(rel, dict) else None
        if not url:
            continue
        release_image_requests.append({
            "endpoint": f"{base_url}/media/image/upload?cover=true",
            "releaseName": rel.get("name"),
            "filename": _safe_filename(url, "release_image", idx + 1),
            "sourceUrl": url,
        })

    artist_requests: List[Dict[str, Any]] = []
    for payload in (artists_payload or []):
        body = copy.deepcopy(payload)
        if isinstance(body, dict):
            body.pop("imageSourceUrl", None)
        artist_requests.append({
            "endpoint": f"{base_url}/artists",
            "body": body
        })

    label_requests = [{
        "endpoint": f"{base_url}/content/label/save",
        "body": copy.deepcopy(payload)
    } for payload in (labels_payload or [])]

    publisher_requests = [{
        "endpoint": f"{base_url}/content/publisher/save",
        "body": copy.deepcopy(payload)
    } for payload in (publishers_payload or [])]

    composer_requests = [{
        "endpoint": f"{base_url}/content/composer/save",
        "body": copy.deepcopy(payload)
    } for payload in (composers_payload or [])]

    release_requests = [{
        "endpoint": f"{base_url}/content/release/save",
        "body": copy.deepcopy(rel)
    } for rel in (releases_payload or [])]

    track_requests: List[Dict[str, Any]] = []
    for upc, track_body in tracks_payload or []:
        entry = {
            "endpoint": f"{base_url}/content/track/save",
            "body": copy.deepcopy(track_body)
        }
        if upc:
            entry["releaseUPC"] = upc
        track_requests.append(entry)

    lines.append("Lookup endpoints called before mutations (GET):")
    lines.append(f"- `{base_url}/content/label/all` (with pagination fallbacks)")
    lines.append(f"- `{base_url}/content/publisher/all`")
    lines.append(f"- `{base_url}/content/composer/all`")
    lines.append(f"- `{base_url}/common/lookup/contributorRoles`")
    lines.append(f"- `{base_url}/common/lookup/languages` (on demand)")
    lines.append(f"- `{base_url}/common/lookup/musicstyles` (on demand)")
    lines.append(f"- `{base_url}/common/lookup/countries` (on demand)")
    lines.append(f"- `{base_url}/content/release/all` (live duplicate detection)")
    lines.append(f"- `{base_url}/content/track/all` (live duplicate detection)")
    lines.append("")

    _add_section(
        "Audio ingest (pull external)",
        "POST",
        "Per track → /media/audio/pullexternal/{ext}",
        audio_requests,
        notes=[
            "Each ISRC uploads the referenced audio before release/track creation.",
            "The isrc field shown below is informational for this report and is not part of the HTTP request body.",
        ],
    )

    _add_section(
        "Artist profile image uploads",
        "POST",
        "Per artist → /media/image/upload?cover=false",
        artist_image_requests,
        notes=["Images are uploaded prior to calling /artists. Live runs send multipart form-data with the listed filename."],
    )

    _add_section(
        "Upsert artists",
        "POST",
        f"{base_url}/artists",
        artist_requests,
        notes=["During live execution, the placeholder image.sourceUrl shown here is replaced once the upload stage returns image.fileId values."]
    )

    _add_section(
        "Save labels",
        "POST",
        f"{base_url}/content/label/save",
        label_requests,
        notes=["Existing labels are detected via GET /content/label/all; only unknown names trigger POST requests."]
    )

    _add_section(
        "Save publishers",
        "POST",
        f"{base_url}/content/publisher/save",
        publisher_requests,
    )

    _add_section(
        "Save composers",
        "POST",
        f"{base_url}/content/composer/save",
        composer_requests,
    )

    _add_section(
        "Release cover image uploads",
        "POST",
        "Per release → /media/image/upload?cover=true",
        release_image_requests,
        notes=["Successful uploads replace imageSourceUrl with image.fileId inside the release payload prior to /content/release/save."],
    )

    _add_section(
        "Create releases",
        "POST",
        f"{base_url}/content/release/save",
        release_requests,
        notes=["On duplicate UPC responses, the script retries without the UPC value (see live logs)."],
    )

    _add_section(
        "Create tracks",
        "POST",
        f"{base_url}/content/track/save",
        track_requests,
    )

    doc = "\n".join(lines).strip()
    return doc + "\n"

# ========= Main pipeline =========

def main():
    parser = argparse.ArgumentParser(description="Catalog spreadsheet parser (dry-run first).")
    parser.add_argument("xlsx", help="Path to the XLSX")
    parser.add_argument("--base-url", default=os.getenv("REVELATOR_BASE_URL", "https://api.revelator.com"))
    parser.add_argument("--token", default=os.getenv("REVELATOR_TOKEN", ""))
    parser.add_argument("--live", action="store_true", help="Execute HTTP calls (otherwise dry-run)")
    parser.add_argument("--role-map", default="roles.json", help="JSON file with RoleName->roleId mapping (optional).")
    args = parser.parse_args()
    start_exercise(Path(args.xlsx), mode="live" if args.live else "dry-run")

    token = args.token or getenv_required("REVELATOR_TOKEN")
    base_url = args.base_url.rstrip("/")
    update_metadata(baseUrl=base_url, workbook=args.xlsx, live=bool(args.live))

    try:
        _run_ingest_pipeline(args=args, token=token, base_url=base_url)
    finally:
        finish_and_write(ARTIFACTS)


def _run_ingest_pipeline(*, args: argparse.Namespace, token: str, base_url: str) -> None:
    # ---- Enterprise/Tenant prompt & validation
    print("Before we start, please provide target account identifiers.")
    ent = input("EnterpriseId: ").strip()
    ten = input("TenantId: ").strip()
    if not ent.isdigit() or not ten.isdigit():
        print("[FATAL] EnterpriseId and TenantId must be integers.")
        sys.exit(2)
    enterpriseId = int(ent); tenantId = int(ten)
    update_metadata(enterpriseId=enterpriseId, tenantId=tenantId)

    # Simple prerequisite: spreadsheet metadata must match provided IDs before proceeding.
    sheet_enterprise_id, sheet_tenant_id = ensure_sheet_account_match(
        args.xlsx,
        "1) Artists list",
        enterpriseId,
        tenantId,
        column_index_one_based=10,
    )

    with requests.Session() as session:
        progress = Progress()
        global HTTP_POST_PAYLOADS, HTTP_GET_PAYLOADS, HTTP_RESPONSES, HTTP_GET_RESPONSES, _HTTP_REQUEST_SEQ
        HTTP_POST_PAYLOADS = []
        HTTP_GET_PAYLOADS = []
        HTTP_RESPONSES = []
        HTTP_GET_RESPONSES = []
        _HTTP_REQUEST_SEQ = 0
        debug_trace: Dict[str, List[Dict[str, Any]]] = {
            "artists_raw": [],
            "publishers_raw": [],
            "releases_raw": [],
            "tracks_raw": [],
            "track_compositions": [],
            "composer_entries": [],
            "composer_warnings": [],
            "final_tracks": [],
            "final_releases": [],
            "release_copyright_debug": [],
            "sheet_account_metadata": [],
            "identifier_warnings": [],
            "artist_binding_warnings": []
        }
        lookup_headers = {"X-EnterpriseId": str(enterpriseId), "X-TenantId": str(tenantId)}
        composer_role_map: Dict[str, Dict[str, Any]] = {}
        publisher_lookup: Dict[str, Dict[str, Any]] = {}

        # Validate enterprise
        with progress.step("Validate enterprise") as s:
            r = http(session, "GET", f"{base_url}/enterprise/clients/{enterpriseId}", token)
            if not r.ok:
                print(f"[FATAL] Enterprise check failed ({r.status_code}): {r.text[:500]}")
                sys.exit(2)
            ent_info = r.json()
            name = ent_info.get("name","?")
            s.info(enterpriseId=enterpriseId, tenantId=tenantId, enterpriseName=name)
            print(f"Resolved EnterpriseId={enterpriseId} → name='{name}'")
            if not yes_no("Proceed ingesting catalog for this enterprise?"):
                print("Aborting as requested.")
                sys.exit(0)
            update_metadata(enterpriseName=name)

        # Load role map
        with progress.step("Load role map") as s:
            # ROLE_FALLBACK is name->id from roles.py
            role_map = ROLE_FALLBACK.copy()
            if Path(args.role_map).exists():
                try:
                    file_map = json.loads(Path(args.role_map).read_text())
                    if isinstance(file_map, dict):
                        # Accept both name->id and id->name; normalize to name->id
                        normalized: Dict[str,int] = {}
                        for k,v in file_map.items():
                            if isinstance(k, str) and isinstance(v, int):
                                nm = (norm_str(k) or "").lower()
                                if nm: normalized[nm] = v
                            elif isinstance(k, (int, float)) and isinstance(v, str):
                                nm = (norm_str(v) or "").lower()
                                if nm: normalized[nm] = int(k)
                        role_map.update(normalized)
                except Exception:
                    print("[WARN] Could not parse roles.json, using fallback map only.")
            s.info(roles=len(role_map))

        with progress.step("Fetch composer roles & publishers") as s:
            try:
                composer_role_map = fetch_contributor_roles(session, base_url, token, headers=lookup_headers)
            except Exception as exc:
                print(f"[WARN] Could not fetch contributor roles: {exc}")
                composer_role_map = {}
            try:
                publisher_lookup = fetch_all_publishers(session, base_url, token, headers=lookup_headers)
            except Exception as exc:
                print(f"[WARN] Could not fetch publishers: {exc}")
                publisher_lookup = {}
            try:
                composer_lookup = fetch_all_composers(session, base_url, token, headers=lookup_headers)
            except Exception as exc:
                print(f"[WARN] Could not fetch composers: {exc}")
                composer_lookup = {}
            # Countries lookup (single call) to support countryId and countryOfResidenceId mapping
            countries_lookup: Dict[str, Dict[str, Any]] = {}
            try:
                resp = http(session, "GET", f"{base_url}/common/lookup/countries", token, headers=lookup_headers)
                if resp.ok:
                    for it in resp.json() or []:
                        nm = str(it.get("name") or "").strip().lower()
                        if nm:
                            countries_lookup[nm] = it
            except Exception as exc:
                print(f"[WARN] Could not fetch countries lookup: {exc}")
            s.info(contributor_roles=len(composer_role_map), publishers=len(publisher_lookup), composers=len(composer_lookup))

        # ===== Read sheets
        xlsx_path = args.xlsx

        s1 = "1) Artists list"
        s2 = "2) Labels list"
        s3 = "3) Release_Label"
        s4 = "4) Release_Artist(s)"
        s5 = "5) Release_Track"
        s6 = "6) Track_Artist(s)"
        s7 = "7) Comp ContributorPublisher li"
        s8 = "8) Track_Composition(s)"
        s9 = "9) Audio_Properties"

        release_role_autofill: List[Dict[str, Any]] = []
        with progress.step("Read sheets") as s:
            try:
                df_art, req_art = df_from_sheet(xlsx_path, s1)
                df_lab, req_lab = df_from_sheet(xlsx_path, s2)
                df_rel, req_rel = df_from_sheet(xlsx_path, s3)
                df_relart, req_relart = df_from_sheet(xlsx_path, s4)
                df_reltrk, req_reltrk = df_from_sheet(xlsx_path, s5)
                df_trkart, req_trkart = df_from_sheet(xlsx_path, s6)
                df_comp_masters, req_comp_masters = df_from_sheet(xlsx_path, s7)
                df_trkcomp, req_trkcomp = df_from_sheet(xlsx_path, s8)
                df_props, req_props = df_from_sheet(xlsx_path, s9)
            except KeyError as e:
                print(f"[FATAL] Sheet not found: {e}")
                raise
            # capture header/data start from req maps
            def meta(req):
                return {
                    "header_row": req.get("_header_row_value"),
                    "data_start": req.get("_data_start_value"),
                    "effective_start": 5,
                }
            # Column peeks: show the first data point found under key columns
            art_name_col, art_name_first = resolve_and_peek(df_art, "Artist Name", "ARTIST NAME", "Artist")
            lab_name_col, lab_name_first = resolve_and_peek(df_lab, "Label Name", "LABEL NAME", "Label")
            upc_rel_col, upc_rel_first = resolve_and_peek(df_rel, "UPC / EAN / JAN")
            upc_reltrk_col, upc_reltrk_first = resolve_and_peek(df_reltrk, "UPC / EAN / JAN")
            isrc_reltrk_col, isrc_reltrk_first = resolve_and_peek(df_reltrk, "ISRC/vISRC")
            isrc_trkart_col, isrc_trkart_first = resolve_and_peek(df_trkart, "ISRC/vISRC")
            artist_trkart_col, artist_trkart_first = resolve_and_peek(df_trkart, "ARTIST")
            role_trkart_col, role_trkart_first = resolve_and_peek(df_trkart, "ARTIST ROLE")
            isrc_trkcomp_col, isrc_trkcomp_first = resolve_and_peek(df_trkcomp, "ISRC/vISRC")
            comp_trkcomp_col, comp_trkcomp_first = resolve_and_peek(df_trkcomp, "COMPOSITION CONTRIBUTOR")
            share_trkcomp_col, share_trkcomp_first = resolve_and_peek(df_trkcomp, "SHARE%")
            isrc_props_col, isrc_props_first = resolve_and_peek(df_props, "ISRC/vISRC")
            s.info(
                artists=len(df_art), labels=len(df_lab), releases=len(df_rel), rel_artists=len(df_relart), rel_tracks=len(df_reltrk), track_artists=len(df_trkart), comps=len(df_trkcomp), props=len(df_props),
                artists_meta=meta(req_art), labels_meta=meta(req_lab), releases_meta=meta(req_rel),
                peek={
                    "artists": {"col": art_name_col, "first": art_name_first},
                    "labels": {"col": lab_name_col, "first": lab_name_first},
                    "releases": {"upc_col": upc_rel_col, "first_upc": upc_rel_first},
                    "rel_tracks": {"upc_col": upc_reltrk_col, "first_upc": upc_reltrk_first, "isrc_col": isrc_reltrk_col, "first_isrc": isrc_reltrk_first},
                    "track_artists": {"isrc_col": isrc_trkart_col, "first_isrc": isrc_trkart_first, "artist_col": artist_trkart_col, "first_artist": artist_trkart_first, "role_col": role_trkart_col, "first_role": role_trkart_first},
                    "track_comps": {"isrc_col": isrc_trkcomp_col, "first_isrc": isrc_trkcomp_first, "comp_col": comp_trkcomp_col, "first_comp": comp_trkcomp_first, "share_col": share_trkcomp_col, "first_share": share_trkcomp_first},
                    "props": {"isrc_col": isrc_props_col, "first_isrc": isrc_props_first}
                }
            )

            # Dump headers snapshot for debugging
            headers_snapshot = {
                s1: {"raw": list(df_art.columns), "norm": [norm_colkey(c) for c in df_art.columns]},
                s2: {"raw": list(df_lab.columns), "norm": [norm_colkey(c) for c in df_lab.columns]},
                s3: {"raw": list(df_rel.columns), "norm": [norm_colkey(c) for c in df_rel.columns]},
                s4: {"raw": list(df_relart.columns), "norm": [norm_colkey(c) for c in df_relart.columns]},
                s5: {"raw": list(df_reltrk.columns), "norm": [norm_colkey(c) for c in df_reltrk.columns]},
                s6: {"raw": list(df_trkart.columns), "norm": [norm_colkey(c) for c in df_trkart.columns]},
                s7: {"raw": list(df_comp_masters.columns), "norm": [norm_colkey(c) for c in df_comp_masters.columns]},
                s8: {"raw": list(df_trkcomp.columns), "norm": [norm_colkey(c) for c in df_trkcomp.columns]},
                s9: {"raw": list(df_props.columns), "norm": [norm_colkey(c) for c in df_props.columns]},
            }
            write_json_artifact("reports", "headers.json", headers_snapshot, "Workbook header snapshot for traceability")

        sheet_account_meta = {
            "enterpriseId": sheet_enterprise_id,
            "tenantId": sheet_tenant_id,
            "column_key": None,
            "column_name": None,
            "rows": {"row1": sheet_enterprise_id, "row2": sheet_tenant_id},
            "column_index_one_based": 10,
            "fallback_error": None,
            "enterprise_source": {"row": 1, "column": 10, "value": sheet_enterprise_id},
            "tenant_source": {"row": 2, "column": 10, "value": sheet_tenant_id},
        }
        sheet_metadata_entry = {
            **sheet_account_meta,
            "input_enterpriseId": enterpriseId,
            "input_tenantId": tenantId,
            "sheet": s1,
        }
        debug_trace["sheet_account_metadata"].append(sheet_metadata_entry)

        with progress.step("Validate sheet account metadata") as s:
            sheet_ent = sheet_account_meta.get("enterpriseId")
            sheet_ten = sheet_account_meta.get("tenantId")
            s.info(
                sheet_enterpriseId=sheet_ent,
                sheet_tenantId=sheet_ten,
                input_enterpriseId=enterpriseId,
                input_tenantId=tenantId,
            )

            issues: List[str] = []
            if sheet_ent is not None and sheet_ent != enterpriseId:
                issues.append(f"EnterpriseId mismatch: sheet={sheet_ent} input={enterpriseId}")
            if sheet_ten is not None and sheet_ten != tenantId:
                issues.append(f"TenantId mismatch: sheet={sheet_ten} input={tenantId}")

            if issues:
                sheet_metadata_entry["status"] = "mismatch"
                print("[FATAL] Spreadsheet account metadata mismatch detected:")
                for msg in issues:
                    print(f"  - {msg}")
                print("Please update the Artists sheet metadata rows 1-2 (column 10) or rerun with matching identifiers.")
                sys.exit(2)

            missing: List[str] = []
            if sheet_ent is None:
                missing.append("EnterpriseId (row1, column10)")
            if sheet_ten is None:
                missing.append("TenantId (row2, column10)")

            if missing:
                sheet_metadata_entry["status"] = "missing"
                print("[WARN] Sheet metadata is missing " + ", ".join(missing) + ". Continuing without spreadsheet validation.")
            else:
                sheet_metadata_entry["status"] = "matched"
        # ===== Optional fields audit (visibility in dry-run and live)
        with progress.step("Audit optional fields usage") as s:
            def extract_optional_series(df: pd.DataFrame, entry: Dict[str, Any]) -> Optional[pd.Series]:
                pos = entry.get("column_index")
                ser: Optional[pd.Series] = None
                if isinstance(pos, int) and 0 <= pos < df.shape[1]:
                    candidate = df.iloc[:, pos]
                    if isinstance(candidate, pd.Series):
                        ser = candidate
                    elif isinstance(candidate, pd.DataFrame) and not candidate.empty:
                        occ = entry.get("occurrence_index", 0)
                        ser = candidate.iloc[:, occ] if occ < candidate.shape[1] else candidate.iloc[:, 0]
                if ser is None:
                    label = entry.get("name")
                    if label in df.columns:
                        obj = df[label]
                        if isinstance(obj, pd.DataFrame):
                            occ = entry.get("occurrence_index", 0)
                            ser = obj.iloc[:, occ] if occ < obj.shape[1] else obj.iloc[:, 0]
                        else:
                            ser = obj
                return ser

            def optional_audit(df: pd.DataFrame, req_map: Dict[str, bool]) -> Dict[str, Any]:
                entries = (req_map or {}).get("_optional_columns") or []
                covered = []
                unused = []
                missing_cols = []
                for entry in entries:
                    if not entry.get("optional"):
                        continue
                    series = extract_optional_series(df, entry)
                    if series is None:
                        missing_cols.append(entry.get("name"))
                        continue
                    series = series.astype(object)
                    count_nonempty = int(series.map(lambda v: 0 if is_nan(v) else 1).sum()) if len(series) else 0
                    total = int(len(series))
                    pct = (count_nonempty / total) if total else 0.0
                    report_entry = {
                        "column": entry.get("name"),
                        "column_key": entry.get("key"),
                        "notes": entry.get("row_texts"),
                        "rows_with_values": count_nonempty,
                        "rows_total": total,
                        "fill_rate": round(pct, 4)
                    }
                    if count_nonempty > 0:
                        covered.append(report_entry)
                    else:
                        unused.append(report_entry)
                covered = sorted(covered, key=lambda e: (-e["rows_with_values"], e["column"]))
                unused = sorted(unused, key=lambda e: e["column"])
                return {"covered": covered, "unused": unused, "missing": missing_cols}

            def optional_values_extract(df: pd.DataFrame, req_map: Dict[str, bool], cap_per_column: int = 50) -> Dict[str, Any]:
                entries = (req_map or {}).get("_optional_columns") or []
                start_row = int((req_map or {}).get("_effective_data_start") or 5)
                out: Dict[str, Any] = {}
                for entry in entries:
                    if not entry.get("optional"):
                        continue
                    series = extract_optional_series(df, entry)
                    if series is None:
                        continue
                    series = series.astype(object)
                    vals = []
                    uniques = []
                    seen = set()
                    taken = 0
                    for idx, v in series.items():
                        if is_nan(v):
                            continue
                        excel_row = start_row + int(idx)
                        sval = str(v)
                        if taken < cap_per_column:
                            vals.append({"row": excel_row, "value": sval})
                            taken += 1
                        if sval not in seen:
                            seen.add(sval)
                            if len(uniques) < 10:
                                uniques.append(sval)
                    out[entry.get("key") or entry.get("name")] = {
                        "column": entry.get("name"),
                        "values_sample": vals,
                        "values_sample_capped": len(vals) >= cap_per_column,
                        "unique_values_sample": uniques
                    }
                return out

            optional_report = {
                s1: optional_audit(df_art, req_art),
                s2: optional_audit(df_lab, req_lab),
                s3: optional_audit(df_rel, req_rel),
                s4: optional_audit(df_relart, req_relart),
                s5: optional_audit(df_reltrk, req_reltrk),
                s6: optional_audit(df_trkart, req_trkart),
                s7: optional_audit(df_comp_masters, req_comp_masters),
                s8: optional_audit(df_trkcomp, req_trkcomp),
                s9: optional_audit(df_props, req_props),
            }
            # Write full JSON artifact
            write_json_artifact("reports", "optional_fields_report.json", optional_report, "Optional field coverage by sheet")
            # Also write actual non-empty values samples per optional column
            optional_values = {
                s1: optional_values_extract(df_art, req_art),
                s2: optional_values_extract(df_lab, req_lab),
                s3: optional_values_extract(df_rel, req_rel),
                s4: optional_values_extract(df_relart, req_relart),
                s5: optional_values_extract(df_reltrk, req_reltrk),
                s6: optional_values_extract(df_trkart, req_trkart),
                s7: optional_values_extract(df_comp_masters, req_comp_masters),
                s8: optional_values_extract(df_trkcomp, req_trkcomp),
                s9: optional_values_extract(df_props, req_props),
            }
            write_json_artifact("reports", "optional_fields_values.json", optional_values, "Optional field value samples")
            # Small console summary to keep it skimmable
            summary = {name: {
                "optional_cols": len(rep.get("covered", [])) + len(rep.get("unused", [])),
                "used": len(rep.get("covered", [])),
                "unused": len(rep.get("unused", []))
            } for name, rep in optional_report.items()}
            # Also compute top 5 unused optional columns across all sheets
            unused_flat = []
            for sh, rep in optional_report.items():
                for e in rep.get("unused", [])[:]:
                    unused_flat.append({"sheet": sh, **e})
            # Only show a small sample in s.info
            sample_unused = [{"sheet": u.get("sheet"), "column": u.get("column")} for u in unused_flat[:5]]
            s.info(summary=summary, sample_unused=sample_unused, values_artifact="optional_fields_values.json")

        # ===== All fields audit (every header, not just optional)
        with progress.step("Audit all columns values") as s:
            def all_fields_report(df: pd.DataFrame, req_map: Dict[str, bool]) -> Dict[str, Any]:
                notes = (req_map or {}).get("_notes_row_text") or {}
                cols_sum = []
                for col in [c for c in df.columns if isinstance(c, str)]:
                    obj = df[col]
                    if isinstance(obj, pd.DataFrame):
                        # Duplicate column name: count row as non-empty if any of the dups has a value
                        if len(obj) == 0:
                            count_nonempty = 0; total = 0
                        else:
                            mask = obj.apply(lambda row: any(not is_nan(v) for v in row), axis=1)
                            count_nonempty = int(mask.sum()); total = int(len(obj))
                    else:
                        series = obj
                        count_nonempty = int(series.map(lambda v: 0 if is_nan(v) else 1).sum()) if len(series) else 0
                        total = int(len(series))
                    pct = (count_nonempty / total) if total else 0.0
                    cols_sum.append({
                        "column": col,
                        "notes": notes.get(col),
                        "rows_with_values": count_nonempty,
                        "rows_total": total,
                        "fill_rate": round(pct, 4)
                    })
                # sort descending by rows_with_values then by name
                cols_sum = sorted(cols_sum, key=lambda e: (-e["rows_with_values"], e["column"]))
                return {"columns": cols_sum}

                def all_fields_values(df: pd.DataFrame, req_map: Dict[str, bool], cap_per_column: int = 50, sheet_name: Optional[str] = None) -> Dict[str, Any]:
                    start_row = int((req_map or {}).get("_effective_data_start") or 5)
                    out: Dict[str, Any] = {}
                    for col in [c for c in df.columns if isinstance(c, str)]:
                        obj = df[col]
                        # Simple series case
                        if hasattr(obj, "items") and not isinstance(obj, pd.DataFrame):
                            vals = []; uniques = []; seen = set(); taken = 0
                            for idx, v in obj.items():
                                if is_nan(v):
                                    continue
                                excel_row = start_row + int(idx)
                                sval = str(v)
                                if taken < cap_per_column:
                                    vals.append({"row": excel_row, "value": sval})
                                    taken += 1
                                if sval not in seen:
                                    seen.add(sval)
                                    if len(uniques) < 10:
                                        uniques.append(sval)
                            out[col] = {
                                "values_sample": vals,
                                "values_sample_capped": len(vals) >= cap_per_column,
                                "unique_values_sample": uniques
                            }
                            continue
                        # Duplicate header region
                        sub = obj if isinstance(obj, pd.DataFrame) else df[[col]]
                        subcols = list(sub.columns)
                        special_ipi_mode = sheet_name == s7 and col.strip().lower() in ("ipi/cae","ipi cae","ipi")
                        if special_ipi_mode and len(subcols) > 2:
                            sub = sub.iloc[:, :2]; subcols = list(sub.columns)
                        composite_vals = []; composite_uniques = []; composite_seen = set(); composite_taken = 0
                        per_dup_col_samples: List[Dict[str, Any]] = []
                        for dup_idx in range(len(subcols)):
                            dup_vals = []; dup_uniques = []; dup_seen = set(); dup_taken = 0
                            for ridx, row in sub.iterrows():
                                try:
                                    v = row.iloc[dup_idx]
                                except Exception:
                                    v = None
                                if is_nan(v):
                                    continue
                                excel_row = start_row + int(ridx)
                                sval = str(v)
                                if dup_taken < cap_per_column:
                                    dup_vals.append({"row": excel_row, "value": sval})
                                    dup_taken += 1
                                if sval not in dup_seen:
                                    dup_seen.add(sval)
                                    if len(dup_uniques) < 10:
                                        dup_uniques.append(sval)
                                if composite_taken < cap_per_column and sval not in composite_seen:
                                    composite_vals.append({"row": excel_row, "value": sval, "source_duplicate_index": dup_idx})
                                    composite_taken += 1
                                if sval not in composite_seen:
                                    composite_seen.add(sval)
                                    if len(composite_uniques) < 10:
                                        composite_uniques.append(sval)
                            per_dup_col_samples.append({
                                "values_sample": dup_vals,
                                "values_sample_capped": len(dup_vals) >= cap_per_column,
                                "unique_values_sample": dup_uniques
                            })
                        out[col] = {
                            "values_sample": composite_vals,
                            "values_sample_capped": len(composite_vals) >= cap_per_column,
                            "unique_values_sample": composite_uniques,
                            "duplicates": len(subcols)
                        }
                        if not special_ipi_mode:
                            for dup_idx in range(len(subcols)):
                                suffix_name = f"{col}#{dup_idx+1}"
                                if suffix_name in out:
                                    suffix_name = f"{suffix_name}_{dup_idx+1}"
                                out[suffix_name] = per_dup_col_samples[dup_idx]
                        else:
                            role_names = ["IPI/CAE (Contributor)", "IPI/CAE (Publisher)"]
                            for dup_idx in range(len(subcols)):
                                out[role_names[dup_idx]] = per_dup_col_samples[dup_idx]
                    return out

            all_report = {
                s1: all_fields_report(df_art, req_art),
                s2: all_fields_report(df_lab, req_lab),
                s3: all_fields_report(df_rel, req_rel),
                s4: all_fields_report(df_relart, req_relart),
                s5: all_fields_report(df_reltrk, req_reltrk),
                s6: all_fields_report(df_trkart, req_trkart),
                s7: all_fields_report(df_comp_masters, req_comp_masters),
                s8: all_fields_report(df_trkcomp, req_trkcomp),
                s9: all_fields_report(df_props, req_props),
            }
            write_json_artifact("reports", "all_fields_report.json", all_report, "Spreadsheet field completeness summary")

            all_values = {
                s1: all_fields_values(df_art, req_art, sheet_name=s1),
                s2: all_fields_values(df_lab, req_lab, sheet_name=s2),
                s3: all_fields_values(df_rel, req_rel, sheet_name=s3),
                s4: all_fields_values(df_relart, req_relart, sheet_name=s4),
                s5: all_fields_values(df_reltrk, req_reltrk, sheet_name=s5),
                s6: all_fields_values(df_trkart, req_trkart, sheet_name=s6),
                s7: all_fields_values(df_comp_masters, req_comp_masters, sheet_name=s7),
                s8: all_fields_values(df_trkcomp, req_trkcomp, sheet_name=s8),
                s9: all_fields_values(df_props, req_props, sheet_name=s9),
            }
            write_json_artifact("reports", "all_fields_values.json", all_values, "Sample values for all parsed fields")
            # Log a tiny sample pointing to columns likely to be present (top by fill-rate)
            sample_cols = []
            for sh, rep in all_report.items():
                arr = rep.get("columns", [])
                if arr:
                    top = arr[0]
                    sample_cols.append({"sheet": sh, "column": top.get("column"), "fill_rate": top.get("fill_rate")})
            s.info(values_artifact="all_fields_values.json", sample_top=sample_cols[:5])

        # ===== Preflight validations
        report: List[Dict[str,Any]] = []

        # Required-field checks (yellow)
        for name, df, req in [
            (s1, df_art, req_art),
            (s2, df_lab, req_lab),
            (s3, df_rel, req_rel),
            (s4, df_relart, req_relart),
            (s5, df_reltrk, req_reltrk),
            (s6, df_trkart, req_trkart),
            (s7, df_comp_masters, req_comp_masters),
            (s8, df_trkcomp, req_trkcomp),
            (s9, df_props, req_props),
        ]:
            errs = require_columns(df, req)

            release_role_autofill = autofill_release_artist_roles(df_relart, req_relart)
            if release_role_autofill:
                artifact_path = ARTIFACTS / "release_artist_autofill.json"
                artifact_path.write_text(json.dumps(release_role_autofill, indent=2, ensure_ascii=False))
                s.info(release_artist_autofilled=len(release_role_autofill))
            for rownum, msg in errs:
                report.append({"sheet": name, "row": rownum, "error": msg})

        # Cross-tab keys & integrity
        # Releases must have UPC / EAN / JAN (join key), Tracks must have ISRC/vISRC
        def expect_col(df, name):
            return has_col(df, name)

        UPC_COL = "UPC / EAN / JAN"
        ISRC_COL = "ISRC/vISRC"

        if expect_col(df_rel, UPC_COL):
            cm_rel = make_colmap(df_rel)
            for i, rw in df_rel.iterrows():
                if is_nan(get_val(rw, cm_rel, UPC_COL)):
                    report.append({"sheet": s3, "row": i+2, "error": "Missing release key 'UPC / EAN / JAN'"})

        if expect_col(df_reltrk, UPC_COL) and expect_col(df_reltrk, ISRC_COL):
            cm_reltrk = make_colmap(df_reltrk)
            # also check track order is present
            for i, rw in df_reltrk.iterrows():
                if is_nan(get_val(rw, cm_reltrk, UPC_COL)):
                    report.append({"sheet": s5, "row": i+2, "error": "Release_Track missing UPC to join Release"})
                if is_nan(get_val(rw, cm_reltrk, ISRC_COL)):
                    report.append({"sheet": s5, "row": i+2, "error": "Release_Track missing ISRC/vISRC"})

        if expect_col(df_trkcomp, ISRC_COL) and has_col(df_trkcomp, "SHARE%"):
            # shares sum to 100 by (ISRC) or 1.0 when decimal representation is used
            cm_trkcomp = make_colmap(df_trkcomp)
            by_isrc: Dict[str, float] = {}
            for i, rw in df_trkcomp.iterrows():
                isrc = norm_str(get_val(rw, cm_trkcomp, ISRC_COL))
                share_s = norm_str(get_val(rw, cm_trkcomp, "SHARE%"))
                share = None
                try:
                    share = float(share_s) if share_s is not None else None
                except Exception:
                    pass
                if isrc and share is not None:
                    by_isrc[isrc] = by_isrc.get(isrc, 0.0) + share
            for isrc, total in by_isrc.items():
                # Accept either 100-based or decimal-based totals (1.0 == 100%)
                tol = 1e-3
                ok_100 = abs(total - 100.0) <= tol
                ok_1 = abs(total - 1.0) <= tol
                if not (ok_100 or ok_1):
                    report.append({"sheet": s8, "row": "-", "error": f"Composition shares for ISRC {isrc} sum to {total}, expected ~100 or ~1.0"})

        # Property conflicts
        if expect_col(df_props, ISRC_COL):
            cm_props = make_colmap(df_props)
            for i, rw in df_props.iterrows():
                ids, _ = map_track_properties(rw.to_dict())
                if ids and 1 in ids and len(ids) > 1:
                    report.append({"sheet": s9, "row": i+2, "error": "Track properties: 'None' cannot be combined with other flags"})

        # Stop if any blocking issues
        with progress.step("Preflight validations") as s:
            s.info(issues=len(report))
            if report:
                out = ARTIFACTS / "preflight_report.json"
                out.write_text(json.dumps(report, indent=2))
                print(f"[BLOCKED] Preflight failed with {len(report)} issue(s). See {out.resolve()}")
                progress.write_log()
                sys.exit(1)

    # ===== Build master maps (Artists, Labels, Composers, Publishers)
    artist_image_tasks: List[Dict[str, Any]] = []
    composer_master_by_name: Dict[str, Dict[str, Any]] = {}
    publisher_master_by_name: Dict[str, Dict[str, Any]] = {}
    # Artists
    with progress.step("Build artists & labels & master entities") as s:
        cm_art = make_colmap(df_art)
        # Try robust resolution of the artist name column with several aliases
        art_name_col = resolve_colkey(
            df_art,
            "Artist Name", "ARTIST NAME", "Artist",
            "Artist Full Name", "ArtistName", "Name"
        )
        # Ultimate fallback: choose the first column whose header contains 'artist' and 'name' tokens
        if not art_name_col:
            for c in df_art.columns:
                if isinstance(c, str):
                    key = norm_colkey(c)
                    if "artist" in key and ("name" in key or key.endswith("artist")):
                        art_name_col = c
                        break
        if not art_name_col:
            print("[WARN] Could not resolve 'Artist Name' column; artist list may be empty.")
        artists_payload = []
        artist_name_to_obj = {}
        dropped_art_missing_name = 0
        artists_with_external_ids = 0
        artist_external_source_counts = {"apple": 0, "spotify": 0, "meta": 0, "soundcloud": 0}
        for i, rw in df_art.iterrows():
            # fallback: use resolved column if present
            name = norm_str(rw.get(art_name_col)) if art_name_col else norm_str(get_val(rw, cm_art, "Artist Name", "ARTIST NAME", "Artist"))
            if not name:
                dropped_art_missing_name += 1
                continue
            img_url = norm_str(get_val(rw, cm_art, "Artist Image url", "Artist Image URL"))
            apple = norm_str(get_val(rw, cm_art, "Apple ArtistId"))
            spotify = norm_str(get_val(rw, cm_art, "Spotify Artist URI"))
            spotify_id = extract_spotify_artist_id(spotify) if spotify else None
            meta = norm_str(get_val(rw, cm_art, "Meta ArtistId"))
            sc = norm_str(get_val(rw, cm_art, "SoundCloud ProfileId", "SoundCloud Profile ID"))
            isni_raw = norm_str(get_val(rw, cm_art, "ISNI"))
            isni, isni_issue = normalize_isni(isni_raw)
            isni_payload_val: Optional[str] = None
            if isni:
                isni_payload_val = isni
            elif isni_raw:
                # Preserve contributor-provided ISNI even when formatting checks fail so the API receives the value.
                sanitized = re.sub(r"[^0-9A-Za-z]", "", isni_raw.upper())
                isni_payload_val = sanitized or isni_raw
            ext = []
            sources_used: List[str] = []
            if apple:
                ext.append({"distributorStoreId": 1, "profileId": apple})
                sources_used.append("apple")
            if spotify_id:
                ext.append({"distributorStoreId": 9, "profileId": spotify_id})
                sources_used.append("spotify")
            if sc:
                ext.append({"distributorStoreId": 68, "profileId": sc})
                sources_used.append("soundcloud")
            if meta:
                ext.append({"distributorStoreId": 309, "profileId": meta})
                sources_used.append("meta")
            if ext:
                artists_with_external_ids += 1
                for src in sources_used:
                    if src in artist_external_source_counts:
                        artist_external_source_counts[src] += 1
            payload = {"name": name}
            if ext:
                payload["artistExternalIds"] = ext
            if isni_payload_val is not None:
                payload["isni"] = isni_payload_val
            if isni_raw and isni_issue:
                debug_trace["identifier_warnings"].append({
                    "entity": "artist",
                    "name": name,
                    "field": "isni",
                    "value": isni_raw,
                    "issue": isni_issue
                })
            if img_url:
                artist_image_tasks.append({"name": name, "url": img_url, "payload": payload})
                # Dry-run visibility: include placeholder filename/source
                if not args.live:
                    payload.setdefault("image", {"fileId": None, "filename": _filename_from_url(img_url), "sourceUrl": img_url})
            artists_payload.append(payload)
            artist_name_to_obj[name.lower()] = payload
            add_debug_sample(debug_trace["artists_raw"], {
                "name": name,
                "apple": apple,
                "spotify_raw": spotify,
                "spotify_id": spotify_id,
                "meta": meta,
                "soundcloud": sc,
                "externalIds": ext,
                "isni": payload.get("isni"),
                "imageSourceUrl": img_url
            })

        # Labels
        cm_lab = make_colmap(df_lab)
        lab_name_col = resolve_colkey(df_lab, "Label Name", "LABEL NAME", "Label")
        labels_payload = []
        label_name_to_id = {}
        dropped_lab_missing_name = 0
        for i, rw in df_lab.iterrows():
            lname = norm_str(rw.get(lab_name_col)) if lab_name_col else norm_str(get_val(rw, cm_lab, "Label Name", "LABEL NAME", "Label"))
            if not lname:
                dropped_lab_missing_name += 1
                continue
            labels_payload.append({"name": lname})

        # Publishers & Composers
        cm_cm = make_colmap(df_comp_masters)
        pub_col = resolve_colkey(df_comp_masters, "Publisher Name", "PUBLISHER NAME", "Publisher")
        comp_col = resolve_colkey(df_comp_masters, "Composition Contributor", "COMPOSITION CONTRIBUTOR", "Contributor")
        # Robust dual IPI detection with validity-based scoring.
        # Strategy:
        # 1. Identify publisher name column position to separate contributor vs publisher regions.
        # 2. Collect all columns whose header matches an IPI pattern ("IPI/CAE", "IPI CAE", "IPI").
        # 3. Pre-publisher region: pick earliest valid IPI column as contributor IPI.
        # 4. Post-publisher region: score candidates by number of rows containing a valid normalized IPI; choose highest.
        #    If no valid IPIs found post-publisher but at least one header candidate exists, choose the first candidate as a fallback.
        # 5. Provide rich debug stats (raw headers sample, candidate lists, counts, chosen one-based indices).
        contributor_ipi_col_index: Optional[int] = None
        publisher_ipi_col_index: Optional[int] = None
        publisher_ipi_candidate_indices: List[int] = []
        publisher_ipi_header: Optional[str] = None  # Prefer explicit publisher IPI header if present
        ipi_detection_debug: Dict[str, Any] = {}
        try:
            raw_headers = list(df_comp_masters.columns)
            pub_positions = [i for i, h in enumerate(raw_headers) if isinstance(h, str) and h.strip().lower() in ("publisher name", "publisher", "publisher name ")]
            first_pub_pos = min(pub_positions) if pub_positions else None
            ipi_positions = [i for i, h in enumerate(raw_headers) if isinstance(h, str) and h.strip().lower() in ("ipi/cae", "ipi cae", "ipi")]
            pre_pub_candidates = [i for i in ipi_positions if first_pub_pos is not None and i < first_pub_pos]
            post_pub_candidates = [i for i in ipi_positions if first_pub_pos is None or i >= first_pub_pos]

            def count_valid_ipis(col_index: int) -> int:
                cnt = 0
                try:
                    for _, rw in df_comp_masters.iterrows():
                        raw_val = norm_str(rw.iloc[col_index])
                        if raw_val:
                            norm_val, _issue = normalize_ipi_cae(raw_val)
                            if norm_val:
                                cnt += 1
                except Exception:
                    return 0
                return cnt

            # Simplified hard-coded contributor/publisher IPI columns for sheet 7.
            if s7 == "7) Comp ContributorPublisher li":
                publisher_name_index = first_pub_pos
                if pre_pub_candidates:
                    contributor_ipi_col_index = pre_pub_candidates[0]
                elif ipi_positions:
                    contributor_ipi_col_index = ipi_positions[0]
                else:
                    contributor_ipi_col_index = 3  # fallback (one-based 4)

                if publisher_name_index is not None:
                    publisher_ipi_candidate_indices = [idx for idx in ipi_positions if idx > publisher_name_index]
                else:
                    publisher_ipi_candidate_indices = ipi_positions[1:]

                if publisher_ipi_candidate_indices:
                    publisher_ipi_col_index = publisher_ipi_candidate_indices[0]
                    header_candidate = raw_headers[publisher_ipi_col_index]
                    publisher_ipi_header = header_candidate if isinstance(header_candidate, str) else None
                else:
                    publisher_ipi_col_index = 8  # fallback (one-based 9)
                ipi_detection_debug = {
                    "mode": "hard-coded",
                    "contributor_ipi_col_index_one_based": 4,
                    "publisher_ipi_col_index_one_based": (publisher_ipi_col_index + 1) if publisher_ipi_col_index is not None else None,
                    "publisher_ipi_header": publisher_ipi_header,
                    "publisher_ipi_candidate_indices_one_based": [idx + 1 for idx in publisher_ipi_candidate_indices],
                    "raw_headers_sample": raw_headers[:25]  # include a slice for debugging
                }
            else:
                ipi_detection_debug = {
                    "mode": "not-applicable",
                    "contributor_ipi_col_index_one_based": None,
                    "publisher_ipi_col_index_one_based": None,
                    "publisher_ipi_header": None
                }
        except Exception as e:
            ipi_detection_debug = {"error": f"ipi detection failed: {e}"}

        # Debug visibility of detected columns (overwrite existing structure if present)
        debug_trace["ipi_column_detection"] = ipi_detection_debug
        publishers_payload = []
        composers_payload = []
        publisher_master_by_name.clear()
        composer_master_by_name.clear()
        if pub_col:
            for ridx, rw in df_comp_masters.iterrows():
                pn = norm_str(rw.get(pub_col))
                if not pn:
                    continue
                key = pn.lower()
                existing_pub = publisher_master_by_name.get(key)
                if existing_pub is None:
                    existing_pub = {"name": pn}
                    publisher_master_by_name[key] = existing_pub
                    publishers_payload.append(existing_pub)
                    add_debug_sample(debug_trace["publishers_raw"], {"name": pn})
                if publisher_lookup:
                    lookup_obj = publisher_lookup.get(key)
                    if lookup_obj and "publisherId" in lookup_obj and existing_pub.get("publisherId") is None:
                        try:
                            existing_pub["publisherId"] = int(lookup_obj.get("publisherId"))
                        except Exception:
                            existing_pub["publisherId"] = lookup_obj.get("publisherId")
                # Optional publisher IPI/CAE extraction (9 or 11 digits) strictly from publisher IPI column region
                contrib_ipi_raw = None
                contrib_ipi_norm = None
                if contributor_ipi_col_index is not None:
                    try:
                        contrib_ipi_raw = norm_str(rw.iloc[contributor_ipi_col_index])
                        if contrib_ipi_raw:
                            contrib_ipi_norm, _ = normalize_ipi_cae(contrib_ipi_raw)
                    except Exception:
                        contrib_ipi_raw = None

                candidate_values: List[Dict[str, Any]] = []
                candidate_indices_seen: Set[int] = set()

                if publisher_ipi_header:
                    header_obj = None
                    try:
                        header_obj = rw.loc[publisher_ipi_header]
                    except Exception:
                        header_obj = None
                    if isinstance(header_obj, pd.Series):
                        for dup_idx, val in enumerate(header_obj.tolist()):
                            candidate_values.append({
                                "source": f"header[{dup_idx}]",
                                "index": None,
                                "raw": norm_str(val)
                            })
                    else:
                        candidate_values.append({
                            "source": "header",
                            "index": publisher_ipi_col_index,
                            "raw": norm_str(header_obj)
                        })

                for idx in publisher_ipi_candidate_indices:
                    candidate_indices_seen.add(idx)
                    try:
                        val = norm_str(rw.iloc[idx])
                    except Exception:
                        val = None
                    candidate_values.append({
                        "source": f"index[{idx}]",
                        "index": idx,
                        "raw": val
                    })

                if publisher_ipi_col_index is not None and publisher_ipi_col_index not in candidate_indices_seen:
                    candidate_indices_seen.add(publisher_ipi_col_index)
                    try:
                        val = norm_str(rw.iloc[publisher_ipi_col_index])
                    except Exception:
                        val = None
                    candidate_values.append({
                        "source": "primary_index",
                        "index": publisher_ipi_col_index,
                        "raw": val
                    })

                fallback_idx = 8
                if fallback_idx not in candidate_indices_seen:
                    candidate_indices_seen.add(fallback_idx)
                    try:
                        val = norm_str(rw.iloc[fallback_idx])
                    except Exception:
                        val = None
                    candidate_values.append({
                        "source": "fallback_index8",
                        "index": fallback_idx,
                        "raw": val
                    })

                candidate_debug_entries: List[Dict[str, Any]] = []
                chosen_entry: Optional[Dict[str, Any]] = None
                chosen_norm: Optional[str] = None
                contributor_match_entry: Optional[Tuple[str, Dict[str, Any]]] = None

                for cand in candidate_values:
                    raw_val = cand.get("raw")
                    norm_val = None
                    issue = None
                    if raw_val:
                        norm_val, issue = normalize_ipi_cae(raw_val)
                    cand_debug = {
                        "source": cand.get("source"),
                        "index_zero_based": cand.get("index"),
                        "raw": raw_val,
                        "normalized": norm_val,
                        "issue": issue
                    }
                    candidate_debug_entries.append(cand_debug)

                    if not raw_val or not norm_val:
                        continue
                    if contrib_ipi_norm and norm_val == contrib_ipi_norm:
                        if contributor_match_entry is None:
                            contributor_match_entry = (norm_val, cand)
                        continue
                    chosen_entry = cand
                    chosen_norm = norm_val
                    break

                if chosen_entry is None and contributor_match_entry is not None:
                    chosen_norm, chosen_entry = contributor_match_entry

                add_debug_sample(
                    debug_trace.setdefault("publisher_ipi_candidate_values", []),
                    {
                        "row": ridx + 1,
                        "publisher": pn,
                        "contributor_ipi": contrib_ipi_norm,
                        "candidates": candidate_debug_entries,
                        "chosen_source": chosen_entry.get("source") if chosen_entry else None
                    }
                )

                if chosen_entry and chosen_norm:
                    existing_pub["ipiCae"] = chosen_norm
                    existing_pub["_ipi_source_col"] = chosen_entry.get("index")
                    debug_trace.setdefault("publisher_ipi_assignments", []).append({
                        "name": pn,
                        "action": "assign",
                        "val": chosen_norm,
                        "src_zero_based": chosen_entry.get("index"),
                        "source": chosen_entry.get("source")
                    })
                elif any(cand.get("raw") for cand in candidate_values):
                    debug_trace["identifier_warnings"].append({
                        "entity": "publisher",
                        "name": pn,
                        "field": "ipi_cae",
                        "value": None,
                        "issue": "publisher_ipi_not_found"
                    })
                # Publisher country mapping -> countryId
                pub_country_raw = norm_str(get_val(rw, cm_cm, "Publisher Country", "Publisher country"))
                if pub_country_raw:
                    cid_obj = countries_lookup.get(pub_country_raw.lower())
                    if cid_obj and cid_obj.get("countryId") is not None:
                        try:
                            existing_pub.setdefault("countryId", int(cid_obj.get("countryId")))
                        except Exception:
                            existing_pub.setdefault("countryId", cid_obj.get("countryId"))
                    else:
                        debug_trace["identifier_warnings"].append({
                            "entity": "publisher",
                            "name": pn,
                            "field": "country",
                            "value": pub_country_raw,
                            "issue": "not_found"
                        })
        if comp_col:
            for ridx, rw in df_comp_masters.iterrows():
                cn = norm_str(rw.get(comp_col))
                if not cn:
                    continue
                key = cn.lower()
                comp_entry = composer_master_by_name.get(key)
                if comp_entry is None:
                    comp_entry = {"name": cn}
                    composer_master_by_name[key] = comp_entry
                    composers_payload.append(comp_entry)
                if composer_lookup:
                    lookup_obj = composer_lookup.get(key)
                    if lookup_obj and "composerId" in lookup_obj and comp_entry.get("composerId") is None:
                        try:
                            comp_entry["composerId"] = int(lookup_obj.get("composerId"))
                        except Exception:
                            comp_entry["composerId"] = lookup_obj.get("composerId")
                isni_raw = norm_str(get_val(rw, cm_cm, "ISNI", "Composer ISNI", "Contribution ISNI"))
                if isni_raw:
                    isni_val, isni_issue = normalize_isni(isni_raw)
                    if isni_val:
                        if comp_entry.get("isni") and comp_entry.get("isni") != isni_val:
                            debug_trace["identifier_warnings"].append({
                                "entity": "composer",
                                "name": cn,
                                "field": "isni",
                                "value": isni_raw,
                                "issue": "conflict_with_existing"
                            })
                        else:
                            comp_entry.setdefault("isni", isni_val)
                    else:
                        debug_trace["identifier_warnings"].append({
                            "entity": "composer",
                            "name": cn,
                            "field": "isni",
                            "value": isni_raw,
                            "issue": isni_issue or "invalid"
                        })
                # Contributor IPI value solely from contributor ipi column region (avoid using publisher ipi column)
                ipi_raw = None
                if contributor_ipi_col_index is not None:
                    try:
                        ipi_raw = norm_str(rw.iloc[contributor_ipi_col_index])
                    except Exception:
                        ipi_raw = None
                if ipi_raw:
                    ipi_val, ipi_issue = normalize_ipi_cae(ipi_raw)
                    if ipi_val:
                        if comp_entry.get("ipiCae") and comp_entry.get("ipiCae") != ipi_val:
                            debug_trace["identifier_warnings"].append({
                                "entity": "composer",
                                "name": cn,
                                "field": "ipi_cae",
                                "value": ipi_raw,
                                "issue": "conflict_with_existing"
                            })
                        else:
                            comp_entry.setdefault("ipiCae", ipi_val)
                    else:
                        debug_trace["identifier_warnings"].append({
                            "entity": "composer",
                            "name": cn,
                            "field": "ipi_cae",
                            "value": ipi_raw,
                            "issue": ipi_issue or "invalid"
                        })
                # Contributor country mapping -> countryOfResidenceId
                comp_country_raw = norm_str(get_val(rw, cm_cm, "Contributor Country", "Contribution Country", "Composer Country"))
                if comp_country_raw:
                    cid_obj = countries_lookup.get(comp_country_raw.lower())
                    if cid_obj and cid_obj.get("countryId") is not None:
                        try:
                            comp_entry.setdefault("countryOfResidenceId", int(cid_obj.get("countryId")))
                        except Exception:
                            comp_entry.setdefault("countryOfResidenceId", cid_obj.get("countryId"))
                    else:
                        debug_trace["identifier_warnings"].append({
                            "entity": "composer",
                            "name": cn,
                            "field": "country",
                            "value": comp_country_raw,
                            "issue": "not_found"
                        })
                comp_debug_rec = {
                    "composerName": cn,
                    "source": "masters",
                    "isni": comp_entry.get("isni"),
                    "ipiCae": comp_entry.get("ipiCae"),
                }
                if contributor_ipi_col_index is not None:
                    comp_debug_rec["contributor_ipi_source_col_one_based"] = contributor_ipi_col_index + 1
                add_debug_sample(debug_trace["composer_entries"], comp_debug_rec)
        # Include small samples in the log for quick visibility
        art_samples = []
        if art_name_col:
            for i in range(min(3, len(df_art))):
                art_samples.append(norm_str(df_art.iloc[i].get(art_name_col)))
        lab_samples = []
        if lab_name_col:
            for i in range(min(3, len(df_lab))):
                lab_samples.append(norm_str(df_lab.iloc[i].get(lab_name_col)))
        pub_samples = []
        if pub_col:
            for i in range(min(3, len(df_comp_masters))):
                pub_samples.append(norm_str(df_comp_masters.iloc[i].get(pub_col)))
        comp_samples = []
        if comp_col:
            for i in range(min(3, len(df_comp_masters))):
                comp_samples.append(norm_str(df_comp_masters.iloc[i].get(comp_col)))
        s.info(
            artists=len(artists_payload), labels=len(labels_payload), publishers=len(publishers_payload), composers=len(composers_payload),
            artists_seen=len(df_art), labels_seen=len(df_lab), dropped_art_missing_name=dropped_art_missing_name, dropped_lab_missing_name=dropped_lab_missing_name,
            artist_name_col=art_name_col, label_name_col=lab_name_col, publisher_col=pub_col, composer_col=comp_col,
            artist_samples=art_samples, label_samples=lab_samples, publisher_samples=pub_samples, composer_samples=comp_samples,
            artists_with_external_ids=artists_with_external_ids, external_id_sources=artist_external_source_counts
        )

        release_meta_by_key: Dict[str, Dict[str, Any]] = {}
        release_order: List[str] = []
        release_key_by_upc: Dict[str, str] = {}
        release_key_by_row: Dict[int, str] = {}
        release_keys_in_payload_order: List[str] = []
        track_meta_by_isrc: Dict[str, Dict[str, Any]] = {}
        track_order: List[str] = []
        track_isrcs_by_release_key: Dict[str, List[str]] = defaultdict(list)

        # ===== Releases & Tracks
        with progress.step("Build releases") as s:
            cm_rel = make_colmap(df_rel)
            releases_payload = []
            tracks_payload = []  # list of (release_key, payload)
            upc_dupes_logged = []
            releases_with_upc = 0
            releases_missing_upc = 0
            releases_with_p_line = 0
            releases_with_c_line = 0
            releases_with_version = 0

            # Build effective headers for Release_Label sheet to resolve subheaders (e.g., (P)/(C) Year/Holder)
            rel_effective_by_index: Dict[int, str] = {}
            rel_header_meta: Dict[int, Dict[str, Optional[str]]] = {}
            cols_rel = list(df_rel.columns)
            try:
                wb_rel = load_workbook(xlsx_path, data_only=True)
                ws_rel = wb_rel[s3]
            except Exception:
                ws_rel = None
            for j, c in enumerate(cols_rel):
                h3s: Optional[str] = None
                h4s: Optional[str] = None
                eff: Optional[str] = None
                if ws_rel is not None:
                    try:
                        colnum = j + 1
                        h3 = ws_rel.cell(row=3, column=colnum).value
                        h4 = ws_rel.cell(row=4, column=colnum).value
                        h3s = norm_str(h3)
                        h4s = norm_str(h4)
                        # If row-3 is a grouped header and row-4 is a subheader, combine
                        if h3s and h4s and any(k in h3s.lower() for k in ["copyright", "release title", "cover image", "upc", "label", "genre", "original release date", "title language"]):
                            # Special-case (P)/(C) Copyright groups: build "(P) Copyright Year" style labels
                            if "copyright" in h3s.lower():
                                eff = f"{h3s} {h4s}".strip()
                            else:
                                eff = h3s  # for non-copyright, row3 is specific enough
                        else:
                            eff = h3s or h4s or norm_str(c) or f"col_{j}"
                    except Exception:
                        eff = None
                if not eff:
                    eff = norm_str(c) or f"col_{j}"
                rel_effective_by_index[j] = eff
                rel_header_meta[j] = {"row3": h3s, "row4": h4s, "effective": eff}
            # Build reverse map eff -> indices
            rel_eff_to_idx: Dict[str, List[int]] = {}
            for j, name in rel_effective_by_index.items():
                rel_eff_to_idx.setdefault(name, []).append(j)

            release_debug_entries = debug_trace["release_copyright_debug"]
            release_debug_cap = 200
            release_sheet_row_offset = int(req_rel.get("_effective_data_start", 5))

            def _serializable(val: Any) -> Any:
                if is_nan(val):
                    return None
                if isinstance(val, (str, int, float, bool)) or val is None:
                    return val
                try:
                    coerced = val.item()
                    if is_nan(coerced):
                        return None
                    return coerced
                except Exception:
                    return str(val)

            def log_release_debug(row_index: int, target: str, stage: str, *, idx: Optional[int] = None, value: Any = None, note: Optional[Dict[str, Any]] = None, meta: Optional[Dict[str, Any]] = None) -> None:
                if len(release_debug_entries) >= release_debug_cap:
                    return
                entry: Dict[str, Any] = {
                    "sheet_row": release_sheet_row_offset + row_index,
                    "target": target,
                    "stage": stage,
                }
                if idx is not None:
                    entry["column_index"] = idx
                    entry["effective_header"] = rel_effective_by_index.get(idx)
                    entry["meta"] = meta if meta is not None else rel_header_meta.get(idx)
                if value is not None:
                    value_ser = _serializable(value)
                    if value_ser is not None:
                        entry["value"] = value_ser
                if note:
                    entry["note"] = note
                release_debug_entries.append(entry)

            def rel_val(row: pd.Series, eff_name: str) -> Optional[Any]:
                idxs = rel_eff_to_idx.get(eff_name) or []
                if not idxs:
                    return None
                try:
                    return row.iloc[idxs[0]]
                except Exception:
                    return None
            def rel_pull(
                row: pd.Series,
                row_index: int,
                target: str,
                *,
                row3_tokens: Tuple[str, ...] = (),
                row4_tokens: Tuple[str, ...] = (),
                fallback_names: Tuple[str, ...] = (),
                allow_missing_row3: bool = False,
                allow_missing_row4: bool = False,
                preferred_indices: Tuple[Optional[int], ...] = (),
                validator: Optional[Callable[[Any], bool]] = None
            ) -> Tuple[Optional[Any], Optional[int]]:
                origins: Dict[int, str] = {}
                candidate_idxs: List[int] = []

                def _push(idx: Optional[int], origin: str) -> None:
                    if idx is None:
                        return
                    if idx < 0 or idx >= len(cols_rel):
                        return
                    candidate_idxs.append(idx)
                    origins.setdefault(idx, origin)

                for idx in preferred_indices:
                    _push(idx, "preferred")

                if row3_tokens or row4_tokens:
                    for idx, meta in rel_header_meta.items():
                        text3 = meta.get("row3") if meta else None
                        text4 = meta.get("row4") if meta else None
                        cond3 = not row3_tokens or text3 and all(tok in text3.lower() for tok in row3_tokens)
                        cond4 = not row4_tokens or text4 and all(tok in text4.lower() for tok in row4_tokens)
                        if cond3 and cond4:
                            _push(idx, "primary")
                for name in fallback_names:
                    idx_list = rel_eff_to_idx.get(name)
                    if idx_list:
                        for idx in idx_list:
                            _push(idx, "fallback_effective")
                for name in fallback_names:
                    norm_name = (name or "").strip().lower()
                    if not norm_name:
                        continue
                    for idx, col in enumerate(cols_rel):
                        if not isinstance(col, str):
                            continue
                        col_norm = col.strip().lower()
                        if col_norm != norm_name and not col_norm.startswith(f"{norm_name}."):
                            continue
                        _push(idx, "fallback_scan")

                # Build snapshot for diagnostics
                snapshot: List[Dict[str, Any]] = []
                seen_snapshot: set[int] = set()
                for idx in candidate_idxs:
                    if idx in seen_snapshot:
                        continue
                    seen_snapshot.add(idx)
                    meta = rel_header_meta.get(idx, {})
                    try:
                        raw_val = row.iloc[idx]
                    except Exception:
                        raw_val = None
                    snapshot.append({
                        "column_index": idx,
                        "effective": rel_effective_by_index.get(idx),
                        "meta": meta,
                        "raw": _serializable(raw_val),
                        "origin": origins.get(idx)
                    })
                log_release_debug(row_index, target, "invoke", note={"candidates": snapshot, "row3_tokens": row3_tokens, "row4_tokens": row4_tokens, "fallback_names": fallback_names} if snapshot else None)

                def tokens_match(text: Optional[str], tokens: Tuple[str, ...], allow_missing: bool) -> bool:
                    if not tokens:
                        return True
                    if not text:
                        return allow_missing
                    low = text.lower()
                    return all(tok in low for tok in tokens)

                ordered: List[int] = []
                seen: set[int] = set()
                for idx in candidate_idxs:
                    if idx in seen:
                        continue
                    seen.add(idx)
                    ordered.append(idx)

                for idx in ordered:
                    meta = rel_header_meta.get(idx, {})
                    origin = origins.get(idx, "primary")
                    if origin != "preferred":
                        if row3_tokens and not tokens_match(meta.get("row3"), row3_tokens, allow_missing_row3):
                            continue
                        if row4_tokens and not tokens_match(meta.get("row4"), row4_tokens, allow_missing_row4):
                            continue
                    try:
                        val = row.iloc[idx]
                    except Exception:
                        val = None
                    if is_nan(val):
                        continue
                    if validator and not validator(val):
                        log_release_debug(row_index, target, "candidate_rejected", idx=idx, value=val, note={"origin": origin})
                        continue
                    stage = {
                        "preferred": "preferred_match",
                        "fallback_effective": "fallback_effective",
                        "fallback_scan": "fallback_scan",
                        "primary": "primary_match"
                    }.get(origin, "primary_match")
                    log_release_debug(row_index, target, stage, idx=idx, value=val, note={"origin": origin})
                    return val, idx

                log_release_debug(row_index, target, "no_match")
                return None, None

            def neighbor_scan(
                row: pd.Series,
                row_index: int,
                start_idx: Optional[int],
                *,
                directions: Tuple[int, ...],
                limit: int,
                predicate: Callable[[Any, int], bool],
                target: str,
                note: Optional[Dict[str, Any]] = None
            ) -> Tuple[Optional[Any], Optional[int]]:
                if start_idx is None:
                    return None, None
                for direction in directions:
                    steps = 0
                    idx = start_idx
                    while True:
                        idx += direction
                        if idx < 0 or idx >= len(cols_rel):
                            break
                        steps += 1
                        if limit and steps > limit:
                            break
                        try:
                            val = row.iloc[idx]
                        except Exception:
                            val = None
                        if is_nan(val):
                            continue
                        if predicate(val, idx):
                            extra_note = {"direction": direction, "steps": steps}
                            if note:
                                extra_note.update(note)
                            log_release_debug(row_index, target, "neighbor_fallback", idx=idx, value=val, note=extra_note)
                            return val, idx
                return None, None

            for i,rw in df_rel.iterrows():
                upc = norm_str(get_val(rw, cm_rel, UPC_COL))
                title = norm_str(get_val(rw, cm_rel, "RELEASE TITLE"))
                version = norm_str(get_val(rw, cm_rel, "RELEASE VERSION"))
                title_lang = norm_str(get_val(rw, cm_rel, "TITLE LANGUAGE"))
                img_url = norm_str(get_val(rw, cm_rel, "COVER IMAGE URL", "COVER IMAGE url"))
                vendor_release_id = norm_str(get_val(rw, cm_rel, "RELEASE ID", "Release ID", "RELEASES RELEASE ID", "RELEASES RELEASE ID#", "VENDOR RELEASE ID"))
                grid = norm_str(get_val(rw, cm_rel, "GRID", "GRid", "Release Grid"))
                status_text = norm_str(get_val(rw, cm_rel, "STATUS", "Release Status"))
                parental_flag = norm_str(get_val(rw, cm_rel, "PARENTAL ADVISORY", "Parental Advisory", "EXPLICIT"))
                total_tracks_expected = norm_int(get_val(rw, cm_rel, "TOTAL TRACKS", "TRACK COUNT", "NUMBER OF TRACKS"))
                release_date_raw = norm_str(get_val(rw, cm_rel, "ORIGINAL RELEASE DATE", "ORIGINAL\nRELEASE DATE"))
                # Pull (P)/(C) from subheaders when present
                p_year, p_year_idx = rel_pull(
                    rw,
                    i,
                    "p_copyright_year",
                    row3_tokens=("(p)", "copyright"),
                    row4_tokens=("year",),
                    fallback_names=("(P) Copyright Year", "(P) Copyright"),
                    allow_missing_row4=True
                )
                holder_preferred = (p_year_idx + 1,) if p_year_idx is not None else tuple()
                p_holder, p_holder_idx = rel_pull(
                    rw,
                    i,
                    "p_copyright_holder",
                    row3_tokens=("(p)", "copyright"),
                    row4_tokens=("holder",),
                    fallback_names=("(P) Copyright Holder", "Holder"),
                    allow_missing_row3=True,
                    allow_missing_row4=True,
                    preferred_indices=holder_preferred
                )
                if not norm_str(p_holder):
                    fallback_val, fallback_idx = neighbor_scan(
                        rw,
                        i,
                        p_year_idx,
                        directions=(1,),
                        limit=3,
                        predicate=lambda v, _: bool(norm_str(v)) and norm_int(v) is None,
                        target="p_copyright_holder"
                    )
                    if fallback_val is not None:
                        p_holder, p_holder_idx = fallback_val, fallback_idx

                c_year, c_year_idx = rel_pull(
                    rw,
                    i,
                    "c_copyright_year",
                    row3_tokens=("(c)", "copyright"),
                    row4_tokens=("year",),
                    fallback_names=("(C) Copyright Year", "(C) Copyright"),
                    allow_missing_row4=True,
                    validator=lambda v: norm_int(v) is not None
                )
                if norm_int(c_year) is None:
                    fallback_val, fallback_idx = neighbor_scan(
                        rw,
                        i,
                        c_year_idx if c_year_idx is not None else p_year_idx,
                        directions=(1,),
                        limit=4,
                        predicate=lambda v, _: norm_int(v) is not None,
                        target="c_copyright_year"
                    )
                    if fallback_val is not None:
                        c_year, c_year_idx = fallback_val, fallback_idx

                holder_preferred_c = (c_year_idx - 1,) if c_year_idx is not None else tuple()
                c_holder, c_holder_idx = rel_pull(
                    rw,
                    i,
                    "c_copyright_holder",
                    row3_tokens=("(c)", "copyright"),
                    row4_tokens=("holder",),
                    fallback_names=("(C) Copyright Holder", "Holder"),
                    allow_missing_row3=True,
                    allow_missing_row4=True,
                    preferred_indices=holder_preferred_c
                )
                if not norm_str(c_holder):
                    fallback_val, fallback_idx = neighbor_scan(
                        rw,
                        i,
                        c_year_idx,
                        directions=(-1, 1),
                        limit=4,
                        predicate=lambda v, idx: bool(norm_str(v)) and norm_int(v) is None and idx != p_holder_idx,
                        target="c_copyright_holder"
                    )
                    if fallback_val is not None:
                        c_holder, c_holder_idx = fallback_val, fallback_idx
                p_line = parse_year_holder(p_year, p_holder)
                c_line = parse_year_holder(c_year, c_holder)
                if len(release_debug_entries) < release_debug_cap:
                    release_debug_entries.append({
                        "sheet_row": release_sheet_row_offset + i,
                        "target": "summary",
                        "stage": "resolved",
                        "upc": upc,
                        "title": title,
                        "value": {
                            "p_year": _serializable(p_year),
                            "p_holder": _serializable(p_holder),
                            "c_year": _serializable(c_year),
                            "c_holder": _serializable(c_holder),
                            "p_line": _serializable(p_line),
                            "c_line": _serializable(c_line),
                        }
                    })
                g1 = norm_str(get_val(rw, cm_rel, "GENRE 1")); g2 = norm_str(get_val(rw, cm_rel, "GENRE 2"))
                label_name = norm_str(get_val(rw, cm_rel, "LABEL", "Label Name", "LABEL NAME"))

                lang_id = resolve_language_id(title_lang, session, base_url, token)
                g1_id = resolve_musicstyle_id(g1, session, base_url, token)
                g2_id = resolve_musicstyle_id(g2, session, base_url, token)

                img = ingest_image_by_url(img_url, session, base_url, token) if img_url else None
                rel = {
                    "name": title, "version": version,
                    "previouslyReleased": bool(release_date_raw),
                    "releaseDate": release_date_raw,
                }
                if upc:
                    rel["upc"] = upc
                    releases_with_upc += 1
                else:
                    releases_missing_upc += 1
                if p_line:
                    rel["copyrightP"] = p_line
                    releases_with_p_line += 1
                if c_line:
                    rel["copyrightC"] = c_line
                    releases_with_c_line += 1
                if version:
                    releases_with_version += 1
                if lang_id:
                    # language of the release title also dictates the release-level languageId
                    rel.setdefault("releaseLocals", []).append({"languageId": lang_id, "name": title})
                    rel["languageId"] = lang_id
                if g1_id: rel["primaryMusicStyleId"] = g1_id
                if g2_id: rel["secondaryMusicStyleId"] = g2_id
                if label_name:
                    rel["hasRecordLabel"] = True
                    rel["labelName"] = label_name
                # Only attach image if we have a valid fileId; otherwise, skip to avoid 400s on null GUID
                if img and img.get("fileId"):
                    rel["image"] = {"fileId": img["fileId"], "filename": img["filename"]}
                if img_url:
                    rel["imageSourceUrl"] = img_url

                release_key = vendor_release_id or upc or f"release_row_{release_sheet_row_offset + i}"
                if release_key in release_meta_by_key:
                    release_key = f"{release_key}#{release_sheet_row_offset + i}"
                release_meta = {
                    "key": release_key,
                    "upc": upc,
                    "vendorReleaseId": vendor_release_id or release_key,
                    "grid": grid,
                    "title": title,
                    "versionTitle": version,
                    "artistName": None,
                    "originalReleaseDate": release_date_raw,
                    "status": status_text,
                    "genreDesc": g1 or g2,
                    "metadataLanguageCode": title_lang,
                    "parentalAdvisory": parental_flag,
                    "pLine": p_line,
                    "cLine": c_line,
                    "totalTracks": total_tracks_expected,
                    "contributors_raw": [],
                    "extra": {
                        "previouslyReleased": rel.get("previouslyReleased"),
                        "releaseDate": rel.get("releaseDate"),
                        "languageId": lang_id,
                        "releaseLocals": copy.deepcopy(rel.get("releaseLocals", [])),
                        "primaryMusicStyleId": g1_id,
                        "secondaryMusicStyleId": g2_id,
                        "hasRecordLabel": rel.get("hasRecordLabel", False),
                        "labelName": label_name,
                        "image": copy.deepcopy(rel.get("image")) if rel.get("image") else None,
                        "imageSourceUrl": rel.get("imageSourceUrl"),
                        "artistExternalIds": None,
                        "artistExternalIdsSource": None,
                        "languageName": title_lang,
                        "primaryGenreDesc": g1,
                        "secondaryGenreDesc": g2,
                        "sheetRow": release_sheet_row_offset + i,
                    }
                }
                release_meta_by_key[release_key] = release_meta
                release_order.append(release_key)
                release_key_by_row[i] = release_key
                if upc and upc not in release_key_by_upc:
                    release_key_by_upc[upc] = release_key
                release_keys_in_payload_order.append(release_key)
                releases_payload.append(rel)
                add_debug_sample(debug_trace["releases_raw"], {
                    "upc": upc,
                    "title": title,
                    "version": version,
                    "copyrightP": p_line,
                    "copyrightC": c_line,
                    "languageId": lang_id,
                    "labelName": label_name
                })
            sample_releases = []
            for i in range(min(3, len(releases_payload))):
                rr = releases_payload[i]
                sample_releases.append({k: rr.get(k) for k in ("name","version","upc","labelName")})
            s.info(
                releases=len(releases_payload), sample_releases=sample_releases,
                with_upc=releases_with_upc, missing_upc=releases_missing_upc,
                with_p_line=releases_with_p_line, with_c_line=releases_with_c_line,
                with_version=releases_with_version
            )

        # Release contributors
        with progress.step("Parse release contributors") as s:
            cm_relart = make_colmap(df_relart)
            release_contribs_by_upc: Dict[str,List[Dict[str,Any]]] = {}
            release_primary_artist_by_upc: Dict[str,str] = {}
            if has_col(df_relart, UPC_COL) and has_col(df_relart, "ARTIST") and has_col(df_relart, "ARTIST ROLE"):
                for _,rw in df_relart.iterrows():
                    upc = norm_str(get_val(rw, cm_relart, UPC_COL))
                    artist = norm_str(get_val(rw, cm_relart, "ARTIST"))
                    role = norm_str(get_val(rw, cm_relart, "ARTIST ROLE"))
                    if not upc or not artist or not role:
                        continue
                    role_norm = (role or '').strip().lower()
                    # Special case: 'Main Primary Artist' sets release.artistName, not a contributor
                    if role_norm == "main primary artist":
                        release_primary_artist_by_upc.setdefault(upc, artist)
                        rel_key = release_key_by_upc.get(upc)
                        if rel_key:
                            rel_meta = release_meta_by_key.get(rel_key)
                            if rel_meta is not None:
                                rel_meta["artistName"] = artist
                        continue
                    rid = role_map.get(role_norm, None)
                    if rid is None:
                        print(f"[WARN] Unknown role '{role}' for release UPC {upc}")
                        continue
                    release_contribs_by_upc.setdefault(upc, []).append({
                        "artistName": artist, "roleId": rid, "roleName": role
                    })
                    rel_key = release_key_by_upc.get(upc)
                    if rel_key:
                        rel_meta = release_meta_by_key.get(rel_key)
                        if rel_meta is not None:
                            rel_meta.setdefault("contributors_raw", []).append({
                                "name": artist,
                                "roleName": role,
                                "extra": {"roleId": rid}
                            })
            total = sum(len(v) for v in release_contribs_by_upc.values())
            sample_rel_contribs = []
            for upc, arr in list(release_contribs_by_upc.items())[:2]:
                sample_rel_contribs.append({"upc": upc, "first": arr[0] if arr else None})
            sample_rel_primary = []
            for upc, name in list(release_primary_artist_by_upc.items())[:2]:
                sample_rel_primary.append({"upc": upc, "artistName": name})
            s.info(contributors=total, primaries=len(release_primary_artist_by_upc), sample=sample_rel_contribs, sample_primary=sample_rel_primary)

        # Tracks (by Release_Track)
        audio_url_map: Dict[str, Optional[str]] = {}
        audio_upload_logs: List[Dict[str, Any]] = []
        release_iswc_by_isrc: Dict[str, Set[str]] = {}
        with progress.step("Build tracks from Release_Track") as s:
            cm_reltrk = make_colmap(df_reltrk)
            audio_url_col = resolve_colkey(df_reltrk, "AUDIO FILE URL", "AUDIO URL", "AUDIO DOWNLOAD URL", "AUDIO FILE", "FILE URL", "AUDIO")
            audio_type_col = resolve_colkey(df_reltrk, "AUDIO TYPE", "FILE TYPE", "AUDIO FORMAT", "FORMAT")
            track_rows = []
            first_audio_url_seen = None
            isrc_to_track: Dict[str, Dict[str, Any]] = {}
            tracks_with_preview = 0
            tracks_missing_preview = 0
            tracks_with_version = 0
            for _, rw in df_reltrk.iterrows():
                upc = norm_str(get_val(rw, cm_reltrk, UPC_COL)); isrc = norm_str(get_val(rw, cm_reltrk, ISRC_COL))
                if not upc or not isrc:
                    continue
                t_title = norm_str(get_val(rw, cm_reltrk, "TRACK TITLE")); t_version = norm_str(get_val(rw, cm_reltrk, "TRACK VERSION"))
                lang = norm_str(get_val(rw, cm_reltrk, "LANGUAGE OF LYRICS", "LANGUAGE"))
                metadata_lang = norm_str(get_val(rw, cm_reltrk, "METADATA LANGUAGE", "LANGUAGE OF METADATA"))
                track_genre = norm_str(get_val(rw, cm_reltrk, "GENRE", "TRACK GENRE"))
                length_seconds = parse_duration_seconds(get_val(rw, cm_reltrk, "DURATION", "TRACK DURATION", "LENGTH", "DURATION (S)", "Duration (mm:ss)"))
                track_original_release = norm_str(get_val(rw, cm_reltrk, "TRACK ORIGINAL RELEASE DATE", "ORIGINAL TRACK RELEASE DATE", "ORIGINAL RELEASE DATE"))
                vendor_track_id = norm_str(get_val(rw, cm_reltrk, "TRACK RESOURCE ID", "Resource ID", "TRACK ID", "TRACK RESOURCE", "RESOURCE ID"))
                explicit = norm_bool(get_val(rw, cm_reltrk, "EXPLICIT"))
                ttype = norm_str(get_val(rw, cm_reltrk, "TYPE"))
                ttype_id = {"original": 1, "cover": 2, "public domain": 3}.get((ttype or "").strip().lower())
                iswc_raw_release = norm_str(get_val(rw, cm_reltrk, "COMPOSITION ISWC", "COMPOSITION\n ISWC", "ISWC"))
                if iswc_raw_release:
                    iswc_norm_release, iswc_issue_release = normalize_iswc(iswc_raw_release)
                    if iswc_norm_release:
                        release_iswc_by_isrc.setdefault(isrc, set()).add(iswc_norm_release)
                    elif iswc_issue_release:
                        debug_trace["identifier_warnings"].append({
                            "entity": "track",
                            "isrc": isrc,
                            "field": "iswc",
                            "value": iswc_raw_release,
                            "issue": iswc_issue_release
                        })
                audio_url_raw = norm_str(rw.get(audio_url_col)) if audio_url_col else norm_str(get_val(rw, cm_reltrk, "AUDIO FILE URL"))
                audio_url = normalize_audio_url(audio_url_raw)
                audio_type = norm_str(rw.get(audio_type_col)) if audio_type_col else norm_str(get_val(rw, cm_reltrk, "AUDIO TYPE"))
                preview = norm_int(get_val(rw, cm_reltrk, "TRACK PREVIEW", "PREVIEW START"))
                trknum = norm_int(get_val(rw, cm_reltrk, "TRACK", "TRACK #", "TRACK NUMBER"))  # track number
                lang_id = resolve_language_id(lang, session, base_url, token)
                if preview is not None:
                    tracks_with_preview += 1
                else:
                    tracks_missing_preview += 1
                if t_version:
                    tracks_with_version += 1

                audio = ingest_audio_by_url(audio_url, audio_type, session, base_url, token, args.live, headers={"X-EnterpriseId": str(enterpriseId), "X-TenantId": str(tenantId)}, isrc=isrc, upload_log=audio_upload_logs) if audio_url else None
                if first_audio_url_seen is None and audio_url:
                    first_audio_url_seen = audio_url
                if isrc:
                    audio_url_map[isrc] = audio_url
                parental_text = None
                if explicit is True:
                    parental_text = "Explicit"
                elif explicit is False:
                    parental_text = "Not Explicit"
                track = {
                    "name": t_title,
                    "version": t_version,
                    "languageId": lang_id,
                    "explicit": explicit,
                    "trackType": ttype_id,
                    "trackNumber": trknum,
                    "previewStartSeconds": preview,
                    "trackRecordingVersions": [{
                        "isrc": isrc,
                        "recordingVersionType": ttype_id,
                        # Only attach audioFiles objects when we have an uploaded audioId.
                        # External URLs are not accepted directly here.
                        "audioFiles": ([{
                            "audioId": audio["audioId"],
                            "audioFilename": audio.get("audioFilename"),
                            "fileFormat": audio.get("fileFormat")
                        }] if (audio and audio.get("audioId")) else [])
                    }]
                }
                release_key = release_key_by_upc.get(upc)
                track_meta = {
                    "release_key": release_key,
                    "vendorTrackId": vendor_track_id,
                    "isrc": isrc,
                    "numberInRelease": trknum,
                    "title": t_title,
                    "versionTitle": t_version,
                    "artistName": None,
                    "lengthSeconds": length_seconds,
                    "audioLanguageCode": lang,
                    "metadataLanguageCode": metadata_lang or lang,
                    "genreDesc": track_genre,
                    "parentalAdvisory": parental_text,
                    "originalReleaseDate": track_original_release,
                    "contributors_raw": [],
                    "composer_raw": [],
                    "extra": {
                        "languageId": lang_id,
                        "explicit": explicit,
                        "trackType": ttype_id,
                        "previewStartSeconds": preview,
                        "trackRecordingVersions": copy.deepcopy(track.get("trackRecordingVersions", [])),
                        "artistExternalIds": None,
                        "trackProperties": None,
                        "compositions": [],
                        "composerContentsDTO": [],
                        "trackLocals": None,
                        "audioUrl": audio_url,
                        "audioType": audio_type,
                        "upc": upc,
                    }
                }
                track_meta_by_isrc[isrc] = track_meta
                track_order.append(isrc)
                if release_key:
                    track_isrcs_by_release_key[release_key].append(isrc)
                track_rows.append((upc, isrc, track))
                isrc_to_track[isrc] = track
                add_debug_sample(debug_trace["tracks_raw"], {
                    "upc": upc,
                    "isrc": isrc,
                    "title": t_title,
                    "version": t_version,
                    "previewStartSeconds": preview,
                    "trackNumber": trknum,
                    "languageId": lang_id,
                    "trackType": ttype_id,
                    "audioUrl": audio_url,
                    "audioType": audio_type
                })
            sample_tracks = []
            for i in range(min(3, len(track_rows))):
                upc, isrc, track = track_rows[i]
                sample_tracks.append({"upc": upc, "isrc": isrc, "name": track.get("name"), "trackNumber": track.get("trackNumber")})
            s.info(
                tracks=len(track_rows), sample_tracks=sample_tracks,
                first_audio_url=first_audio_url_seen, audio_url_col=audio_url_col, audio_type_col=audio_type_col,
                preview_with=tracks_with_preview, preview_missing=tracks_missing_preview,
                tracks_with_version=tracks_with_version
            )

        # Track contributors
        with progress.step("Parse track contributors") as s:
            cm_trkart = make_colmap(df_trkart)
            track_contribs_by_isrc: Dict[str,List[Dict[str,Any]]] = {}
            track_primary_artist_by_isrc: Dict[str,str] = {}
            if has_col(df_trkart, ISRC_COL) and has_col(df_trkart, "ARTIST") and has_col(df_trkart, "ARTIST ROLE"):
                for _,rw in df_trkart.iterrows():
                    isrc = norm_str(get_val(rw, cm_trkart, ISRC_COL))
                    artist = norm_str(get_val(rw, cm_trkart, "ARTIST"))
                    role = norm_str(get_val(rw, cm_trkart, "ARTIST ROLE"))
                    if not isrc or not artist or not role:
                        continue
                    role_norm = (role or '').strip().lower()
                    if role_norm == "main primary artist":
                        track_primary_artist_by_isrc.setdefault(isrc, artist)
                        meta = track_meta_by_isrc.get(isrc)
                        if meta is not None:
                            meta["artistName"] = artist
                        continue
                    rid = role_map.get(role_norm, None)
                    if rid is None:
                        print(f"[WARN] Unknown role '{role}' for track ISRC {isrc}")
                        continue
                    track_contribs_by_isrc.setdefault(isrc, []).append({
                        "artistName": artist, "roleId": rid
                    })
                    meta = track_meta_by_isrc.get(isrc)
                    if meta is not None:
                        meta.setdefault("contributors_raw", []).append({
                            "name": artist,
                            "roleName": role,
                            "extra": {"roleId": rid}
                        })
            total = sum(len(v) for v in track_contribs_by_isrc.values())
            sample_trk_contribs = []
            for isrc, arr in list(track_contribs_by_isrc.items())[:2]:
                sample_trk_contribs.append({"isrc": isrc, "first": arr[0] if arr else None})
            sample_trk_primary = []
            for isrc, name in list(track_primary_artist_by_isrc.items())[:2]:
                sample_trk_primary.append({"isrc": isrc, "artistName": name})
            s.info(contributors=total, primaries=len(track_primary_artist_by_isrc), sample=sample_trk_contribs, sample_primary=sample_trk_primary)

        # Track compositions
        with progress.step("Parse track compositions") as s:
            cm_trkcomp = make_colmap(df_trkcomp)
            trk_comp_by_isrc: Dict[str,List[Dict[str,Any]]] = {}
            comp_warnings: List[Dict[str, Any]] = []
            for _,rw in df_trkcomp.iterrows():
                isrc = norm_str(get_val(rw, cm_trkcomp, ISRC_COL)); comp = norm_str(get_val(rw, cm_trkcomp, "COMPOSITION CONTRIBUTOR"))
                role = norm_str(get_val(rw, cm_trkcomp, "ROLE")); share_s = norm_str(get_val(rw, cm_trkcomp, "SHARE%"))
                rights = norm_str(get_val(rw, cm_trkcomp, "PUBLISHING")); publisher = norm_str(get_val(rw, cm_trkcomp, "PUBLISHER"))
                if not isrc or not comp or not role or not share_s:
                    continue
                share_num = None
                try:
                    share_num = float(share_s)
                except Exception:
                    share_num = None
                rightsId, rights_reason = resolve_rights_id(rights)
                if rights_reason:
                    issue = "missing_rights" if rights_reason == "missing" else "unrecognized_rights"
                    comp_warnings.append({
                        "isrc": isrc,
                        "composer": comp,
                        "issue": issue,
                        "raw_rights": rights,
                        "defaulted_to": rightsId
                    })
                iswc_raw = norm_str(get_val(rw, cm_trkcomp, "COMPOSITION ISWC", "COMPOSITION ISWC", "ISWC"))
                iswc_norm, iswc_issue = normalize_iswc(iswc_raw)
                role_key = normalize_role_key(role)
                role_rec = composer_role_map.get(role_key)
                role_id = int(role_rec["roleId"]) if role_rec else None
                if role_id is None:
                    comp_warnings.append({"isrc": isrc, "composer": comp, "role": role, "issue": "unknown_role"})
                    continue
                entry: Dict[str, Any] = {
                    "composerName": comp,
                    "roleName": role,
                    "roleId": role_id,
                    "share": share_s,
                    "share_num": share_num,
                    "rightsId": rightsId,
                }
                if iswc_norm:
                    entry["iswc"] = iswc_norm
                elif iswc_issue and iswc_raw:
                    comp_warnings.append({
                        "isrc": isrc,
                        "composer": comp,
                        "issue": f"iswc_{iswc_issue}",
                        "raw_iswc": iswc_raw
                    })
                if rightsId == 2:
                    if publisher:
                        pub_key = publisher.lower()
                        pub_obj = publisher_lookup.get(pub_key) if publisher_lookup else None
                        publisher_id: Optional[int] = None
                        if pub_obj:
                            try:
                                publisher_id = int(pub_obj.get("publisherId") or pub_obj.get("id"))
                            except Exception:
                                publisher_id = None
                        if publisher_id is None:
                            publisher_id = 0
                        entry["publisherName"] = publisher
                        entry["publisherId"] = publisher_id
                    else:
                        comp_warnings.append({"isrc": isrc, "composer": comp, "issue": "missing_publisher_for_published"})
                        continue
                # Capture row-level identifiers (fallback to master sheet when absent)
                row_isni_raw = norm_str(get_val(rw, cm_trkcomp, "COMPOSITION CONTRIBUTOR ISNI", "CONTRIBUTOR ISNI", "ISNI"))
                if row_isni_raw:
                    row_isni, row_isni_issue = normalize_isni(row_isni_raw)
                    if row_isni:
                        entry["isni"] = row_isni
                    else:
                        debug_trace["identifier_warnings"].append({
                            "entity": "track_composer",
                            "isrc": isrc,
                            "composer": comp,
                            "field": "isni",
                            "value": row_isni_raw,
                            "issue": row_isni_issue or "invalid"
                        })
                row_ipi_raw = norm_str(get_val(rw, cm_trkcomp, "COMPOSITION CONTRIBUTOR IPI/CAE", "CONTRIBUTOR IPI/CAE", "IPI/CAE", "IPI CAE", "IPI"))
                if row_ipi_raw:
                    row_ipi, row_ipi_issue = normalize_ipi_cae(row_ipi_raw)
                    if row_ipi:
                        entry["ipiCae"] = row_ipi
                    else:
                        debug_trace["identifier_warnings"].append({
                            "entity": "track_composer",
                            "isrc": isrc,
                            "composer": comp,
                            "field": "ipi_cae",
                            "value": row_ipi_raw,
                            "issue": row_ipi_issue or "invalid"
                        })
                if composer_master_by_name:
                    master_comp = composer_master_by_name.get(comp.lower())
                    if master_comp:
                        if master_comp.get("isni") and not entry.get("isni"):
                            entry["isni"] = master_comp.get("isni")
                        if master_comp.get("ipiCae") and not entry.get("ipiCae"):
                            entry["ipiCae"] = master_comp.get("ipiCae")
                trk_comp_by_isrc.setdefault(isrc, []).append(entry)
                meta = track_meta_by_isrc.get(isrc)
                if meta is not None:
                    composer_extra = {
                        "roleId": role_id,
                        "share": share_s,
                        "shareNumeric": share_num,
                        "rightsId": rightsId,
                        "publisherName": entry.get("publisherName"),
                        "publisherId": entry.get("publisherId"),
                        "composerId": entry.get("composerId"),
                        "iswc": entry.get("iswc"),
                        "isni": entry.get("isni"),
                        "ipiCae": entry.get("ipiCae"),
                    }
                    meta.setdefault("composer_raw", []).append({
                        "name": comp,
                        "roleName": role,
                        "extra": composer_extra
                    })
                debug_comp_entry = {
                    "isrc": isrc,
                    "composerName": comp,
                    "roleName": role,
                    "roleId": role_id,
                    "share": share_s,
                    "rightsId": rightsId,
                    "publisherName": entry.get("publisherName")
                }
                if entry.get("iswc"):
                    debug_comp_entry["iswc"] = entry["iswc"]
                if entry.get("isni"):
                    debug_comp_entry["isni"] = entry["isni"]
                if entry.get("ipiCae"):
                    debug_comp_entry["ipiCae"] = entry["ipiCae"]
                add_debug_sample(debug_trace["track_compositions"], debug_comp_entry)
            total = sum(len(v) for v in trk_comp_by_isrc.values())
            sample_comps = []
            for isrc, arr in list(trk_comp_by_isrc.items())[:2]:
                sample_comps.append({"isrc": isrc, "first": arr[0] if arr else None})
            if comp_warnings:
                debug_trace["composer_warnings"].extend(comp_warnings[:50])
            s.info(compositions=total, sample=sample_comps, warnings=len(comp_warnings))

        # Track properties
        with progress.step("Parse track properties") as s:
            cm_props = make_colmap(df_props)
            # Build effective headers PER COLUMN POSITION using sheet row 3/4 so we disambiguate duplicate
            # 'SPECIAL AUDIO PROPERTIES' columns into unique subheaders (row 4 values).
            try:
                wb_props = load_workbook(xlsx_path, data_only=True)
                ws_props = wb_props[s9]
            except Exception:
                wb_props = None; ws_props = None
            effective_headers_by_index: Dict[int, str] = {}
            mapped_cols: Dict[str, str] = {}
            # Map DF columns to Excel column indices by matching header names by occurrence order
            excel_pos_by_name: Dict[str, List[int]] = {}
            if ws_props is not None:
                try:
                    max_col = ws_props.max_column
                    for colnum in range(1, max_col + 1):
                        h3 = ws_props.cell(row=3, column=colnum).value
                        name = norm_str(h3) or ""
                        key = (name or "").strip().lower()
                        excel_pos_by_name.setdefault(key, []).append(colnum)
                except Exception:
                    excel_pos_by_name = {}
            used_count: Dict[str, int] = {}
            for j, c in enumerate(list(df_props.columns)):
                eff: Optional[str] = None
                excel_colnum: Optional[int] = None
                if ws_props is not None:
                    # try match by header name occurrence
                    cname = norm_str(c) or ""
                    ckey = cname.strip().lower()
                    pos_list = excel_pos_by_name.get(ckey) or []
                    idx = used_count.get(ckey, 0)
                    if idx < len(pos_list):
                        excel_colnum = pos_list[idx]
                        used_count[ckey] = idx + 1
                # Fallback: approximate by DF index position
                if ws_props is not None and excel_colnum is None:
                    excel_colnum = j + 1
                if ws_props is not None and excel_colnum is not None:
                    try:
                        h3 = ws_props.cell(row=3, column=excel_colnum).value
                        h4 = ws_props.cell(row=4, column=excel_colnum).value
                        h3s = norm_str(h3)
                        h4s = norm_str(h4)
                        if h3s and h3s.strip().lower() == "special audio properties" and h4s:
                            eff = h4s
                        else:
                            eff = h3s or h4s
                    except Exception:
                        eff = None
                if not eff:
                    eff = norm_str(c) or f"col_{j}"
                effective_headers_by_index[j] = eff
                mapped_cols[f"{j}:{c}"] = eff

            props_by_isrc: Dict[str,List[int]] = {}
            props_diag: Dict[str, Any] = {}
            for _, rw in df_props.iterrows():
                isrc = norm_str(get_val(rw, cm_props, ISRC_COL))
                if not isrc:
                    continue
                # Build a row dict keyed by effective headers using positional indexing to handle duplicates
                row_dict: Dict[str, Any] = {}
                for j in range(len(df_props.columns)):
                    eh = effective_headers_by_index.get(j)
                    if not eh:
                        continue
                    try:
                        val = rw.iloc[j]
                    except Exception:
                        val = None
                    # Don't overwrite an already set key (keep first occurrence)
                    if eh not in row_dict:
                        row_dict[eh] = val
                ids, diag = map_track_properties(row_dict)
                if ids:
                    props_by_isrc[isrc] = ids
                    meta = track_meta_by_isrc.get(isrc)
                    if meta is not None:
                        meta.setdefault("extra", {}).update({"trackProperties": ids})
                if diag:
                    # include which headers were mapped for extra visibility
                    diag["_effective_headers"] = mapped_cols
                    props_diag[isrc] = diag
            sample_props = []
            for isrc, arr in list(props_by_isrc.items())[:2]:
                sample_props.append({"isrc": isrc, "props": arr})
            s.info(with_properties=len(props_by_isrc), defaulted=sum(1 for d in props_diag.values() if d.get("defaulted_to_none")), sample=sample_props)

        # Attach contributors/compositions/properties
        with progress.step("Attach track contributors/compositions/properties") as s:
            comp_share_diag: Dict[str, Any] = {}
            for idx,(upc,isrc,track) in enumerate(track_rows):
                meta = track_meta_by_isrc.get(isrc)
                # Contributors
                if isrc in track_contribs_by_isrc:
                    applied = []
                    for c in track_contribs_by_isrc[isrc]:
                        # Use full artist object if available
                        artist_name_key = (c.get("artistName") or "").strip().lower()
                        full_artist = artist_name_to_obj.get(artist_name_key)
                        if full_artist:
                            artist_entry = copy.deepcopy(full_artist)
                        else:
                            artist_entry = {"name": c["artistName"]}
                        applied.append({"roleId": c["roleId"], "artist": artist_entry})
                    if applied:
                        track["contributors"] = applied
                        if meta is not None:
                            meta.setdefault("extra", {}).update({"contributorsPayload": copy.deepcopy(applied)})
                # Primary artistName from Track_Artist(s)
                if 'artistName' not in track and isrc in track_primary_artist_by_isrc:
                    track['artistName'] = track_primary_artist_by_isrc[isrc]
                if meta is not None and track.get('artistName'):
                    meta["artistName"] = track.get('artistName')
                # If we have a primary artistName and that artist has known external IDs, attach them (optional field)
                try:
                    nm = (track.get('artistName') or '').strip()
                    if nm:
                        ext_map = locals().get('artist_external_by_name', {}) or {}
                        ext = ext_map.get(nm.lower())
                        if not ext:
                            art_obj = artist_name_to_obj.get(nm.lower())
                            ext = (art_obj or {}).get('artistExternalIds')
                        if ext:
                            track['artistExternalIds'] = copy.deepcopy(ext)
                            if meta is not None:
                                meta.setdefault("extra", {}).update({"artistExternalIds": copy.deepcopy(ext)})
                except Exception:
                    pass
                # Compositions
                compositions_out: List[Dict[str, Any]] = []
                seen_iswc: Set[str] = set()
                release_iswcs = release_iswc_by_isrc.get(isrc)
                if release_iswcs:
                    for iswc_val in sorted(release_iswcs):
                        if iswc_val and iswc_val not in seen_iswc:
                            seen_iswc.add(iswc_val)
                            compositions_out.append({"iswc": iswc_val})
                if isrc in trk_comp_by_isrc:
                    # Determine scale and convert to 0-100 percentages for API
                    entries = trk_comp_by_isrc[isrc]
                    for cc in entries:
                        iswc_val = cc.get("iswc")
                        if not iswc_val or iswc_val in seen_iswc:
                            continue
                        seen_iswc.add(iswc_val)
                        compositions_out.append({"iswc": iswc_val})
                    nums = [cc.get("share_num") for cc in entries if cc.get("share_num") is not None]
                    total = sum(nums) if nums else None
                    tol = 1e-3
                    scale = "unknown"
                    if total is not None:
                        if abs(total - 1.0) <= tol:
                            scale = "unit"
                        elif abs(total - 100.0) <= tol:
                            scale = "percent"
                        else:
                            scale = "mixed"
                    def fmt_pct(x: float) -> str:
                        # Format as int if near-integer, else keep up to 4 decimals
                        rx = round(x)
                        if abs(x - rx) <= 1e-6:
                            return str(int(rx))
                        return ("%0.4f" % x).rstrip("0").rstrip(".")
                    comp_out = []
                    out_vals = []
                    for cc in entries:
                        s_num = cc.get("share_num")
                        if s_num is None:
                            # fallback: attempt parse again
                            try:
                                s_num = float(cc.get("share") or 0)
                            except Exception:
                                s_num = 0.0
                        pct_val = s_num*100.0 if scale == "unit" else s_num
                        role_id_out = cc.get("roleId")
                        rights_id_out = cc.get("rightsId")
                        if role_id_out is None or rights_id_out is None:
                            comp_warnings_entry = {
                                "isrc": isrc,
                                "composer": cc.get("composerName"),
                                "issue": "missing_role_or_rights_at_attach"
                            }
                            comp_share_diag.setdefault(isrc, {}).setdefault("warnings", []).append(comp_warnings_entry)
                            continue
                        out_vals.append(pct_val)
                        item: Dict[str, Any] = {
                            "share": fmt_pct(pct_val),
                            "composerName": cc["composerName"],
                            "roleId": role_id_out,
                            "rightsId": rights_id_out
                        }
                        if cc.get("composerId"):
                            item["composerId"] = cc["composerId"]
                        if cc.get("isni"):
                            item["isni"] = cc["isni"]
                        if cc.get("ipiCae"):
                            item["ipiCae"] = cc["ipiCae"]
                        if rights_id_out == 2:
                            item["publisherName"] = cc.get("publisherName")
                            if cc.get("publisherId") is not None:
                                item["publisherId"] = cc.get("publisherId")
                        # Composer locals (use track language and version when available)
                        comp_local = {}
                        if track.get("languageId"):
                            comp_local["languageId"] = track.get("languageId")
                        comp_local_name = cc.get("composerName")
                        if comp_local_name:
                            comp_local["name"] = comp_local_name
                        track_version = track.get("version")
                        if track_version:
                            comp_local["version"] = track_version
                        if comp_local:
                            item["composersLocals"] = [comp_local]
                        add_debug_sample(debug_trace["composer_entries"], {
                            "isrc": isrc,
                            "composerName": cc.get("composerName"),
                            "roleId": role_id_out,
                            "rightsId": rights_id_out,
                            "publisherName": cc.get("publisherName"),
                            "isni": cc.get("isni"),
                            "ipiCae": cc.get("ipiCae")
                        })
                        comp_out.append(item)
                    comp_share_diag[isrc] = {
                        "in_values": nums,
                        "in_total": total,
                        "scale_detected": scale,
                        "out_values": out_vals,
                        "out_total": sum(out_vals) if out_vals else None
                    }
                    if compositions_out:
                        track["compositions"] = compositions_out
                    if comp_out:
                        track["composerContentsDTO"] = comp_out
                    if meta is not None:
                        if compositions_out:
                            meta.setdefault("extra", {}).update({"compositions": copy.deepcopy(compositions_out)})
                        if comp_out:
                            meta.setdefault("extra", {}).update({"composerContentsDTO": copy.deepcopy(comp_out)})
                else:
                    if compositions_out:
                        track["compositions"] = compositions_out
                        if meta is not None:
                            meta.setdefault("extra", {}).update({"compositions": copy.deepcopy(compositions_out)})
                # Properties
                if isrc in props_by_isrc:
                    track["trackProperties"] = props_by_isrc[isrc]
                    if meta is not None:
                        meta.setdefault("extra", {}).update({"trackProperties": props_by_isrc[isrc]})
                else:
                    # Default to NONE APPLY when properties row is missing
                    track["trackProperties"] = [1]
                    if meta is not None:
                        meta.setdefault("extra", {}).update({"trackProperties": [1]})
                # Ensure required keys exist even if empty
                track.setdefault("contributors", [])
                track.setdefault("composerContentsDTO", [])
                # Optional trackLocals: include when we have at least name and languageId; version is optional
                if "trackLocals" not in track:
                    tl_name = track.get("name")
                    tl_lang = track.get("languageId")
                    if tl_name and tl_lang:
                        loc = {"name": tl_name, "languageId": tl_lang}
                        if track.get("version"):
                            loc["version"] = track.get("version")
                        track["trackLocals"] = [loc]
                        if meta is not None:
                            meta.setdefault("extra", {}).update({"trackLocals": copy.deepcopy(track["trackLocals"])})
                elif meta is not None and track.get("trackLocals"):
                    meta.setdefault("extra", {}).update({"trackLocals": copy.deepcopy(track["trackLocals"])})
                add_debug_sample(debug_trace["final_tracks"], {
                    "upc": upc,
                    "isrc": isrc,
                    "previewStartSeconds": track.get("previewStartSeconds"),
                    "version": track.get("version"),
                    "composerCount": len(track.get("composerContentsDTO") or []),
                    "compositionsCount": len(track.get("compositions") or []),
                    "contributorsCount": len(track.get("contributors") or []),
                    "trackProperties": track.get("trackProperties"),
                    "artistExternalIds": track.get("artistExternalIds"),
                    "recordingVersions": [
                        {
                            "isrc": rv.get("isrc"),
                            "recordingVersionType": rv.get("recordingVersionType"),
                            "audioFiles": rv.get("audioFiles")
                        }
                        for rv in track.get("trackRecordingVersions", [])
                    ]
                })
                tracks_payload.append((upc, track))
            s.info(tracks=len(tracks_payload))

        # Attach release contributors
        with progress.step("Attach release contributors") as s:
            for rel_index, rel in enumerate(releases_payload):
                upc = rel.get("upc")
                release_key = None
                if rel_index < len(release_keys_in_payload_order):
                    release_key = release_keys_in_payload_order[rel_index]
                if not release_key and upc:
                    release_key = release_key_by_upc.get(upc)
                rel_meta = release_meta_by_key.get(release_key) if release_key else None
                if not upc and release_key is None:
                    continue
                # Apply primary artistName from Release_Artist(s)
                if 'artistName' not in rel and upc in release_primary_artist_by_upc:
                    rel['artistName'] = release_primary_artist_by_upc[upc]
                if rel_meta is not None and rel.get('artistName'):
                    rel_meta["artistName"] = rel.get('artistName')
                # If we now have an artistName, include their external IDs if available
                try:
                    nm = (rel.get('artistName') or '').strip()
                    if nm:
                        ext_map = locals().get('artist_external_by_name', {}) or {}
                        ext = ext_map.get(nm.lower())
                        if not ext:
                            art_obj = artist_name_to_obj.get(nm.lower())
                            ext = (art_obj or {}).get('artistExternalIds')
                        if ext:
                            rel['artistExternalIds'] = copy.deepcopy(ext)
                            if rel_meta is not None:
                                rel_meta.setdefault("extra", {}).update({"artistExternalIds": copy.deepcopy(ext)})
                except Exception:
                    pass
                if upc in release_contribs_by_upc:
                    applied = []
                    for c in release_contribs_by_upc[upc]:
                        # Use full artist object if available
                        artist_name_key = (c.get("artistName") or "").strip().lower()
                        full_artist = artist_name_to_obj.get(artist_name_key)
                        if full_artist:
                            artist_entry = copy.deepcopy(full_artist)
                        else:
                            artist_entry = {"name": c["artistName"]}
                        applied.append({"roleId": c["roleId"], "artist": artist_entry})
                    if applied:
                        rel["contributors"] = applied
                        if rel_meta is not None:
                            rel_meta.setdefault("extra", {}).update({"contributorsPayload": copy.deepcopy(applied)})
                add_debug_sample(debug_trace["final_releases"], {
                    "upc": upc,
                    "name": rel.get("name"),
                    "version": rel.get("version"),
                    "copyrightP": rel.get("copyrightP"),
                    "copyrightC": rel.get("copyrightC"),
                    "artistName": rel.get("artistName"),
                    "labelName": rel.get("labelName"),
                    "contributorsCount": len(rel.get("contributors") or []),
                    "hasReleaseLocals": bool(rel.get("releaseLocals"))
                })
            s.info(releases=len(releases_payload))

        # Pre-populate artist_name_to_obj with image filenames from artist_image_tasks
        # so catalog_to_api_payloads can reference full artist objects with images
        for task in artist_image_tasks:
            art_name = task.get("name")
            if art_name:
                name_key = art_name.lower()
                if name_key in artist_name_to_obj:
                    # Add image filename reference (will be enriched with fileId later during upload)
                    img_url = task.get("url")
                    if img_url and "image" not in artist_name_to_obj[name_key]:
                        artist_name_to_obj[name_key]["image"] = {
                            "filename": _filename_from_url(img_url),
                            "fileId": None  # Will be filled during image upload
                        }

        with progress.step("Build canonical catalog model") as s:
            catalog = build_catalog_from_meta(
                release_order,
                release_meta_by_key,
                track_meta_by_isrc,
                track_isrcs_by_release_key,
            )
            catalog_path = write_json_artifact("raw", "catalog.json", catalog.to_dict(), "Canonical catalog snapshot used for payload generation")
            catalog_releases_payload, catalog_tracks_payload = catalog_to_api_payloads(catalog, artist_name_to_obj)
            if len(catalog_releases_payload) != len(releases_payload):
                print(
                    f"[WARN] Canonical release count mismatch: catalog={len(catalog_releases_payload)} raw={len(releases_payload)}"
                )
            releases_payload = catalog_releases_payload
            tracks_payload = catalog_tracks_payload
            release_keys_in_payload_order = release_order.copy()
            s.info(releases=len(releases_payload), tracks=len(tracks_payload), catalog_path=str(catalog_path))

        # ===== Emit dry-run artifacts
        with progress.step("Write dry-run artifacts") as s:
            # Enrich labels with existing labelId for transparency
            try:
                headers_tmp = {"X-EnterpriseId": str(enterpriseId), "X-TenantId": str(tenantId)}
                existing_labels = fetch_all_labels(session, base_url, token, headers_tmp)
                # Persist a compact view of existing label lookup for transparency
                try:
                    write_json_artifact(
                        "reports",
                        "labels_lookup.json",
                        {k: v.get("labelId") for k, v in existing_labels.items()},
                        "Existing label name → id mapping",
                    )
                except Exception:
                    pass
                if not existing_labels:
                    try:
                        write_json_artifact(
                            "raw",
                            "labels_lookup.debug.json",
                            {
                                "enterpriseId": enterpriseId,
                                "tenantId": tenantId,
                                "endpoints_tried": [
                                    f"{base_url}/content/label/all?pageNumber=1&pageSize=100",
                                    f"{base_url}/content/label/all",
                                    f"{base_url}/content/labels/all",
                                ],
                            },
                            "Label lookup fallback endpoints attempted",
                        )
                    except Exception:
                        pass
                labels_artifact = []
                for it in labels_payload:
                    nm = (it.get("name") or "").strip()
                    lid = None
                    if nm:
                        ex = existing_labels.get(nm.lower())
                        if ex:
                            try:
                                lid = int(ex.get("labelId"))
                            except Exception:
                                lid = ex.get("labelId")
                    rec = {"name": nm}
                    if lid:
                        rec["labelId"] = lid
                    labels_artifact.append(rec)
            except Exception:
                labels_artifact = labels_payload
            write_json_artifact("raw", "artists.json", artists_payload, "Compiled artist payloads for dry-run review")
            write_json_artifact("raw", "labels.json", labels_artifact, "Label payloads enriched with existing IDs")
            write_json_artifact("raw", "publishers.json", publishers_payload, "Publisher payloads prepared for ingest")
            write_json_artifact("raw", "composers.json", composers_payload, "Composer payloads prepared for ingest")
            write_json_artifact("raw", "releases.json", releases_payload, "Release payloads generated from workbook")
            write_json_artifact(
                "raw",
                "tracks.json",
                [{"upc": u, **t} for u, t in tracks_payload],
                "Track payloads generated from workbook",
            )
            # Aid troubleshooting: dump resolved audio URLs per ISRC
            try:
                if 'audio_url_map' in globals() or 'audio_url_map' in locals():
                    write_json_artifact("raw", "audio_urls.json", audio_url_map, "Resolved audio URLs by ISRC")
            except Exception:
                pass
            # Composition share diagnostics
            try:
                if 'comp_share_diag' in locals():
                    write_json_artifact("reports", "composition_share_analysis.json", comp_share_diag, "Composer share diagnostics")
            except Exception:
                pass
            # Track properties diagnostics
            try:
                if 'props_diag' in locals():
                    write_json_artifact("reports", "track_properties_analysis.json", props_diag, "Track property diagnostics")
            except Exception:
                pass
            try:
                write_json_artifact("raw", "debug_field_trace.json", debug_trace, "Expanded debug trace for parsing stage")
            except Exception:
                pass
            try:
                payload_doc = build_dry_run_payload_doc(
                    base_url,
                    enterpriseId,
                    tenantId,
                    artist_image_tasks=artist_image_tasks,
                    artists_payload=artists_payload,
                    labels_payload=labels_payload,
                    publishers_payload=publishers_payload,
                    composers_payload=composers_payload,
                    releases_payload=releases_payload,
                    tracks_payload=tracks_payload,
                    audio_url_map=audio_url_map if 'audio_url_map' in locals() else None,
                )
                write_text_artifact("reports", "dry_run_payloads.md", payload_doc, "Dry-run HTTP walkthrough")
            except Exception as exc:
                print(f"[WARN] Failed to write payload simulation doc: {exc}")
            try:
                lookup_count = len(existing_labels) if 'existing_labels' in locals() else None
            except Exception:
                lookup_count = None
            s.info(artists=len(artists_payload), labels=len(labels_artifact) if 'labels_artifact' in locals() else len(labels_payload), labels_lookup=lookup_count, publishers=len(publishers_payload), composers=len(composers_payload), releases=len(releases_payload), tracks=len(tracks_payload))
            print(f"[OK] Dry-run artifacts written under {ARTIFACTS.resolve()}")

        if not args.live:
            progress.write_log()
            print("Dry-run complete. Re-run with --live to execute API calls.")
            return

        # ===== Live execution (upserts + creation)
        headers_common = {
            "X-EnterpriseId": str(enterpriseId),
            "X-TenantId": str(tenantId),
        }

        http_errors: List[Dict[str, Any]] = []
        image_upload_logs: List[Dict[str, Any]] = []
        image_file_records: List[Dict[str, Any]] = []
        # Artist profile artwork must exist before attempting POST /artists
        with progress.step("Download and upload artist images") as s:
            total = len(artist_image_tasks)
            if not artist_image_tasks:
                s.info(total=0, downloaded=0, uploaded=0, skipped=0)
            else:
                downloaded = 0; uploaded = 0; skipped = 0
                failed_names: List[str] = []
                for idx, task in enumerate(artist_image_tasks):
                    name = task.get("name")
                    url = task.get("url")
                    payload = task.get("payload")
                    if not url or not payload:
                        skipped += 1
                        failed_names.append(name or f"artist_{idx+1}")
                        image_upload_logs.append({
                            "when": "artist_image_missing_payload",
                            "entity": "artist",
                            "artistName": name,
                            "sourceUrl": url,
                            "error": "missing url or payload"
                        })
                        continue
                    safe_name = _filename_from_url(url) or f"artist_image_{idx+1}.jpg"
                    tmp_path, tmp_name, err = download_file(session, url)
                    if err or not tmp_path:
                        failed_names.append(name or safe_name)
                        image_upload_logs.append({
                            "when": "artist_image_download_failed",
                            "entity": "artist",
                            "artistName": name,
                            "sourceUrl": url,
                            "error": err or "download failed"
                        })
                        continue
                    downloaded += 1
                    filename_on_upload = tmp_name or safe_name
                    fid = None
                    up_log: Dict[str, Any] = {}
                    try:
                        fid, up_log = upload_image_file(
                            session,
                            base_url,
                            token,
                            headers_common,
                            tmp_path,
                            filename_on_upload,
                            cover=False,
                        )
                    finally:
                        try:
                            os.remove(tmp_path)
                        except Exception:
                            pass
                    if up_log:
                        up_log.update({
                            "entity": "artist",
                            "artistName": name,
                            "sourceUrl": url
                        })
                        image_upload_logs.append(up_log)
                    if fid:
                        uploaded += 1
                        payload["image"] = {"filename": filename_on_upload, "fileId": fid}
                        payload.pop("imageSourceUrl", None)
                        # Also update artist_name_to_obj so contributors can reference the complete artist object
                        if name:
                            name_key = name.strip().lower()
                            if name_key in artist_name_to_obj:
                                artist_name_to_obj[name_key]["image"] = {"filename": filename_on_upload, "fileId": fid}
                        image_file_records.append(
                            {
                                "entity": "artist",
                                "name": name,
                                "fileId": str(fid),
                                "fileName": filename_on_upload,
                                "sourceUrl": url,
                            }
                        )
                    else:
                        failed_names.append(name or filename_on_upload)
                        http_errors.append({
                            "when": "artist_image_upload",
                            "endpoint": f"{base_url}/media/image/upload",
                            "status": up_log.get("status") if isinstance(up_log, dict) else None,
                            "request": {"filename": filename_on_upload, "artist": name},
                            "response": up_log.get("responseText")[:1000] if isinstance(up_log, dict) else None
                        })
                s.info(total=total, downloaded=downloaded, uploaded=uploaded, skipped=skipped, failed=len(failed_names))
                if failed_names:
                    names = sorted({fn for fn in failed_names if fn}) or ["unknown"]
                    raise SystemExit("Failed to upload artist images for: " + ", ".join(names))

        # Update contributors in releases_payload and tracks_payload with enriched image fileIds
        # Now that artist images have been uploaded and artist_name_to_obj has the fileIds
        with progress.step("Enrich contributor image references") as s:
            updated_contributors = 0
            
            def update_contributor_images(contributors: List[Dict[str, Any]]) -> int:
                """Update image.fileId in contributor artist objects from artist_name_to_obj."""
                count = 0
                for contrib in contributors:
                    if not isinstance(contrib, dict):
                        continue
                    artist_ref = contrib.get("artist")
                    if not isinstance(artist_ref, dict):
                        continue
                    # If contributor has full artist object with name, sync image from artist_name_to_obj
                    artist_name = artist_ref.get("name")
                    if artist_name:
                        name_key = artist_name.strip().lower()
                        full_artist = artist_name_to_obj.get(name_key)
                        if full_artist and "image" in full_artist:
                            # Update the image reference with the uploaded fileId
                            artist_ref["image"] = copy.deepcopy(full_artist["image"])
                            count += 1
                return count
            
            # Update release contributors
            if releases_payload:
                for release in releases_payload:
                    if not isinstance(release, dict):
                        continue
                    contribs = release.get("contributors")
                    if isinstance(contribs, list):
                        updated_contributors += update_contributor_images(contribs)
                    # Also update track contributors within releases
                    tracks = release.get("tracks")
                    if isinstance(tracks, list):
                        for track in tracks:
                            if isinstance(track, dict):
                                track_contribs = track.get("contributors")
                                if isinstance(track_contribs, list):
                                    updated_contributors += update_contributor_images(track_contribs)
            
            # Update standalone track contributors
            if tracks_payload:
                for _, track in tracks_payload:
                    if not isinstance(track, dict):
                        continue
                    contribs = track.get("contributors")
                    if isinstance(contribs, list):
                        updated_contributors += update_contributor_images(contribs)
            
            s.info(contributors_updated=updated_contributors)

        # Upsert masters and create others
        # Collect created entity responses in `created_entities` for the final report
        created_entities: Dict[str, List[Dict[str, Any]]] = {}
        with progress.step("Upsert masters (artists/labels) and create publishers/composers") as s:
            label_map, l_created, l_reused, l_failed = create_or_reuse_labels(session, base_url, token, headers_common, labels_payload, http_errors)
            artist_map, artist_external_by_name, a_created, a_reused, a_failed, artist_failures = create_or_reuse_artists(session, base_url, token, headers_common, enterpriseId, artists_payload, http_errors)
            for name_key, ext_list in (artist_external_by_name or {}).items():
                art_obj = artist_name_to_obj.get(name_key)
                if art_obj is not None:
                    if ext_list:
                        art_obj["artistExternalIds"] = copy.deepcopy(ext_list)
                    else:
                        art_obj.pop("artistExternalIds", None)
            publisher_map, p_created, p_reused, p_failed, p_objs = create_or_reuse_publishers(session, base_url, token, headers_common, publishers_payload, http_errors, existing=publisher_lookup)
            composer_map, c_created, c_reused, c_failed, c_objs = create_or_reuse_composers(session, base_url, token, headers_common, composers_payload, http_errors, existing=composer_lookup)
            s.info(labels_created=l_created, labels_reused=l_reused, labels_failed=l_failed,
                   artists_created=a_created, artists_reused=a_reused, artists_failed=a_failed,
                   publishers_created=p_created, publishers_reused=p_reused, publishers_failed=p_failed,
                   composers_created=c_created, composers_reused=c_reused, composers_failed=c_failed)
            if p_objs:
                created_entities["publishers"] = p_objs
            if c_objs:
                created_entities["composers"] = c_objs

        if a_failed:
            detail_str = ", ".join(
                f"{entry.get('name', 'unknown')} (status {entry.get('status')})" for entry in (artist_failures or [])
            ) or "unknown artists"
            raise SystemExit(
                "Aborting before release creation. Artist upsert did not return 200 OK for: "
                + detail_str
                + ". Check artifacts/http_errors.json for full responses."
            )

        # Add artistId to contributors now that artists have been created
        # This ensures contributors have both full artist object AND artistId
        with progress.step("Add artistId to contributors") as s:
            artistid_added = 0
            
            def add_artist_id_to_contributors(contributors: List[Dict[str, Any]]) -> int:
                """Add artistId to contributor artist objects using artist_map."""
                count = 0
                for contrib in contributors:
                    if not isinstance(contrib, dict):
                        continue
                    artist_ref = contrib.get("artist")
                    if not isinstance(artist_ref, dict):
                        continue
                    # If contributor has artist name, lookup artistId from artist_map
                    artist_name = artist_ref.get("name")
                    if artist_name and "artistId" not in artist_ref:
                        name_key = artist_name.strip().lower()
                        artist_id = artist_map.get(name_key)
                        if artist_id:
                            artist_ref["artistId"] = artist_id
                            count += 1
                return count
            
            # Update release contributors
            if releases_payload:
                for release in releases_payload:
                    if not isinstance(release, dict):
                        continue
                    contribs = release.get("contributors")
                    if isinstance(contribs, list):
                        artistid_added += add_artist_id_to_contributors(contribs)
                    # Also update track contributors within releases
                    tracks = release.get("tracks")
                    if isinstance(tracks, list):
                        for track in tracks:
                            if isinstance(track, dict):
                                track_contribs = track.get("contributors")
                                if isinstance(track_contribs, list):
                                    artistid_added += add_artist_id_to_contributors(track_contribs)
            
            # Update standalone track contributors
            if tracks_payload:
                for _, track in tracks_payload:
                    if not isinstance(track, dict):
                        continue
                    contribs = track.get("contributors")
                    if isinstance(contribs, list):
                        artistid_added += add_artist_id_to_contributors(contribs)
            
            s.info(artistIds_added=artistid_added)

        # Enrich contributors with full artist objects now that images are uploaded
        # This ensures release/save and track/save payloads include complete artist data
        def enrich_contributors_with_full_artists(contributors: List[Dict[str, Any]]) -> None:
            """Replace artistId-only refs with full artist objects from artist_name_to_obj."""
            for contrib in contributors:
                if not isinstance(contrib, dict):
                    continue
                artist_ref = contrib.get("artist")
                if not isinstance(artist_ref, dict):
                    continue
                # If we have artistId but not full object, try to enrich from artist_name_to_obj
                if "artistId" in artist_ref and "artistExternalIds" not in artist_ref:
                    # Try to find by matching artistId in the lookup
                    for name_key, full_artist in artist_name_to_obj.items():
                        # Match by name lookup (best effort - contributor name may not be normalized)
                        # Better: iterate artist_map to find name by artistId, then lookup
                        pass
                    # Fallback: search for artist name in contributor if available
                    # For now, look through artist_map to find the name
                    target_artist_id = artist_ref.get("artistId")
                    artist_name_found = None
                    for name_key, full_obj in artist_name_to_obj.items():
                        # We need to check if this artist matches the artistId
                        # But artist_name_to_obj doesn't have artistId yet...
                        # Instead, reverse-lookup through artist_map
                        pass
                    # Actually, we need a different approach: use artist_map which has name->id
                    for art_name, art_id in (artist_map or {}).items():
                        if art_id == target_artist_id:
                            name_key = art_name.strip().lower()
                            full_artist = artist_name_to_obj.get(name_key)
                            if full_artist:
                                # Replace with full object
                                contrib["artist"] = copy.deepcopy(full_artist)
                            break

        if releases_payload:
            for release in releases_payload:
                if not isinstance(release, dict):
                    continue
                contribs = release.get("contributors")
                if isinstance(contribs, list):
                    enrich_contributors_with_full_artists(contribs)
                # Also enrich track contributors within releases
                tracks = release.get("tracks")
                if isinstance(tracks, list):
                    for track in tracks:
                        if isinstance(track, dict):
                            track_contribs = track.get("contributors")
                            if isinstance(track_contribs, list):
                                enrich_contributors_with_full_artists(track_contribs)

        if tracks_payload:
            for _, track in tracks_payload:
                if not isinstance(track, dict):
                    continue
                contribs = track.get("contributors")
                if isinstance(contribs, list):
                    enrich_contributors_with_full_artists(contribs)

        # Update track composer payloads with resolved publisher/composer IDs when available
        resolved_publishers = locals().get("publisher_map") or {}
        resolved_composers = locals().get("composer_map") or {}
        if (resolved_publishers or resolved_composers) and tracks_payload:
            for _, track in tracks_payload:
                if not isinstance(track, dict):
                    continue
                comps = track.get("composerContentsDTO") or []
                if not isinstance(comps, list):
                    continue
                for comp_entry in comps:
                    if not isinstance(comp_entry, dict):
                        continue
                    if resolved_publishers and comp_entry.get("publisherName") and not comp_entry.get("publisherId"):
                        key = comp_entry.get("publisherName", "").strip().lower()
                        pid = resolved_publishers.get(key)
                        if pid:
                            comp_entry["publisherId"] = pid
                    if resolved_composers and comp_entry.get("composerName") and not comp_entry.get("composerId"):
                        key = comp_entry.get("composerName", "").strip().lower()
                        cid = resolved_composers.get(key)
                        if cid:
                            comp_entry["composerId"] = cid

        # Prepare cover images: download to temp storage, build mapping, then upload to get fileIds
        with progress.step("Prepare and upload cover images") as s:
            downloaded = 0; uploaded = 0; skipped = 0
            for idx, rel in enumerate(releases_payload):
                src = rel.get("imageSourceUrl")
                if not src:
                    skipped += 1
                    continue
                try:
                    # Download each image to temp storage with deterministic filename
                    _, name = os.path.split(_filename_from_url(src))
                    safe_name = name or f"image_{idx}.jpg"
                    tmp_path, fn, err = download_file(session, src)
                    if err or not tmp_path:
                        image_upload_logs.append({
                            "when": "download_image",
                            "entity": "release",
                            "releaseName": rel.get("name"),
                            "sourceUrl": src,
                            "error": err or "unknown"
                        })
                        continue
                    downloaded += 1
                    # Upload
                    upload_name = fn or safe_name
                    fid = None
                    up_log: Dict[str, Any] = {}
                    try:
                        fid, up_log = upload_image_file(
                            session,
                            base_url,
                            token,
                            headers_common,
                            tmp_path,
                            upload_name,
                        )
                    finally:
                        try:
                            os.remove(tmp_path)
                        except Exception:
                            pass
                    if up_log:
                        up_log.update({
                            "entity": "release",
                            "releaseName": rel.get("name"),
                            "sourceUrl": src
                        })
                        image_upload_logs.append(up_log)
                    if fid:
                        uploaded += 1
                        image_file_records.append(
                            {
                                "entity": "release",
                                "name": rel.get("name"),
                                "upc": rel.get("upc"),
                                "fileId": str(fid),
                                "fileName": upload_name,
                                "sourceUrl": src,
                            }
                        )
                        rel["image"] = {"fileId": fid, "filename": upload_name}
                        # Once we have a fileId, drop the source URL to avoid ambiguous payloads
                        if "imageSourceUrl" in rel:
                            rel.pop("imageSourceUrl", None)
                except Exception as e:
                    image_upload_logs.append({
                        "when": "image_pipeline_exception",
                        "entity": "release",
                        "releaseName": rel.get("name"),
                        "sourceUrl": src,
                        "error": str(e)
                    })
            # Persist mapping
            try:
                write_json_artifact(
                    "reports",
                    "image_file_map.json",
                    image_file_records,
                    "Image uploads resolved to fileIds for artists and releases",
                )
            except Exception:
                pass
            s.info(downloaded=downloaded, uploaded=uploaded, skipped=skipped, mapped=len(image_file_records))

        # Inject known IDs into release/track payloads before creation
        with progress.step("Wire labelId/artistId into payloads") as s:
            # Labels on releases
            try:
                for rel in releases_payload:
                    lname = (rel.get("labelName") or "").strip()
                    if lname:
                        lid = None
                        try:
                            lid = label_map.get(lname.lower())
                        except Exception:
                            lid = None
                        if lid:
                            rel["labelId"] = int(lid)
                            rel["hasRecordLabel"] = True
                            # Keep labelName for readability; API should prefer labelId
                # Contributors on releases
                for rel in releases_payload:
                    contribs = rel.get("contributors") or []
                    for c in contribs:
                        art = c.get("artist") or {}
                        nm = (art.get("name") or "").strip()
                        if nm:
                            aid = None
                            try:
                                aid = artist_map.get(nm.lower())
                            except Exception:
                                aid = None
                            if aid:
                                art["artistId"] = int(aid)
                # Contributors on tracks
                for i, (upc, track) in enumerate(tracks_payload):
                    contribs = track.get("contributors") or []
                    for c in contribs:
                        art = c.get("artist") or {}
                        nm = (art.get("name") or "").strip()
                        if nm:
                            aid = None
                            try:
                                aid = artist_map.get(nm.lower())
                            except Exception:
                                aid = None
                            if aid:
                                art["artistId"] = int(aid)
                # Primary artists on releases
                for rel in releases_payload:
                    primary_name = (rel.get("artistName") or "").strip()
                    primary_id: Optional[Any] = None
                    if primary_name:
                        try:
                            primary_id = artist_map.get(primary_name.lower())
                        except Exception:
                            primary_id = None
                    if primary_id is not None:
                        try:
                            rel["artistId"] = int(primary_id)
                        except Exception:
                            rel["artistId"] = primary_id
                        rel.pop("artistName", None)
                    elif primary_name:
                        debug_trace["artist_binding_warnings"].append({
                            "scope": "release",
                            "upc": rel.get("upc"),
                            "artistName": primary_name,
                            "issue": "artist_not_resolved"
                        })
                    contribs = rel.get("contributors") or []
                    contribs_with_ids = sum(1 for c in contribs if isinstance(c.get("artist"), dict) and "artistId" in c.get("artist", {}))
                    add_debug_sample(debug_trace["final_releases"], {
                        "upc": rel.get("upc"),
                        "labelId": rel.get("labelId"),
                        "primaryArtistId": rel.get("artistId"),
                        "primaryArtistName": primary_name,
                        "artistExternalIds": rel.get("artistExternalIds"),
                        "contributorsCount": len(contribs),
                        "contributorsWithIds": contribs_with_ids
                    })
                # Primary artists on tracks
                for upc, track in tracks_payload:
                    primary_name = (track.get("artistName") or "").strip()
                    primary_id: Optional[Any] = None
                    if primary_name:
                        try:
                            primary_id = artist_map.get(primary_name.lower())
                        except Exception:
                            primary_id = None
                    first_isrc: Optional[str] = None
                    versions = track.get("trackRecordingVersions") or []
                    if versions:
                        try:
                            first = versions[0] if isinstance(versions[0], dict) else None
                            if first:
                                first_isrc = first.get("isrc")
                        except Exception:
                            first_isrc = None
                    if primary_id is not None:
                        try:
                            track["artistId"] = int(primary_id)
                        except Exception:
                            track["artistId"] = primary_id
                        track.pop("artistName", None)
                    elif primary_name:
                        debug_trace["artist_binding_warnings"].append({
                            "scope": "track",
                            "upc": upc,
                            "isrc": first_isrc,
                            "artistName": primary_name,
                            "issue": "artist_not_resolved"
                        })
                    contribs = track.get("contributors") or []
                    contribs_with_ids = sum(1 for c in contribs if isinstance(c.get("artist"), dict) and "artistId" in c.get("artist", {}))
                    add_debug_sample(debug_trace["final_tracks"], {
                        "upc": upc,
                        "isrc": first_isrc,
                        "previewStartSeconds": track.get("previewStartSeconds"),
                        "primaryArtistId": track.get("artistId"),
                        "primaryArtistName": primary_name,
                        "artistExternalIds": track.get("artistExternalIds"),
                        "contributorsCount": len(contribs),
                        "contributorsWithIds": contribs_with_ids
                    })
            finally:
                # Summaries
                counted_rel_label_ids = sum(1 for rel in releases_payload if rel.get("labelId"))
                counted_rel_contrib_ids = sum(1 for rel in releases_payload for c in (rel.get("contributors") or []) if isinstance(c.get("artist"), dict) and "artistId" in c.get("artist", {}))
                counted_trk_contrib_ids = sum(1 for _, track in tracks_payload for c in (track.get("contributors") or []) if isinstance(c.get("artist"), dict) and "artistId" in c.get("artist", {}))
                counted_release_primary_ids = sum(1 for rel in releases_payload if rel.get("artistId"))
                counted_track_primary_ids = sum(1 for _, track in tracks_payload if track.get("artistId"))
                s.info(
                    release_label_ids=counted_rel_label_ids,
                    release_primary_ids=counted_release_primary_ids,
                    release_contrib_ids=counted_rel_contrib_ids,
                    track_primary_ids=counted_track_primary_ids,
                    track_contrib_ids=counted_trk_contrib_ids
                )

        # Block if any track lacks a valid uploaded audioId
        with progress.step("Validate required media (audio)") as s:
            missing = []
            for upc, track in tracks_payload:
                trk_recs = track.get("trackRecordingVersions") or []
                isrc = trk_recs[0].get("isrc") if trk_recs else None
                afs = (trk_recs[0].get("audioFiles") if trk_recs else None) or []
                has_audio_id = any(isinstance(af, dict) and af.get("audioId") for af in afs)
                if not has_audio_id:
                    missing.append({
                        "upc": upc,
                        "isrc": isrc,
                        "audioUrl": audio_url_map.get(isrc) if 'audio_url_map' in locals() or 'audio_url_map' in globals() else None
                    })
            s.info(missing=len(missing))
            if missing:
                out = ARTIFACTS/"missing_audio_ids.json"
                out.write_text(json.dumps(missing, indent=2, ensure_ascii=False))
                print(f"[BLOCKED] Missing audioId for {len(missing)} track(s). Upload audio to storage and retry. See {out.resolve()}")
                progress.write_log()
                sys.exit(1)

    # Releases (with UPC duplicate handling)
        with progress.step("Create releases") as s:
            # Build tracks grouped by release UPC
            tracks_by_upc: Dict[str, List[Dict[str, Any]]] = {}
            try:
                for upc_key, trk in tracks_payload:
                    if upc_key:
                        tracks_by_upc.setdefault(upc_key, []).append(trk)
            except Exception:
                tracks_by_upc = {}
            upc_to_release_id: Dict[str,str] = {}
            rel_created = 0; rel_failed = 0
            release_search_cache: Dict[str, List[Dict[str, Any]]] = {}
            release_override_answers: Dict[str, bool] = {}
            for rel in releases_payload:
                body = dict(rel)  # copy
                # ensure UPC is serialized as string per new API contract
                if body.get("upc") is not None:
                    try:
                        body["upc"] = str(body.get("upc"))
                    except Exception:
                        body["upc"] = body.get("upc")
                # Attach tracks for this release if available
                upc_val = body.get("upc")
                release_name = body.get("name")
                if upc_val and upc_val in tracks_by_upc:
                    body["tracks"] = tracks_by_upc[upc_val]
                if args.live:
                    duplicate = find_release_duplicate(session, base_url, token, headers_common, release_name, upc_val, release_search_cache)
                    if duplicate is not None:
                        upc_key = _normalize_code(upc_val) or upc_val
                        decision = release_override_answers.get(upc_key)
                        if decision is None:
                            dup_id = duplicate.get("releaseId") or duplicate.get("id")
                            dup_name = duplicate.get("name") or release_name
                            dup_upc = duplicate.get("upc") or duplicate.get("UPC") or upc_val
                            print(f"[ALERT] Release '{dup_name}' with UPC '{dup_upc}' already exists (releaseId {dup_id}).")
                            decision = yes_no("Proceed to override this release?")
                            release_override_answers[upc_key] = decision
                        if not decision:
                            print("[ABORT] Execution stopped to avoid overriding existing release as requested.")
                            progress.write_log()
                            sys.exit(1)
                url = f"{base_url}/content/release/save"
                resp = http(session, "POST", url, token, json_body=body, headers=headers_common)
                if not resp.ok:
                    txt = (resp.text or "").lower()
                    # If duplicate UPC error → retry without upc and log
                    if "upc" in txt and ("exist" in txt or "duplicate" in txt or resp.status_code in (400,409)):
                        upc_removed = body.pop("upc", None)
                        upc_dupes_logged.append(upc_removed)
                        print(f"[INFO] UPC '{upc_val}' appears to exist; retrying without UPC as requested.")
                        resp = http(session, "POST", url, token, json_body=body, headers=headers_common)
                ensure_http_ok(resp, context=f"create_release[{upc_val or rel.get('name') or 'unknown'}]", endpoint=url, request=body, http_errors=http_errors)
                rel_created += 1
                try:
                    data = resp.json() or {}
                except Exception:
                    data = {}
                expected_release_fields = extract_expected_fields("release", body)
                record_entity_attempt(
                    kind="release",
                    endpoint=url,
                    request_payload=copy.deepcopy(body),
                    response_payload=data if isinstance(data, dict) else {},
                    status_code=resp.status_code,
                    success=bool(resp.ok),
                    identifiers={"releaseId": data.get("releaseId"), "upc": upc_val, "name": release_name},
                    expected=expected_release_fields,
                    request_id=getattr(resp, "_catalog_request_id", None),
                )
                rid = data.get("releaseId")
                # store creation info for report
                created_entities.setdefault("releases", []).append({"request": body, "response": data})
                if rid and rel.get("upc"):
                    upc_to_release_id[rel["upc"]] = rid
            s.info(created=rel_created, failed=rel_failed)

        # Tracks per release
        with progress.step("Create tracks") as s:
            created = 0; failed = 0
            track_search_cache: Dict[str, List[Dict[str, Any]]] = {}
            track_override_answers: Dict[str, bool] = {}
            for upc, track in tracks_payload:
                track_name = track.get("name") if isinstance(track, dict) else None
                try:
                    isrc_ref = (track.get("trackRecordingVersions") or [{}])[0].get("isrc") if isinstance(track, dict) else None
                except Exception:
                    isrc_ref = None
                if args.live:
                    duplicate_track = find_track_duplicate(session, base_url, token, headers_common, track_name, isrc_ref, track_search_cache)
                    if duplicate_track is not None:
                        isrc_key = _normalize_code(isrc_ref) or isrc_ref
                        decision = track_override_answers.get(isrc_key)
                        if decision is None:
                            dup_id = duplicate_track.get("trackId") or duplicate_track.get("id")
                            dup_name = duplicate_track.get("name") or track_name
                            dup_isrc = duplicate_track.get("isrc") or duplicate_track.get("ISRC") or isrc_ref
                            print(f"[ALERT] Track '{dup_name}' with ISRC '{dup_isrc}' already exists (trackId {dup_id}).")
                            decision = yes_no("Proceed to make a duplicate of this track?")
                            track_override_answers[isrc_key] = decision
                        if not decision:
                            print("[ABORT] Execution stopped to avoid overriding existing track as requested.")
                            progress.write_log()
                            sys.exit(1)
                # If we know releaseId, include association if API needs it; otherwise the endpoint may infer.
                t_url = f"{base_url}/content/track/save"
                t_resp = http(session, "POST", t_url, token, json_body=track, headers=headers_common)
                ensure_http_ok(t_resp, context=f"create_track[{isrc_ref or track.get('name') or 'unknown'}]", endpoint=t_url, request=track, http_errors=http_errors)
                created += 1
                try:
                    tdata = t_resp.json() or {}
                except Exception:
                    tdata = {}
                expected_track_fields = extract_expected_fields("track", track if isinstance(track, dict) else {})
                record_entity_attempt(
                    kind="track",
                    endpoint=t_url,
                    request_payload=copy.deepcopy(track) if isinstance(track, dict) else {},
                    response_payload=tdata if isinstance(tdata, dict) else {},
                    status_code=t_resp.status_code,
                    success=bool(t_resp.ok),
                    identifiers={"trackId": tdata.get("trackId"), "name": track_name, "isrc": isrc_ref, "upc": upc},
                    expected=expected_track_fields,
                    request_id=getattr(t_resp, "_catalog_request_id", None),
                )
                created_entities.setdefault("tracks", []).append({"request": track, "response": tdata})
            s.info(created=created, failed=failed)

        if upc_dupes_logged:
            (ARTIFACTS/"upc_skipped_for_duplicates.json").write_text(json.dumps(upc_dupes_logged, indent=2))
            print(f"[INFO] UPCs skipped (already existed): {len(upc_dupes_logged)} → logged to upc_skipped_for_duplicates.json")

        if http_errors:
            (ARTIFACTS/"http_errors.json").write_text(json.dumps(http_errors, indent=2, ensure_ascii=False))
            print(f"[LOG] Wrote HTTP error details to {(ARTIFACTS/ 'http_errors.json').resolve()}")

        if args.live and HTTP_POST_PAYLOADS:
            try:
                path = ARTIFACTS/"http_post_payloads.json"
                path.write_text(json.dumps(HTTP_POST_PAYLOADS, indent=2, ensure_ascii=False))
                print(f"[LOG] Wrote POST payload audit to {path.resolve()}")
            except Exception as e:
                print(f"[WARN] Failed to write POST payload artifact: {e}")

        if args.live and HTTP_GET_PAYLOADS:
            try:
                path = ARTIFACTS/"http_get_payloads.json"
                path.write_text(json.dumps(HTTP_GET_PAYLOADS, indent=2, ensure_ascii=False))
                print(f"[LOG] Wrote GET payload audit to {path.resolve()}")
            except Exception as e:
                print(f"[WARN] Failed to write GET payload artifact: {e}")

        if args.live and HTTP_RESPONSES:
            try:
                path = ARTIFACTS/"http_responses.json"
                path.write_text(json.dumps(HTTP_RESPONSES, indent=2, ensure_ascii=False))
                print(f"[LOG] Wrote HTTP responses to {path.resolve()}")
            except Exception as e:
                print(f"[WARN] Failed to write HTTP responses artifact: {e}")

        if args.live and HTTP_GET_RESPONSES:
            try:
                path = ARTIFACTS/"http_get_responses.json"
                path.write_text(json.dumps(HTTP_GET_RESPONSES, indent=2, ensure_ascii=False))
                print(f"[LOG] Wrote GET responses to {path.resolve()}")
            except Exception as e:
                print(f"[WARN] Failed to write GET responses artifact: {e}")

        # Upload logs for audio
        try:
            if audio_upload_logs:
                (ARTIFACTS/"audio_uploads.json").write_text(json.dumps(audio_upload_logs, indent=2, ensure_ascii=False))
                print(f"[LOG] Wrote audio upload logs to {(ARTIFACTS/ 'audio_uploads.json').resolve()}")
        except Exception:
            pass

        # Upload logs for images
        try:
            if image_upload_logs:
                (ARTIFACTS/"image_uploads.json").write_text(json.dumps(image_upload_logs, indent=2, ensure_ascii=False))
                print(f"[LOG] Wrote image upload logs to {(ARTIFACTS/ 'image_uploads.json').resolve()}")
        except Exception:
            pass

        tracker_obj = get_current_tracker()
        attempts_iterable = tracker_obj.entity_attempts if tracker_obj else []
        try:
            verification_summary = run_verification(
                session=session,
                base_url=base_url,
                token=token,
                headers=headers_common,
                enterprise_id=enterpriseId,
                http_call=http,
                attempts=attempts_iterable,
            )
        except Exception as exc:
            verification_summary = {"status": "failed", "error": str(exc)}
            set_verification_summary(verification_summary)

        # Write a creation report summarizing created entities and mappings
        try:
            report = {
                "generated": round(time.time(), 3),
                "artists_by_name": (locals().get("artist_map") or {}),
                "labels_by_name": (locals().get("label_map") or {}),
                "publishers_by_name": (locals().get("publisher_map") or {}),
                "composers_by_name": (locals().get("composer_map") or {}),
                "upc_skipped": (locals().get("upc_dupes_logged") or []),
                "created_entities": (locals().get("created_entities") or {}),
                "verification_summary": copy.deepcopy(verification_summary),
            }
            if tracker_obj:
                report["exercise"] = copy.deepcopy(tracker_obj.exercise)
                report["exercise_metadata"] = copy.deepcopy(tracker_obj.metadata)
            out_path = ARTIFACTS/"creation_report.json"
            out_path.write_text(json.dumps(report, indent=2, ensure_ascii=False))
            print(f"[LOG] Wrote creation report to {out_path.resolve()}")
        except Exception as e:
            print(f"[WARN] Failed to write creation report: {e}")

        progress.write_log()
        print("[DONE] Live execution finished.")


if __name__ == "__main__":
    main()
